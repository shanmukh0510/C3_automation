import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
print(Element)
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_Help_page_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def click_element_with_wait(driver_setup, by, xpath):
    element = WebDriverWait(driver_setup, 10).until(
        EC.element_to_be_clickable((by, xpath))
    )
    element.click()
    time.sleep(3)

def verify_help_page(driver_setup):
    try:
        click_element_with_wait(driver_setup, By.XPATH, ElementXpath['Help'])

        grl_support = WebDriverWait(driver_setup, 10).until(
            EC.visibility_of_element_located((By.XPATH, ElementXpath['support']))
        )
        support = yaml_msg("support")

        if grl_support.text == support:
            logging.info("Help page is loaded")
        else:
            raise Exception("Help Page is not loaded")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return "Fail", str(e)  # Include the error message as the remark in case of failure

    return "Pass", " "

def tc2_verify_help_page(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("Verify that after clicking the help page, it should load and land on the Help Page.")
    pass_or_fail, remark = verify_help_page(driver_setup)

    test_description = "Verify that after clicking the help page, it should load and land on the Help Page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name,test_description ,pass_or_fail, remark)
    
def tc3_grl_supportDesk(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("Please confirm the presence of the support desk and ensure it should load the ticket page.")
    pass_or_fail = "Pass"
    remark = ""

    try:
        grl_support = driver_setup.find_element(By.XPATH, ElementXpath['support'])
        support = yaml_msg("support")

        if grl_support.text == support:
            logging.info("GRL Support Desk is present")
        else:
            logging.error("GRL Support Desk is not present")
            pass_or_fail = "Fail"
            remark = "GRL Support Desk is not present"
    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "Error finding GRL Support Desk element"

    try:
        grl_support_url = driver_setup.find_element(By.XPATH, ElementXpath.get('url'))
        logging.info('The website is accessible.')
    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "The support URL element was not found."

    test_description = "Please confirm the presence of the support desk and ensure it should load the ticket page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_email_support(driver_setup):
    logging.info("*************** TC-4 ***************")
    logging.info("Please confirm the presence of the email_support and ensure it should load the email page.")
    pass_or_fail = "Pass"
    remark = ""

    try:
        email = driver_setup.find_element(By.XPATH, ElementXpath['email'])
        emailsupport = yaml_msg("email")

        if email.text == emailsupport:
            logging.info("Email Customer Support is present")
        else:
            logging.error('Email Customer support is not present')
            pass_or_fail = "Fail"
            remark = "Email Customer support is not present"
    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "Error finding Email Customer Support element"

    test_description = "Please confirm the presence of the email_support and ensure it should load the email page"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc5_download_logs(driver_setup):
    logging.info("*************** TC-5 ***************")
    logging.info("Please confirm the presence of the download log and ensure it should download.")
    pass_or_fail = "Pass"
    remark = ""

    try:
        debuglog = driver_setup.find_element(By.XPATH, ElementXpath['debuglog'])
        time.sleep(2)
        debuglog.click()
        time.sleep(2)
        debug = yaml_msg("debug")

        if debuglog.text == debug:
            logging.info("Download debug log is present and it's downloaded")
        else:
            logging.error('Download debug log is not present')
            pass_or_fail = "Fail"
            remark = "Download debug log is not present"
    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "Error finding Download debug log element"

    test_description = "Please confirm the presence of the download log and ensure it should download"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc6_info(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Please confirm the info@graniteriverlabs.com should be present.")
    pass_or_fail = "Pass"
    remark = ""

    try:
        infoemail = driver_setup.find_element(By.XPATH, ElementXpath['info'])
        info = yaml_msg("info")

        if infoemail.text == info:
            logging.info("Info emailId is present")
        else:
            logging.error('Info emailID is not present')
            pass_or_fail = "Fail"
            remark = "Info emailID is not present"
    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "Error finding Info emailID element"

    test_description = "Please confirm the info@graniteriverlabs.com should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
