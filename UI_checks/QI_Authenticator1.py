import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from pynput.keyboard import Key, Controller
import random
import string
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
print(Element)
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_QI_Auth_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_check_QI_Authenticator(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("QI_Authenticator")
    pass_or_fail = "Pass"
    remark = ""
    try:
        qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        time.sleep(8)
        qi.click()
        Qi1 = driver_setup.find_element(By.XPATH, ElementXpath['Certificate'])
        # Verify if the text is present in the element
        if Qi1.text ==  "Certificate Validation":
            logging.info("QI-Authenticator page is loaded successfully")
        else:
            logging.error("Can't able to load the QI-Authenticator page")
            pass_or_fail = "Fail"
            remark = "Can't able to load the QI-Authenticator page"
    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        logging.error(f"The attempt failed to load the QiAuthenticator page because button is not clickable at the moment. {str(e)}")
        remark = "The attempt failed to load the QiAuthenticator page because button is not clickable at the moment/unable to load the page."

    testdescription = "Check that QI-Authenticator page should is loaded successfully "
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc3_upload_and_validate_Authfile(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("Validate the trace")
    pass_or_fail = "Pass"
    remark = ""

    try:
        loadtrace(driver_setup)
        time.sleep(10)
        filter_packet_image = driver_setup.find_element(By.XPATH, ElementXpath['filter'])
        # Click on the filter packet image
        filter_packet_image.click()
        time.sleep(2)
        # Wait for the checkbox to be clickable
        checkbox = driver_setup.find_element(By.ID, "authentication_message")
        # Click on the checkbox
        checkbox.click()
        # Click on the filter packet image to close
        filter_packet_image.click()
        time.sleep(5)
        align = driver_setup.find_element(By.XPATH, ElementXpath['align'])
        # Click on the filter packet image
        align.click()

        # First Occurrence
        try:
            pass_or_fail = "Pass"
            second_occurrence_xpath = '(//div[@class="ReactVirtualized__Grid ReactVirtualized__List C3_packet_list"])[2]'

            # Locate all elements with the text "0: CHALLENGE_AUTH" within the second occurrence
            challenge_auth_xpath = f'{second_occurrence_xpath}//div[@class=" packetString truncate_text textLeft ccpacket-font-style" and contains(text(), "0: CHALLENGE_AUTH")]'
            challenge_auth_elements = driver_setup.find_elements(By.XPATH, challenge_auth_xpath)

            # Now you can print the number of occurrences
            num_occurrences = len(challenge_auth_elements)
            time.sleep(2)
            # Locate the "Authenticate" element that follows the "0: CHALLENGE_AUTH" element
            authenticate_xpath = f'{challenge_auth_xpath}/following::span[@class="auth-custom-tag custom-tag"][1]'
            authenticate_element = driver_setup.find_element(By.XPATH, authenticate_xpath)

            try:
                authenticate_element.click()
            except Exception as e:
                pass_or_fail = "Fail"
                remark = "Check the XPATH can't able to interact with element"

            time.sleep(5)
            valid = driver_setup.find_element(By.XPATH, ElementXpath['Challenge_Signature_Valid'])
            if valid.text == 'Challenge Signature Valid':
                logging.info("Challenge Signature Valid - Challenge Auth Passed")
            else:
                logging.info("Challenge Signature is not Valid - Challenge Auth Failed")
                pass_or_fail = "Fail"
                remark = "Challenge Signature is not Valid - Challenge Auth Failed"
            time.sleep(2)
        except Exception as e:
            pass

        # Switch to Result tab
        try:
            element = WebDriverWait(driver_setup, 10).until(
                EC.element_to_be_clickable((By.XPATH, ElementXpath['Result_tab'])))
            element.click()
        except TimeoutException as e:
            logging.error("Timeout waiting for Result tab")
            pass_or_fail = "Fail"
            remark = "Timeout waiting for Result tab"

        # Second Occurrence
        try:
            time.sleep(2)
            second_occurrence_xpath = '(//div[@class="ReactVirtualized__Grid ReactVirtualized__List C3_packet_list"])[2]'

            # Locate all elements with the text "0: CHALLENGE_AUTH" within the second occurrence
            challenge_auth_xpath = f'{second_occurrence_xpath}//div[@class=" packetString truncate_text textLeft ccpacket-font-style" and contains(text(), "0: CHALLENGE_AUTH")]'
            challenge_auth_elements = driver_setup.find_elements(By.XPATH, challenge_auth_xpath)

            # Now you can print the number of occurrences
            num_occurrences = len(challenge_auth_elements)
            time.sleep(2)
            # Locate the "Authenticate" element that follows the "0: CHALLENGE_AUTH" element
            authenticate_xpath = f'{challenge_auth_xpath}/following::span[@class="auth-custom-tag custom-tag"][2]'
            authenticate_element = driver_setup.find_element(By.XPATH, authenticate_xpath)

            try:
                authenticate_element.click()
            except Exception as e:
                pass_or_fail = "Fail"
                remark = "Check the XPATH can't able to interact with element"

            time.sleep(5)
            valid = driver_setup.find_element(By.XPATH, ElementXpath['Challenge_Signature_Valid'])
            if valid.text == 'Challenge Signature Valid':
                logging.info("Challenge Signature Valid - Challenge Auth Passed")
            else:
                logging.info("Challenge Signature is not Valid - Challenge Auth Failed")
                pass_or_fail = "Fail"
                remark = "Challenge Signature is not Valid - Challenge Auth Failed"
        except Exception as e:
            pass

    except Exception as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that QI Challenge Authenticator should have Valid Signature"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)


def tc4_verify_signature_notValid(driver_setup):
    logging.info("*************** TC-4 ***************")
    logging.info("Verify that the signature is not valid")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        # Find the textarea element using id
        textarea_element = driver_setup.find_element(By.ID, 'comment')

        # Capture the text from the textarea element
        captured_text = textarea_element.get_attribute('value')

        # Modify the text by adding a random number
        modified_text = captured_text + generate_random_number()

        # Pass the modified text back to the textarea element
        textarea_element.clear()
        textarea_element.send_keys(modified_text)

        time.sleep(2)
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['validatebutton'])
        label_element.click()
        time.sleep(2)
        valid = driver_setup.find_element(By.XPATH, ElementXpath['Challenge_Signature_Not_Valid'])
        if valid.text == 'Challenge Signature Not-Valid':
            logging.info("Challenge Signature Not-Valid - Challenge Auth Passed")
        else:
            raise NoSuchElementException("Challenge Signature is Valid - Challenge Auth Failed")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Edited QI raw certificate, should verify the challenge Auth and confirm that it does not have a valid signature"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

    
def generate_random_number():
    return ''.join(random.choices(string.digits, k=4))

def tc5_upload_PTX_Auth(driver_setup):
    logging.info("*************** TC-5 ***************")
    logging.info('Verify Upload Ptx Auth')
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['QI1'])
        if label_element:
            logging.info("Sequence 1 is present: Upload PTX's Auth Certificate Chain")
        else:
            raise NoSuchElementException("Sequence 1 is not present: Upload PTX's Auth Certificate Chain")

        label_element = driver_setup.find_element(By.XPATH, ElementXpath['qi_Download'])
        if label_element.text == 'Download':
            logging.info("Download button is present")
        else:
            raise NoSuchElementException("Download button is not present")

        label_element = driver_setup.find_element(By.XPATH, ElementXpath['qi_Update'])
        if label_element.text == 'Update':
            logging.info("Update button is present")
        else:
            raise NoSuchElementException("Update button is not present")

        label_element = driver_setup.find_element(By.XPATH, ElementXpath['qi_clear'])
        if label_element.text == 'Clear':
            logging.info("Clear button is present")
        else:
            raise NoSuchElementException("Clear button is not present")

        label_element = driver_setup.find_element(By.XPATH, ElementXpath['choose_file'])
        if label_element.text == 'Choose File':
            logging.info("Choose File button is present")
        else:
            raise NoSuchElementException("Choose File button is not present")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that Upload PTX's Auth Certificate Chain and the Element should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc6_nonce(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Verify the Nonce is present")
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['QI3'])
        if label_element:
            logging.info("Sequence 3 is present: Nonce")
        else:
            raise NoSuchElementException("Sequence 3 is not present: Nonce")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that Nonce and the input box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc7_challenge_Auth(driver_setup):
    logging.info("*************** TC-7 ***************")
    logging.info("Verify Challenge Auth is present")
    pass_or_fail = "Pass"
    remark = ""
    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['QI4'])
        if label_element:
            logging.info("Sequence 5 is present: Challenge Auth")
        else:
            raise NoSuchElementException("Sequence 5 is not present: Challenge Auth")
        
    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that Challenge Auth and the input box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc8_decode_certificate(driver_setup):
    logging.info("*************** TC-8 ***************")
    logging.info("Verify the Decode Certificate")
    pass_or_fail = "Pass"
    remark = ""

    def check_element(xpath, expected_text):
        label_element = driver_setup.find_element(By.XPATH, xpath)
        if label_element.text == expected_text:
            logging.info(f"{expected_text} is present")
        else:
            logging.error(f"{expected_text} is not present")
            nonlocal pass_or_fail, remark
            pass_or_fail = "Fail"
            remark = f"{expected_text} is not present"

    Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
    Qi.click()

    check_element(ElementXpath['decode'], 'Decoded Certificate')
    check_element(ElementXpath['qi_save'], 'Save Changes')
    check_element(ElementXpath['qi_Download_cert'], 'Download Certificate')

    # Check the Root Certificate hash
    logging.info("Sub: Root Certificate hash")
    check_element(ElementXpath['roothash'], 'Root Certificate Hash')

    dd = driver_setup.find_element(By.XPATH, ElementXpath['roothash_DD'])
    dd.click()
    time.sleep(2)
    check_element(ElementXpath['choose_file2'], 'Choose File')
    check_element(ElementXpath['qi_clear2'], 'Clear')

    du = driver_setup.find_element(By.XPATH, ElementXpath['DU'])
    du.click()
    time.sleep(2)

    # Manufacturer CA Certificate
    logging.info("Sub: Manufacturer CA Certificate")
    check_element(ElementXpath['manufacturerCA'], 'Manufacturer CA Certificate')

    # Product Unit Certificate
    logging.info("Sub: Product Unit Certificate")
    check_element(ElementXpath['productCA'], 'Product Unit Certificate')

    dd = driver_setup.find_element(By.XPATH, ElementXpath['manufacturerCA_DD'])
    dd.click()
    logging.info("----------------Value to be under the Manufacturer CA Certificate----------------")
    pass_or_fail,remark = common(driver_setup, 'MCAC')
    du = driver_setup.find_element(By.XPATH, ElementXpath['DU'])
    du.click()
    time.sleep(2)
    
    dd = driver_setup.find_element(By.XPATH, ElementXpath['productCA_DD'])
    dd.click()
    logging.info("----------------Value to be under the Product Unit Certificate----------------")
    pass_or_fail,remark = common(driver_setup, 'PUC')
    du = driver_setup.find_element(By.XPATH, ElementXpath['DU'])
    du.click()
    time.sleep(2)

    testdescription = "Check that Decoded Certificate and their functionalities should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def common(driver_setup, value):
    pass_or_fail = "Pass"
    remark = ""

    def check_element(xpath, expected_text):
        nonlocal pass_or_fail, remark
        try:
            time.sleep(1)
            label_element = driver_setup.find_element(By.XPATH, xpath)
            if label_element.text == expected_text:
                logging.info(f"{expected_text} text is present")
            else:
                raise NoSuchElementException(f"{expected_text} text is not present")
        except NoSuchElementException as e:
            if expected_text == 'ValidityNotAfter1' and "no such element" in str(e).lower():
                logging.warning(f"{expected_text} is not present, but this might be expected.")
            else:
                logging.error(str(e))
                pass_or_fail = "Fail"
                remark = f"{expected_text} is not present"

    try:
        check_element(ElementXpath['rawbytes'], 'RawBytes')
        check_element(ElementXpath['version'], 'Version')
        check_element(ElementXpath['serialnumber'], 'SerialNumber')
        check_element(ElementXpath['sign'], 'Signature')
        check_element(ElementXpath['issuer'], 'Issuer')
        check_element(ElementXpath['notafter'], 'ValidityNotAfter')
        check_element(ElementXpath['notbefore'], 'ValidityNotBefore')
        check_element(ElementXpath['pkia'], 'SubjectPublicKeyInfoAlgorithm')
        check_element(ElementXpath['pkia1'], 'SubjectPublicKeyInfoAlgorithm1')
        check_element(ElementXpath['pkisp'], 'SubjectPublicKeyInfoSubjectPublicKey')
        check_element(ElementXpath['Extensions1'], 'Extensions1')
        check_element(ElementXpath['ve1critical'], 'Extensions1Critical')
        check_element(ElementXpath['signalgorithm'], 'SignatureAlgorithm')
        check_element(ElementXpath['signvalue'], 'SignatureValue')

        if value == 'MCAC':
            check_element(ElementXpath['sub'], 'Subject')
            check_element(ElementXpath['extensions2'], 'Extensions2')
            check_element(ElementXpath['2critical'], 'Extensions2Critical')
            check_element(ElementXpath['2extnvalue'], 'Extensions2ExtnValue')
            check_element(ElementXpath['1extnvalueca'], 'Extensions1ExtnValueCA')
            check_element(ElementXpath['1extnvalue'], 'Extensions1ExtnValuePathLenConstraint')

        elif value == 'PUC':
            check_element(ElementXpath['subjectAttribute1'], 'SubjectAttribute1')
            check_element(ElementXpath['subjectAttribute2'], 'SubjectAttribute2')
            check_element(ElementXpath['subjectAttribute3'], 'SubjectAttribute3')

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        pass_or_fail = "Fail"
        remark = str(e)

    return pass_or_fail, remark

def tc9_certificate_chain(driver_setup):
    logging.info("*************** TC-9 ***************")
    logging.info("Verify the Certificate Chain")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['digest'])
        if label_element.text == 'Certificate chain Digest':
            logging.info("Certificate chain Digest is present")
        else:
            raise NoSuchElementException("Certificate chain Digest is not present")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that Certificate chain Digest and their functionalities should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc10_tbsAuth(driver_setup):
    logging.info("*************** TC-10 ***************")
    logging.info("Verify the TBS Auth")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['TBS'])
        if label_element.text == 'TBS Auth':
            logging.info("TBS Auth is present")
        else:
            raise NoSuchElementException("TBS Auth is not present")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that TBS Auth and their functionalities should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc11_validateauth_button(driver_setup):
    logging.info("*************** TC-11 ***************")
    logging.info("Verify the Auth Button")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        label_element = driver_setup.find_element(By.XPATH, ElementXpath['validatebutton'])
        if label_element.text == 'Validate Challenge Auth':
            logging.info("Validate Challenge Auth button is present")
        else:
            raise NoSuchElementException("Validate Challenge Auth button is not present")

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Check that Validate Challenge Auth button should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc12_download_RawCertificate(driver_setup):
    logging.info("*************** TC-12 ***************")
    logging.info("Verify the download the Raw Certificate")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()
        pass_or_fail = file_to_find(driver_setup, ty='txt', base_filename1='raw-certificate', xpath='qi_Download')

    except NoSuchElementException as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Upon clicking Download (Raw Certificate), verify that the file should saved in the download path."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def tc13_download_Manufacturer_certificate(driver_setup):
    logging.info("*************** TC-13 ***************")
    logging.info("Verify the download the Manufacturer Certificate")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()

        download_button = driver_setup.find_element(By.XPATH, ElementXpath['qi_Download_cert'])
        download_button.click()

        pass_or_fail = file_to_find(driver_setup, ty='pem', base_filename1='ManufacturerCertificate', xpath='manufacturer_certificate')

    except (NoSuchElementException, FileNotFoundError) as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Upon clicking Download (Manufacturer_certificate), verify that the file should saved in the download path."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)


def tc14_download_product_certificate(driver_setup):
    logging.info("*************** TC-14 ***************")
    logging.info("Verify the download the Product_certificate")
    pass_or_fail = "Pass"
    remark = ""

    try:
        Qi = driver_setup.find_element(By.XPATH, ElementXpath['Qi'])
        Qi.click()

        download_button = driver_setup.find_element(By.XPATH, ElementXpath['qi_Download_cert'])
        download_button.click()

        pass_or_fail = file_to_find(driver_setup, ty='pem', base_filename1='ProductUnitCertificate', xpath='product_certificate')

    except (NoSuchElementException, FileNotFoundError) as e:
        logging.error(str(e))
        pass_or_fail = "Fail"
        remark = str(e)

    testdescription = "Upon clicking Download (Product_certificate), verify that the file should saved in the download path."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, testdescription, pass_or_fail, remark)

def file_to_find(driver_setup, ty, base_filename1, xpath):
    time.sleep(2)
    pass_or_fail = "Pass"
    # Set the download directory path
    download_directory = r"C:\Users\GRL\Downloads"
    # Check for existing files with similar names
    base_filename = base_filename1
    existing_files = [f for f in os.listdir(download_directory) if f.startswith(base_filename)]

    # Extract numbers from filenames, ignoring those that don't follow the expected pattern
    existing_numbers = []
    for file in existing_files:
        try:
            number = int(file.split("(")[-1].split(")")[0])
            existing_numbers.append(number)
        except ValueError:
            pass
    # print(existing_files)
    file = len(existing_files)
    # print(file)
    # Form the new filename

    if file == 0:
        #print("1st")
        new_filename = f"{base_filename}.{ty}"
    else:
        #print("2nd")
        # Find the latest number or default to 0 if none are found
        latest_number = max(existing_numbers, default=0)

        # Form the new filename
        new_filename = f"{base_filename} ({latest_number + 1}).{ty}"

    # Find and click the download button
    download_button = driver_setup.find_element(By.XPATH, ElementXpath[xpath])
    download_button.click()

    # Wait for the file to be downloaded (adjust the timeout as needed)
    time.sleep(10)
    downloaded_file_path = os.path.join(download_directory, new_filename)

    # Wait for the file to be downloaded
    os.path.exists(downloaded_file_path)

    # Check if the file exists
    if os.path.exists(downloaded_file_path):
        logging.info(f"Download successful! File '{new_filename}' is present.")
    else:
        logging.error(f"Download failed! File '{new_filename}' is not present.")
        pass_or_fail = "Fail"

    return pass_or_fail

def loadtrace(driver_setup):
    element = WebDriverWait(driver_setup, 10).until(
            EC.element_to_be_clickable((By.XPATH, ElementXpath['Result_tab'])))
    element.click()
    # Click on the image with the specified XPath
    image = driver_setup.find_element(By.XPATH, '(//img[@src="./images/chartIcons/PNG/load.png"])[2]')
    image.click()
    time.sleep(3)
    try:
        popup = driver_setup.find_element(By.XPATH, '//div[@class="custom-scroll-popup "]')
        #logging.warning("Existing test results will be cleared from application if a new capture file is loaded."
        popup_ok = driver_setup.find_element(By.XPATH, '//button[text()="Ok"]')
        popup_ok.click()
        time.sleep(3)
    except Exception as e:
        pass
    Keyboard = Controller()     
    if Element == "TPT":
        pass
    else:
        path = setting['grltrace_path']

    modified_path = path.replace("/", "\\")
    # Provide the file path using send_keys
    Keyboard.type(modified_path)
    time.sleep(5)
    Keyboard.press(Key.enter)
    Keyboard.release(Key.enter)
    time.sleep(8)
    logging.info("Capture file is loaded.")