import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_ConnectionTab_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def tc1_check_readme_file():
    logging.info("*************** TC-1 ***************")
    logging.info("Check the ReadMe File is present")
    pass_or_fail = "Pass"
    remark = ""

    # Set the path to the directory
    if Element == "TPT":
        directory_path = r"C:\GRL\GRL-C3-MP-TPT"
    else:
        directory_path = r"C:\GRL\GRL-C3-MP-TPR"

    # Check if the ReadMe file exists in the directory
    readme_path = os.path.join(directory_path, "ReadMe.txt")

    try:
        if os.path.exists(readme_path):
            logging.info(f"The ReadMe file was found in the directory at: {readme_path}")
        else:
            raise FileNotFoundError(f"The ReadMe file was not found in the directory at: {directory_path}")
    except FileNotFoundError as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    test_description = f"Verify that ReadMe.Txt File should be present in the directory: {directory_path}"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_check_release_file():
    logging.info("*************** TC-2 ***************")
    logging.info("Check the Release Notes File is present")
    pass_or_fail = "Pass"
    remark = ""

    # Set the path to the directory
    if Element == "TPT":
        directory_path = r"C:\GRL\GRL-C3-MP-TPT"
        release_notes_path = os.path.join(directory_path, "GRL-C3-MP-TPT Release Notes.pdf")
    else:
        directory_path = r"C:\GRL\GRL-C3-MP-TPR"
        release_notes_path = os.path.join(directory_path, "GRL-C3-MP-TPR Release Notes.pdf")

    try:
        if os.path.exists(release_notes_path):
            logging.info(f"The Release Notes was found in the directory at: {release_notes_path}")
        else:
            raise FileNotFoundError(f"The Release Notes was not found in the directory at: {directory_path}")
    except FileNotFoundError as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    test_description = f"Verify that release File should be present in the directory: {directory_path}"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_check_eload_file():
    logging.info("*************** TC-3 ***************")
    logging.info("Check the Eload File is present")
    pass_or_fail = "Pass"
    remark = ""

    if Element == "TPT":
        # Set the path to the directory
        directory_path = r"C:\GRL\GRL-C3-MP-TPT\Firmware_Files\MPP"
    else:
        directory_path = r"C:\GRL\GRL-C3-MP-TPR\Firmware_Files"

    # Check if the PPS_ELOAD.bin file exists in the directory
    pps_eload_path = os.path.join(directory_path, "PPS_ELOAD.bin")

    try:
        if os.path.exists(pps_eload_path):
            logging.info(f"PPS_ELOAD.bin was found in the directory at: {pps_eload_path}")
        else:
            raise FileNotFoundError(f"PPS_ELOAD.bin not found in the directory at: {directory_path}")
    except FileNotFoundError as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    test_description = f"Verify that eload File should be present in the directory: {directory_path}"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_check_firmware_file():
    logging.info("*************** TC-4 ***************")
    logging.info("Check all the Firmware Files are present")
    pass_or_fail = "Pass"
    remark = ""

    if Element == "TPT":
        # Set the path to the directory
        directory_path = r"C:\GRL\GRL-C3-MP-TPR\Firmware_Files"
    else:
        directory_path = r"C:\GRL\GRL-C3-MP-TPR\Firmware_Files"

    # List of files to check
    files_to_check = ["BOOT.BIN", "image.ub", "start.sh", "ShortFixture.bin"]

    try:
        missing_files = []

        # Check if each file exists in the directory
        for file_name in files_to_check:
            file_path = os.path.join(directory_path, file_name)
            if not os.path.exists(file_path):
                missing_files.append(file_name)

        # Print missing files
        if missing_files:
            logging.info("The following files are missing:")
            for file_name in missing_files:
                logging.error(file_name)
        else:
            logging.info(f"All files are present{files_to_check}.")
    except FileNotFoundError as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    test_description = f"Verify that Firmware Files should be present in the directory: {directory_path}"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc5_open_browser(driver_setup, mode):
    logging.info("*************** TC-5 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc6_browser_tab_title(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Browser Tab Title")
    pass_or_fail = "Pass"
    remark = ""
    time.sleep(4)
    try:
        actual_title = BrowserTitle(driver_setup)
        if actual_title == 'GRL-C3-TPT Software':
            expected_title = yaml_msg("BPP")
        else:
            if Element == "TPT":
                expected_title = yaml_msg("TITLE")
            else:
                expected_title = yaml_msg("TITLE1")

        if actual_title == expected_title:
            logging.info(f"The browser title in the UI tab is: {actual_title}")
        else:
            logging.critical(
                f"The browser title in the UI tab Value is Mismatched. Actual Value: {actual_title}, Expected Value: {expected_title}")
            pass_or_fail = "Fail"
        time.sleep(5)
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    test_description = f"Verify the browser Title {actual_title}"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def BrowserTitle(driver_setup):
    act_title = driver_setup.title
    return act_title

def tc7_scannetwork_enablestate(driver_setup):
    logging.info("*************** TC-7 ***************")
    logging.info("Check that the Scan Network Button is present and whether it's clickable or not")

    pass_or_fail = "Pass"
    remark = ""  # Initialize remark as an empty string

    try:
        # Use the WebDriver to find the button element
        button_elements = driver_setup.find_elements(By.XPATH, ElementXpath['scanbutton'])
        
        # Check if the button is available
        if button_elements:
            logging.info("The Scan Network button is visible on the UI.")
            # Select the first button element from the list (assuming there's only one)
            button_element = button_elements[0]
            
            # Check if the button is clickable
            if button_element.is_displayed() and button_element.is_enabled():
                logging.info("The Scan Network button is clickable.")
            else:
                logging.critical("The Scan Network button is not clickable at the moment.")
                pass_or_fail = "Fail"
                remark = "The Scan Network button is not clickable at the moment."
        else:
            logging.critical("The Scan Network button is not visible on the UI.")
            pass_or_fail = "Fail"
            remark = "The Scan Network button is not visible on the UI."
    except NoSuchElementException as e:
        logging.error(f"Element not found: {str(e)}")
        pass_or_fail = "Fail"
        remark = f"Element not found: {str(e)}"
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"

    test_description = "Please ensure that the Scan Network button is both displayed and functional, allowing users to click on it."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_scanNetworkstate(driver_setup):
    logging.info("*************** TC-8 ***************")
    logging.info("Check that after clicking the Scan network button, the loading icon is visible, and the connect button should be disabled, and vice versa")
    
    pass_or_fail = "Pass"
    remark = ""
    input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])
    # Click on the input element to focus it
    input_element.click()
    time.sleep(3)
    # Select all text in the input field and delete it
    input_element.send_keys(Keys.CONTROL + "a")  # Select all text
    input_element.send_keys(Keys.DELETE)  # Delete the selected text
    # Enter the new IP 
    in_valid = "192.168.255.2"
    input_element.send_keys(in_valid)
    time.sleep(3)
    button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
    button_element.click()
    time.sleep(5)
    try:
        pass_or_fail = "Pass"
        # Find the "Scan Network" and "Connect" buttons
        scan_button = driver_setup.find_element(By.XPATH,  ElementXpath['scanbutton'])
        connect_button = driver_setup.find_element(By.XPATH, ElementXpath['connect_button'])

        # Click the "Scan Network" button
        scan_button.click()

        # Define a wait for the loading icon to appear
        wait = WebDriverWait(driver_setup, 40)  # Adjust the timeout as needed

        # Wait for the loading icon to appear
        try:
            loading_icon = wait.until(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            
            # Check if the "Connect" button is disabled
            if not connect_button.is_enabled():
                #logging.info("'Connect' button is disabled as expected")
                pass
            else:
                #logging.error("Error: 'Connect' button is not disabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not disabled as expected."
            
        except Exception as e:
            logging.error("Loading icon did not appear within the specified time or is not visible.")
            pass_or_fail = "Fail"
            remark= "Loading icon did not appear within the specified time or is not visible."
        # Wait for the loading icon to disappear
        try:
            wait.until_not(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            logging.info("Loading icon has disappeared")
            
            # Check if the "Connect" button is enabled
            if connect_button.is_enabled():
                #logging.info("'Connect' button is enabled as expected.")
                pass
            else:
                #logging.error("Error: 'Connect' button is not enabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not enabled as expected."

        except Exception as e:
            logging.error("Loading icon did not disappear within the specified time or is still visible.")
            pass_or_fail = "Fail"
            remark = "Loading icon did not disappear within the specified time or is still visible." 

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        logging.error(f" The Scan Network button not visible on the UI: {str(e)}")
        remark = " The Scan Network button not visible on the UI"

    test_description = "Check that after clicking the Scan network button, the loading icon is visible, and the connect button should be disabled, and vice versa"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc9_devicedetails(driver_setup):
    logging.info("*************** TC-9 ***************")
    logging.info("Please ensure that the device details keys are correctly present on the connection setup page")
    pass_or_fail = "Pass"
    remark = ""
    method_name = inspect.currentframe().f_code.co_name
    
    try:
        elements_to_check = {
            'Tester_status': 'expected_tester_status',
            'Serial_number': 'expected_serial_number',
            'Firmware_version': 'expected_firmware_version',
            'Next_calibration_date': 'expected_next_calibration_date',
            'Tester_ip_address': 'expected_tester_ip_address',
            'Port': 'expected_port'
        }

        for element_locator, expected_value_key in elements_to_check.items():
            element = driver_setup.find_element(By.XPATH, ElementXpath[element_locator])
            actual_value = element.text
            
            # Log the values with logger information
            logging.info(f"{element_locator.replace('_', ' ').title()}: {actual_value}")

            # Add assert statements to check the values
            assert actual_value == yaml_msg(expected_value_key), f"{element_locator.replace('_', ' ').title()} mismatch: Expected '{yaml_msg(expected_value_key)}', but got '{actual_value}'"

    except Exception as e:
        logging.error(f"An exception occurred: {str(e)}")
        remark = f"An exception occurred: {str(e)}"
        # Handle the exception as needed, for example, take a screenshot, log additional details, or raise it again.
        pass_or_fail = "Fail"
        raise e

    test_description = " Check the Device details table should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc10_addressText(driver_setup):
    logging.info("*************** TC-10 ***************")
    logging.info("Verify the address Text")

    pass_or_fail = "Pass"
    remark = ""

    try:
        time.sleep(2)
        # Find the element by its CSS selector
        element = driver_setup.find_element(By.XPATH, ElementXpath['TPTAddressText'])

        # Get the text content of the element
        element_text = element.text

        act_title = driver_setup.title
        if act_title == 'GRL-C3-TPT Software':
            desired_text = yaml_msg('IP_Address_Title_BPP')
        else:
            # Get the desired text based on the Element variable
            desired_text_key = "IP_Address_Title" if Element == 'TPT' else "IP_Address_Title1"
            desired_text = yaml_msg(desired_text_key)

        # Check if the desired text is present in the element's text
        if desired_text == element_text:
            logging.info(f"Desired text '{element_text}' is present.")
        else:
            logging.error(f"Desired text '{element_text}' is not present.")
            pass_or_fail = "Fail"
            remark = f"Desired text '{element_text}' is not present."

    except NoSuchElementException as e:
        pass_or_fail = "Fail"
        remark = f"Element not found: {str(e)}"
        logging.error(f"Element not found: {str(e)}")

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(f"An unexpected error occurred: {str(e)}")

    test_description = "Verify that the Address Text should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc11_defaultIp(driver_setup):
    logging.info("*************** TC-11 ***************")
    if Element == 'TPT':
        logging.info("Clicking the scan network, verify default IP is present in the TPT IP Address input box and connect")
    else:
        logging.info("Clicking the scan network, verify default IP is present in the TPR IP Address input box and connect")

    check = setting["tester_connected"]
    check1 = setting["static_dynamic"]
    remark = ""
    if check == True:
        if check1 == "Static":
            try:
                pass_or_fail = "Pass"
                ip = 'DefaultIP'
                elements = driver_setup.find_elements(By.CSS_SELECTOR, ElementXpath['Ipdropdown'])
                element_req = None
                for element in elements:
                    ip_element = element.find_element(By.TAG_NAME,ElementXpath['option'])
                    ip_address = ip_element.get_attribute('value')
                    Defaultip = yaml_msg("Defaultip")
                    if ip_address == Defaultip:
                        element_req = element
                logging.info(f"The Default IP is :{Defaultip}")
                if element_req is not None:  # Check if element_req is assigned
                    time.sleep(2)
                    element_req.click()   
                connected = False
                while not connected:
                    try:
                        # Try to find the image element using the provided XPath
                        image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
                        
                        # If the image is found, set connected to True and break out of the loop
                        connected = True
                        logging.info("Tester is connected with Default IP")

                    except Exception as e:
                        try:
                            # If the image is not found, try to click the button
                            button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                            button_element.click()
                            #print("Clicked the connection setup button")
                            time.sleep(7)
                            # You can add additional logic here if needed after clicking the button
                            
                        except Exception as e:
                            # If both the image and button are not found, print a message and wait before trying again
                            logging.info("c3 is not connected, retrying...")
                            time.sleep(5)  # Adjust the sleep duration as needed

            except Exception as e:
                # If OpenBrowser raises an exception, it's a fail
                pass_or_fail = "Fail"
                logging.error(f"DefaultIp failed with error: {str(e)}")
                remark = f"DefaultIp failed with error: {str(e)}"
            # Get the calling method's name using inspect
        else:
            logging.info("Tester is not connected with Default IP so the test is get skipped")
            pass_or_fail = "Skip"
            remark = "Tester is not connected with Default IP so the test is get skipped"
    else:
        logging.info("Tester is not connected so the test is get skipped")
        pass_or_fail = "Skip"
        remark = "Tester is not connected so the test is get skipped"

    test_description = "Pass Default IP Address and Verify the Tester Status"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc12_dynamicIp(driver_setup):
    logging.info("*************** TC-12 ***************")
    logging.info("Verify Connection with Dynamic IP and Verify the Tester Status")
    pass_or_fail = "Pass"
    method_name = inspect.currentframe().f_code.co_name
    remark = ""
    try:
        check = setting["tester_connected"]
        check1 = setting["static_dynamic"]
        if check == True:
            if check1 == "Dynamic":
                ip = 'DynamicIP'
                # Find the input element using its class name
                input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])

                # Click on the input element to focus it
                # input_element.click()

                # # Select all text in the input field and delete it
                # input_element.clear()

                # Enter the new IP
                in_valid = setting["DynamicIP"]
                input_element.send_keys(Keys.CONTROL + "a")  # Select all text
                input_element.send_keys(Keys.DELETE)  # Delete the selected text
                time.sleep(5)
                input_element.send_keys(in_valid)
                # # Click the connection setup button
                # button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                # button_element.click()

                # # Wait for the connection image to appear
                # WebDriverWait(driver_setup, 30).until(
                #     EC.visibility_of_element_located((By.XPATH, ElementXpath['connectonimg']))
                # )

                # logging.info("c3 is connected")

                # TesterStatus(driver_setup, ip)
                # pass_or_fail, remark = check_license(driver_setup)

                connected = False
                while not connected:
                    try:
                        
                        # button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                        # button_element.click()
                        # Try to find the image element using the provided XPath
                        image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])

                        # If the image is found, set connected to True and break out of the loop
                        connected = True
                        logging.info("c3 is connected")

                    except NoSuchElementException:
                        try:
                            # If the image is not found, try to click the button
                            button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                            button_element.click()
                            time.sleep(15)
                        except NoSuchElementException:
                            # If both the image and button are not found, print a message and wait before trying again
                            remark = "c3 is not connected, retrying..."
                            logging.info(remark)
                            time.sleep(5)  # Adjust the sleep duration as needed

                TesterStatus(driver_setup, ip)
                pass_or_fail, remark = check_license(driver_setup)
            
            else:
                logging.info("Tester is not connected with Dynamic IP so the test is get skipped")
                pass_or_fail = "Skip"
                remark = "Tester is not connected with Dynamic IP so the test is get skipped"

        else:
            logging.info("Tester is not connected, so the test is skipped")
            pass_or_fail = "Skip"
            remark = "Tester is not connected, so the test is skipped"

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Pass dynamicIP Address and Verify the Tester Status"
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc13_invalidIP(driver_setup):
    logging.info("*************** TC-13 ***************")
    logging.info("Pass Invalid IP Address and Verify the Tester Status")
    driver_setup.refresh()

    try:
        pass_or_fail = "Pass"
        remark = ""
        ip = 'invalid'

        # Find the input element using its class name
        input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])

        # # Click on the input element to focus it
        input_element.click()

        # # Select all text in the input field and delete it
        input_element.clear()
        
        # Enter the new IP "1.1.1"
        in_valid = yaml_msg("invalid")
        input_element.send_keys(Keys.CONTROL + "a")  # Select all text
        input_element.send_keys(Keys.DELETE)  # Delete the selected text
        time.sleep(2)
        input_element.send_keys(in_valid)
        logging.info(f"Entered invalid IP: {in_valid}")

        time.sleep(2)
        # Click the connection setup button
        driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton']).click()

        # Wait for the connection off image to appear
        WebDriverWait(driver_setup, 20).until(
            EC.visibility_of_element_located((By.XPATH, ElementXpath['connectoffimg']))
        )

        logging.info("C3 is not Connected")

        # Additional wait if needed
        time.sleep(3)

        TesterStatus(driver_setup, ip)

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        remark = str(e)
        logging.error(f"InvalidIP failed with error: {remark}")

    # Get the calling method's name using inspect
    test_description = "Pass Invalid IP Address and Verify the Tester Status"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc14_unreachableIP(driver_setup):
    logging.info("*************** TC-14 ***************")
    logging.info("Pass Unreachable IP Address and Verify the Tester Status")
    
    try:
        pass_or_fail = "Pass"
        remark = ""
        ip = 'unreachable'

        # Find the input element using its class name
        input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])

        # Click on the input element to focus it
        input_element.click()

        # Select all text in the input field and delete it
        input_element.clear()

        # Enter the new IP "192.168.255.2"
        un_reach = yaml_msg("unreach")
        input_element.send_keys(Keys.CONTROL + "a")  # Select all text
        input_element.send_keys(Keys.DELETE)  # Delete the selected text
        time.sleep(2)
        input_element.send_keys(un_reach)
        logging.info(f"Entered unreachable IP: {un_reach}")

        # Click the connection setup button
        driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton']).click()

        # Wait for the connection off image to appear
        WebDriverWait(driver_setup, 10).until(
            EC.visibility_of_element_located((By.XPATH, ElementXpath['connectoffimg']))
        )

        logging.info("C3 is not Connected")

        # Additional wait if needed
        time.sleep(3)

        TesterStatus(driver_setup, ip)

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        remark = str(e)
        logging.error(f"UnreachableIP failed with error: {remark}")

    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    test_description = "Pass Unreachable IP Address and Verify the Tester Status"
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def tc15_ImageCompare(driver_setup): 
    logging.info("*************** TC-15 ***************")
    logging.info("Verify the Setup Diagram")

    if Element == 'TPT':
        reference_image_path = "Resources\\img\\MPP\\TPT.png"
    else:
        reference_image_path = "Resources\\img\\MPP\\TPR.png"

    threshold = 10  # Define your threshold here
    screenshot_path = "Resources\\img\\screenshot\\screenshot.png"
    pass_or_fail = "Pass"
    remark = ""

    try:
        driver_setup.find_element(By.XPATH, ElementXpath['SetupDiagram_button']).click()
        time.sleep(5)
        image_element = driver_setup.find_element(By.XPATH, ElementXpath['SetupDiagram_image'])
        if image_element.is_displayed():
            logging.info("Image is present")
            time.sleep(5)
            driver_setup.find_element(By.XPATH, ElementXpath['imageok']).click()
        else:
            logging.error("Image is not present")
            remark = "Image is not present"
            pass_or_fail = "Fail"
        
    #     # Capture a screenshot of the image
    #     image_element.screenshot(screenshot_path)

    #     # Load the reference image from the provided file path
    #     reference_image = cv2.imread(reference_image_path, cv2.IMREAD_GRAYSCALE)

    #     # Load the captured screenshot and convert it to grayscale
    #     screenshot_image = cv2.imread(screenshot_path, cv2.IMREAD_GRAYSCALE)

    #     # Compare the two images
    #     difference = cv2.absdiff(reference_image, screenshot_image)

    #     # Check if the images are similar
    #     time.sleep(3)
    #     if cv2.countNonZero(difference) < threshold:
    #         logging.info("Images are similar.")
    #     else:
    #         logging.error("Images are different.")
    #         pass_or_fail = "Fail"

    #         # Optionally, save the difference image for further analysis
    #         cv2.imwrite("Resources\\img\\difference_image.png", difference)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    # finally:
    #     # Delete the screenshot file when done (optional)
    #     try:
    #         if os.path.exists(screenshot_path):
    #             os.remove(screenshot_path)
    #             logging.info(f"Screenshot file deleted: {screenshot_path}")
    #     except Exception as e:
    #         logging.error(f"Failed to delete the screenshot file: {str(e)}")

    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    test_description = "Verify the Setup Diagram"
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc16_shortFixture(driver_setup):
    logging.info("*************** TC-16 ***************")
    logging.info("Verify the Short Fixture Setup")
    pass_or_fail = "Pass"
    remark = ""
    try:
        driver_setup.find_element(By.XPATH, ElementXpath['shortfixture'])
        logging.info("Short Fixture Setup is present on the webpage.")
        try:
            driver_setup.find_element(By.XPATH, ElementXpath['shortconn'])
            logging.info("Short Fixture Connection is present on the webpage.")
            try:
                driver_setup.find_element(By.XPATH, ElementXpath['disconnect'])
                logging.info("Short Fixture Disconnection is present on the webpage.")
            except NoSuchElementException:
                logging.error("Short Fixture Disconnection is present on the webpage.")
        except NoSuchElementException:
            logging.error("Short Fixture Connection is present on the webpage.")
    except NoSuchElementException:
        logging.error("Short Fixture Setup is not present on the webpage.")
        
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    test_description = "Verify the Short Fixture Setup"
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def TesterStatus(driver_setup,value):
    logging.info("Sub_Test_Case: Verify the device detail once it connected to network")
    GUI = []
    b_elements = driver_setup.find_elements(By.CSS_SELECTOR, ElementXpath['Right_space'])

    # Extract and print the values inside <b> </b> elements
    for b_element in b_elements:
        stat = b_element.text.strip()
        GUI.append(stat)
    #logging.info(f"GUI: {GUI}")

    # Initialize values_list with default values
    values_list = ['NA', 'NA', 'NA', 'NA', 'NA', 'NA']

    if value == 'DefaultIP' or value == 'DynamicIP':
        # Initialize a dictionary to map values to their replacements
        value_mappings = {
            0: 'Connected',
            1: 'Disconnected'
        }

        # Initialize an empty list to store the values
        values_list = []
        if Element == "TPT":
            MPP = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"
        else:
            MPP = "C:\\GRL\\GRL-C3-MP-TPR\\AppData\\AppProperty.json"

        try:
            # Open the file for reading
            with open(MPP, 'r') as file:
                # Read the entire contents of the file
                file_contents = file.read()
                
                try:
                    # Try to parse the file contents as JSON
                    data = json.loads(file_contents)
                    
                    # Specify the keys you want to retrieve
                    keys_to_retrieve = ["CS_TesterStatus", "CS_SerialNumber", "CS_FirmwareVersion","C3_CalibrationDueData","CS_C3IpAddress", "CS_C3PortNumber"]

                    # Extract and append the values to the list
                    for key in keys_to_retrieve:
                        if key in data:
                            property_value = data[key]['PropertyValue']
                            # Map 'Connected' and 'Disconnected' to 0 and 1
                            if property_value in value_mappings:
                                property_value = value_mappings[property_value]
                            # Replace empty values with 'NA'
                            if not property_value:
                                property_value = 'NA'
                            values_list.append(property_value)
                        else:
                            values_list.append('NA')  # If the key is not found, append 'NA'
                        
                except json.JSONDecodeError as e:
                    logging.error(f"Error decoding JSON: {str(e)}")
                
        except FileNotFoundError:
            logging.error(f"File not found at path: {MPP}")

        # Print the values list
        logging.info(f"values_list: {values_list}")

    elif value == 'invalid':
        in_validlist = yaml_msg("in_validlist")
        values_list = in_validlist
        logging.info(f"Values List: {values_list}")

    elif value == 'unreachable':
        un_reach = yaml_msg("un_reach")
        values_list = un_reach
        logging.info(f"Values List: {values_list}")

    logging.info("Checking the key pair value are present as per expected")
    names = ['Tester Status', 'Serial Number', 'Firmware Version', 'C3_CalibrationDueData', 'IpAddress', 'PortNumber']
    # Iterate through the lists and compare the elements with assert
    for i in range(len(GUI)):
        element1 = GUI[i]
        element2 = values_list[i]
        are_equal = element1 == element2
        result = "Equal" if are_equal else "Not Equal"
        logging.info(f"{names[i]}: {result}")
        print(element1,element2)
        assert are_equal, f"{names[i]} is not equal: {element1} (List 1) != {element2} (List 2)"

    #print(GUI,values_list)
    # If all assertions pass, no errors will be raised
    logging.info("All elements are equal.")

def check_license(driver_setup):
    logging.info("Sub_Test_Case: Verify the License info to check whether the License is enabled or not")
    pass_or_fail = "Pass"
    remark = ""

    try:
        try:
            # Task 1: Check if the License Info section is present
            license_info_section = WebDriverWait(driver_setup, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, ElementXpath["license1"]))
            )
            logging.info("License Info section is present")

        except Exception as e:
            logging.error(f"Error while checking License Info section: {str(e)}")
            pass_or_fail = "Fail"
            remark = str(e)

        try:
            # Task 2: Check for specific elements in the License Info section
            license_headers = driver_setup.find_elements(By.CLASS_NAME, ElementXpath["license2"])
            found_texts = [header.text for header in license_headers]

            if "Module Name" in found_texts and "License Type" in found_texts:
                logging.info("Module Name and License Type are present")
            else:
                logging.error("Module Name and/or License Type are not present")
                pass_or_fail = "Fail"
                remark = "Module Name and/or License Type are not present"

        except Exception as e:
            logging.error(f"Error while checking License elements: {str(e)}")
            pass_or_fail = "Fail"
            remark = str(e)

        act_title = driver_setup.title
        if act_title == "GRL-C3-TPT Software":
            module_names = []
            license_types = []

            try:
                # Find module names
                module_name_elements = driver_setup.find_elements(By.XPATH, ElementXpath["ModuleName"])
                for element in module_name_elements:
                    module_names.append(element.text)

                # Find license types
                license_type_elements = driver_setup.find_elements(By.XPATH, ElementXpath["license_type"])
                for element in license_type_elements:
                    license_types.append(element.text)

                # Log module names and license types together
                for module_name, license_type in zip(module_names, license_types):
                    logging.info("%s: %s", module_name, license_type)

            except NoSuchElementException as e:
                logging.error("Element not found:", e)

        else:
            print("1st else condition")
            try:
                # Task 3: Check for specific elements with text and attributes
                mpp_compliance_element = driver_setup.find_element(By.XPATH, ElementXpath["license3"])
                perm_element = driver_setup.find_element(By.XPATH, ElementXpath["license4"])

                # Check if PERM is in a green color box
                if "box-PERM" in perm_element.get_attribute("class"):
                    logging.info("PERM is in a green color box")
                else:
                    logging.error("PERM is not in a green color box")
                    pass_or_fail = "Fail"
                    remark = "PERM is not in a green color box"
                
                if Element == 'TPT':
                    logging.info("Both MPP PRx Compliance and PERM are present and meet the conditions")
                else:
                    logging.info("Both MPP PTx Compliance and PERM are present and meet the conditions")


            except Exception as e:
                logging.error(f"Error while checking License conditions: {str(e)}")
                pass_or_fail = "Fail"
                remark = str(e)

    except Exception as e:
        logging.error(f"license failed with error: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)

    return pass_or_fail, remark