import json
import time
import yaml
from selenium.webdriver.common.by import By
import logging
import os
import re
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

# Configure logging
log_file = 'ConnectionTab_automation.log'

logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_layout_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_check_grl_image(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("Verify that the GRL Image is present in the left corner of the page")

    pass_or_fail = "Pass"
    remark = ""  # Initializing remark as an empty string
    time.sleep(1)
    try:
        # Find the image element using XPath
        img_element = driver_setup.find_element(By.XPATH, ElementXpath['grlimg'])

        # Check if the image element is present
        if img_element:
            logging.info("Image is present on the page")
        else:
            logging.error("Image is not present on the page")
            pass_or_fail = "Fail"
            remark = "Image is not present on the page"
    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Verify that GRL Image should be present in the left corner of the page"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_check_menu_button(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("Verify that Menu Button should be present in the left corner of the page")
    time.sleep(1)
    pass_or_fail = "Pass"
    remark = ""

    try:
        menu = driver_setup.find_element(By.XPATH, ElementXpath['menubutton'])

        # Check if the menu button is present
        if menu:
            logging.info("Menu Found")
            
            # First time click
            logging.info("1st time, the menu is clicked")
            menu.click()
            time.sleep(2)

            # Check if the menu is closed
            onClose = driver_setup.find_element(By.XPATH, ElementXpath['menuclose'])
            if onClose:
                logging.info("The menu is onClose")
            else:
                logging.error("The menu is not onClose")
                pass_or_fail = "Fail"
                remark = "The menu is not onClose"

            # Second time click
            logging.info("2nd time, the menu is clicked")
            menu.click()
            time.sleep(2)

            # Check if the menu is open
            onOpen = driver_setup.find_element(By.XPATH, ElementXpath['menuopen'])
            if onOpen:
                logging.info("The menu is onOpen")
            else:
                logging.error("The menu is not onOpen")
                pass_or_fail = "Fail"
                remark = "The menu is not onOpen"
        else:
            logging.error("Menu not Found")
            pass_or_fail = "Fail"
            remark = "Menu not Found"

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Verify that Menu Button should be present in the left corner of the page"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_check_swupdate(driver_setup):
    logging.info("*************** TC-4 ***************")
    logging.info("Check that the Software Update Notification Icon should be present")

    pass_or_fail = "Pass"
    remark = ""

    try:
        sw_update = driver_setup.find_element(By.XPATH, ElementXpath['swupdate'])

        # Check if the Software Update button is present
        if sw_update:
            logging.info("Software Update button is present")
        else:
            logging.error("Software Update button is not present")
            pass_or_fail = "Fail"
            remark = "Software Update button is not present"

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Check that the Software Update Notification Icon should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def tc5_swVersion_swName(driver_setup):
    logging.info("*************** TC-5 ***************")
    logging.info("Please ensure that you have installed the latest version of the software and that it is currently running.")

    pass_or_fail = "Pass"
    remark = ""

    try:
        # Find the software name element using XPath
        swName = driver_setup.find_element(By.XPATH, "//p[@class='navbar-primaryText']")
        name = swName.text

        # Use regular expression to extract text inside brackets
        matches = re.search(r'\((.*?)\)', name)

        if matches:
            result = matches.group(1)
            logging.info("Retrieve the Software version")
        else:
            logging.error("Can't able to Retrieve the Software version")
            pass_or_fail = "Fail"
            remark = "Can't able to Retrieve the Software version"

        readJson = get_software_version()

        if result == readJson:
            logging.info(f"The software version is match: GUI:{result} {Element}, JSON:{readJson} {Element}")
        else:
            logging.error(f"The software version is not match: GUI: {result} {Element}, JSON: {readJson} {Element}")
            pass_or_fail = "Fail"
            remark = f"The software version is not match: GUI: {result} {Element}, JSON: {readJson} {Element}"

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Please ensure that you have installed the latest version of the software and that it is currently running."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc6_layout(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Please ensure that all the layout are present")
    pass_or_fail = "Pass"
    remark = ""
    try:
        list1 = ['ConnectionTab','Qi_exe_tab', 'TestConfigurationTab', 'Result_tab', 'Report_tab', 'Report_analyser','Qi', 'Help']
        for value in list1:
            layout = driver_setup.find_element(By.XPATH, ElementXpath[value])
            if layout:
                layout.click()
                logging.info(f"{value}- Tab is present")
            else:
                logging.error(f"{value}- Tab is not present")
                pass_or_fail = "Fail"
                remark = f"{value}- Tab is not present"
    except NoSuchElementException as e:
        pass_or_fail = "Fail"
        remark = f"Element is not present:{e}"
    test_description = "Please ensure that all the layout are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc7_mpp_bpp_switchTab(driver_setup):
    logging.info("*************** TC-7 ***************")
    logging.info("Please ensure that MPP and BPP Switch Tab can be switchable.")
    pass_or_fail = "Pass"
    remark = ""
    try:
        # Find all checkbox elements with the class name "react-switch-handle"
        checkboxes = driver_setup.find_elements(By.CLASS_NAME, ElementXpath['switchtab'])

        # Select the first checkbox element
        checkbox = checkboxes[0]
        # Click the checkbox
        checkbox.click()
        time.sleep(2)
        logging.info("Tab is switched from MPP to BPP")
        # Find the "Cancel" button element
        cancel_button = driver_setup.find_element(By.CLASS_NAME, "popupButton_Cancel")

        # Click the "Cancel" button
        cancel_button.click()
        time.sleep(3)
    except NoSuchElementException as e:
        pass_or_fail = "Fail"
        remark = f"Element is not present:{e}"
    test_description = "Please ensure that MPP and BPP Switch Tab can be switchable."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def get_software_version():
    #Specify the path to the directory and the filename
    if Element == "TPT":
        directory_path = r'C:\GRL\GRL-C3-MP-TPT'
    else:
        directory_path = r'C:\GRL\GRL-C3-MP-TPR'
    readme_filename = 'ReadMe.txt'
    # Create the full path to the README file
    readme_file_path = os.path.join(directory_path, readme_filename)
    try:
        # Check if the file exists
        if os.path.exists(readme_file_path):
            # Open the README file and read its content
            with open(readme_file_path, 'r') as readme_file:
                # Read the content of the file
                readme_content = readme_file.read()

                # Use regular expression to extract the software version
                version_matches = re.search(r'Software Version\s*:\s*([\d.]+)', readme_content)

                if version_matches:
                    software_version = version_matches.group(1)
                    return software_version
                else:
                    return "Software version not found in README file."
        else:
            return"README file not found in the specified directory."

    except Exception as e:
        return f"An error occurred: {str(e)}"