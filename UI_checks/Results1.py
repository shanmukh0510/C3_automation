import json
import time
import yaml
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
import os
import re
import inspect
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill
from pynput.keyboard import Key, Controller
from selenium.common.exceptions import WebDriverException

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

TPTTestcases = read_file('json/TPTTest_Case.json')

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_ResultsTab_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")


def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_Result_tab(driver):
    logging.info("*************** TC-2 ***************")
    logging.info("Check that result tab is visbile and the corresponding page should be loaded.")
    pass_or_fail = "Pass"
    remark = " "
    try:
        # Use WebDriverWait to wait for the element to become clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, ElementXpath['Result_tab']))
        )

        # Check if the element is visible
        if element.is_displayed():
            # Click the element
            element.click()
            logging.info("The Result tab is visible")
            time.sleep(3)
        else:
            logging.error("Result tab is not visible")
            pass_or_fail = "Fail"
            remark = "Result tab is not visible"

        # Find elements matching the updated XPath expression
        test_result_elements = driver.find_element(By.XPATH, ElementXpath['Testresult_text'])
        Ts = test_result_elements.text
        Test_Results = yaml_msg('Test_Results')

        pattern = re.compile(r'^\s*|\s*$')
        Ts = pattern.sub('', Ts)
        Test_Results = pattern.sub('', Test_Results)

        if Ts == Test_Results:
            logging.info(f"The Result corresponding page is load.")
        else:
            logging.error(f"Error: Can't able to load the result page.")
            pass_or_fail = "Fail"
            remark = f"Error: Can't able to load the result page."

        # Find elements matching the updated XPath expression
        Test_Progress_elements = driver.find_element(By.XPATH, ElementXpath['Test_progress'])
        Tp = Test_Progress_elements.text
        Test_Progress = yaml_msg('Test_Progress')

        pattern = re.compile(r'^\s*|\s*$')
        Tp = pattern.sub('', Tp)
        Test_Progress = pattern.sub('', Test_Progress)

        if Tp == Test_Progress:
            logging.info(f"Test Results progress is present.")
        else:
            logging.error(f"Test Results progress is not found.")
            pass_or_fail = "Fail"
            remark = f"Test Results progress is not found."
        
    except Exception as e:
        logging.error("Element is not clickable or not found:", str(e))
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Element is not clickable or not found:", str(e)

    test_description = "Check that result tab is visbile and the corresponding page should be loaded."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_uploadproject(driver):
    logging.info("*************** TC-3 ***************")
    logging.info("Upload the project and check that the file should be uploaded successfully.")
    try:
        pass_or_fail = "Pass"
        remark = ""
        driver.refresh()
        time.sleep(2)
        driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
        time.sleep(2)
        path = setting['CustomLocation']
        upload_project_button = driver.find_element(By.XPATH, ElementXpath['Uploadproject']).click()
        radio_button = driver.find_element(By.ID, ElementXpath['customLocation'])
        radio_button.click()
        element = driver.find_element(By.XPATH, ElementXpath['Filepath_input'])
        element.click()
        element.send_keys(path)
        driver.find_element(By.XPATH, ElementXpath['UploadTestSequence']).click()
        driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit']).click()
        time.sleep(15)
        logging.info("File Uploaded successfully")

    except Exception as e:
        logging.error(f"File to load the file: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"File to load the file: {str(e)}"

    test_description = "Upload the project and check that the file should be uploaded successfully."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_verfiy_Passcount(driver):
    logging.info("*************** TC-4 ***************")
    logging.info("Pass Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = " "
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_Pass'])))
    # Check if the element is visible
    if img_element.is_displayed():
                
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'PASS' element is visible and it's clicked")
        else:
            logging.error("The 'PASS' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'PASS' element is not clickable."
    else:
        logging.error("The 'PASS' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'PASS' element is not visible."

    xpath_img = ElementXpath['Pass_alt']

    # Find matching 'img' elements
    img_elements = driver.find_elements(By.XPATH, xpath_img)

    # Initialize a set to store the unique test case names
    unique_test_case_names = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element in img_elements:
        alt_value = img_element.get_attribute("alt")
        span_title_element = img_element.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name = span_title_element.text.strip()
        
        # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name and test_case_name not in unique_test_case_names:
            unique_test_case_names.add(test_case_name)

    # Print the count of unique test case names
    unique_test_case_name_count = len(unique_test_case_names)
    logging.info(f"Number of Pass test case names: {unique_test_case_name_count}")

    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Pass element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count

def tc5_verfiy_Failcount(driver):
    logging.info("*************** TC-5 ***************")
    logging.info("Fail Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = " "
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_Fail'])))
    # Check if the element is visible
    if img_element.is_displayed():        
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'FAIL' element is visible and it's clicked")
        else:
            logging.error("The 'FAIL' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'FAIL' element is not clickable."
    else:
        logging.error("The 'FAIL' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'FAIL' element is not visible."
    time.sleep(2)
    xpath_img1 = ElementXpath['Fail_alt']

    # Find matching 'img' elements
    img_elements1 = driver.find_elements(By.XPATH, xpath_img1)

    # Initialize a set to store the unique test case names
    unique_test_case_names1 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element1 in img_elements1:
        #time.sleep(1)
        alt_value1 = img_element1.get_attribute("alt")
        span_title_element1 = img_element1.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name1 = span_title_element1.text.strip()

        # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name1 and test_case_name1 not in unique_test_case_names1:
            unique_test_case_names1.add(test_case_name1)

    # Print the count of unique test case names
    unique_test_case_name_count1 = len(unique_test_case_names1)
    logging.info(f"Number of FAIL test case names: {unique_test_case_name_count1}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Fail element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count1

def tc6_verfiy_Inconclusivecount(driver):
    logging.info("*************** TC-6 ***************")
    logging.info("Inconclusive count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_inconclusive'])))
    # Check if the element is visible
    if img_element.is_displayed():       
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'Inconclusive' element is visible and it's clicked")
        else:
            logging.error("The 'INCONCLUSIVE' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'INCONCLUSIVE' element is not clickable."
    else:
        logging.error("The 'INCONCLUSIVE' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'INCONCLUSIVE' element is not visible."
    time.sleep(2)
    xpath_img2 = ElementXpath['INCONCLUSIV_alt']

    # Find matching 'img' elements
    img_elements2 = driver.find_elements(By.XPATH, xpath_img2)

    # Initialize a set to store the unique test case names
    unique_test_case_names2 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element2 in img_elements2:
        #time.sleep(1)
        alt_value2 = img_element2.get_attribute("alt")
        span_title_element2 = img_element2.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name2 = span_title_element2.text.strip()

    # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name2 and test_case_name2 not in unique_test_case_names2:
            unique_test_case_names2.add(test_case_name2)

    # Print the count of unique test case names
    unique_test_case_name_count2 = len(unique_test_case_names2)
    logging.info(f"Number of INCONCLUSIVE test case names: {unique_test_case_name_count2}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Inconclusive element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count2

def tc7_verfiy_Notruncount(driver):          
    logging.info("*************** TC-7 ***************")
    logging.info("NotRun Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_notrun'])))
    # Check if the element is visible
    if img_element.is_displayed():
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'NotRun' element is visible and it's clicked")
        else:
            logging.error("The 'NOTRUN' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'NOTRUN' element is not clickable."
    else:
        logging.error("The 'NOTRUN' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'NOTRUN' element is not visible."
    time.sleep(2)
    xpath_img3 = ElementXpath['Notrun_alt']

    # Find matching 'img' elements
    img_elements3 = driver.find_elements(By.XPATH, xpath_img3)

    # Initialize a set to store the unique test case names
    unique_test_case_names3 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element3 in img_elements3:
        #time.sleep(1)
        alt_value3 = img_element3.get_attribute("alt")
        span_title_element3 = img_element3.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name3 = span_title_element3.text.strip()

    # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name3 and test_case_name3 not in unique_test_case_names3:
            unique_test_case_names3.add(test_case_name3)

    # Print the count of unique test case names
    unique_test_case_name_count3 = len(unique_test_case_names3)
    logging.info(f"Number of NOTRUN test case names: {unique_test_case_name_count3}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Notrun element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count3

def tc8_no_testcase(driver):
    unique_test_case_name_count= tc4_verfiy_Passcount(driver)
    unique_test_case_name_count1 = tc5_verfiy_Failcount(driver)
    unique_test_case_name_count2= tc6_verfiy_Inconclusivecount(driver)
    unique_test_case_name_count3 = tc7_verfiy_Notruncount(driver)
    logging.info("*************** TC-8 ***************")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    totalcase = unique_test_case_name_count + unique_test_case_name_count1 + unique_test_case_name_count2 + unique_test_case_name_count3
    logging.info("Total Number of Test Cases: %s", totalcase)
    
    # Initialize a list to store the counts
    counts = []
    # Define the XPath expression to select the 'strong' element
    xpath_strong = ElementXpath['strong']
    strong_element = driver.find_elements(By.XPATH, xpath_strong)
    # Iterate through the 'strong' elements and extract the counts
    for strong in strong_element:
        # Extract the text from the 'strong' element
        text = strong.text.strip()
        # Split the text to get the count
        parts = text.split('/')
        count = parts[1].strip() if len(parts) > 1 else None
        # Append the count to the list if it's not None
        if count is not None:
            counts.append(count)
    # Print the list of counts
    TP = ['Pass','Fail', 'Inconclusive','Notrun']
    # Compare each index with the total
    for count, TPs in zip(counts, TP):
            if int(count) == totalcase:
                pass
                #logging.info(f"Count {count} matches the total {TPs} count ({totalcase}).")
            else:
                logging.error(f"Count {count} does not match the total {TPs} count ({totalcase}).")
                pass_or_fail = "Fail"
                remark = f"Count {count} does not match the total {TPs} count ({totalcase})."

    counts1 = []
    xpath_strong = ElementXpath['strong']
    strong_element = driver.find_elements(By.XPATH, xpath_strong)
    # Iterate through the 'strong' elements and extract the counts
    for strong in strong_element:
        # Extract the text from the 'strong' element
        text = strong.text.strip()
        # Extract just the "28" from the text
        result = text.split('/')[0].strip()
        # Print the result
        counts1.append(result)

    my_list = counts1[:-1]  # Remove the last element
    #logging.info(my_list)

    my_list1 = [unique_test_case_name_count, unique_test_case_name_count1, unique_test_case_name_count2, unique_test_case_name_count3]
    #logging.info(my_list1)

    labels = ['PASS', 'FAIL', 'INCONCLUSIVE', 'NOTRUN']

    # Iterate through the lists and generate the desired output
    for label, result, string_result in zip(labels, my_list1, my_list):
        if result == int(string_result):
            logging.info(f"The Count of {string_result} matches the total number of {label}, which is also ({string_result}).")
        else:
            logging.error(f"The Count of {string_result} doesn't matches the total number of {label}, which is({string_result}).")
            pass_or_fail = "Fail"
            remark = f"The Count of {string_result} doesn't matches the total number of {label}, which is({string_result})."

    # Get the calling method's name using inspect
    test_description = "Please verify whether the total number of executed test cases matches the expected value."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc9_expand_collapse(driver):
    logging.info("*************** TC-9 ***************")
    logging.info("Expand and Collapse")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    collapse_xpath = ElementXpath['Tg_collapse']
    expand_xpath = ElementXpath['Tg_expand']          
    # Wait for and click the collapse element
    try:
        collapse_element = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, collapse_xpath)))
        collapse_element.click()
        logging.info("All the Test cases is minimized")
    except:
        logging.error("The collapse element is not available or clickable within the timeout.")
        pass_or_fail = "Fail"
        remark = "The collapse element is not available or clickable within the timeout."

    # Wait for and click the expand element
    try:
        expand_element = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, expand_xpath)))
        expand_element.click()
        logging.info("All the Test cases is expand")
    except:
        logging.error("The expand element is not available or clickable within the timeout.")
        pass_or_fail = "Fail"
        remark = "The expand element is not available or clickable within the timeout."

    test_description = "Please verify whether the expand and collapse element is present and accessible or not."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_10_chart_header(driver):
    logging.info("*************** TC-10 ***************")
    logging.info("Chart-Header")
    pass_or_fail = "Pass"
    remark = ""
    try:
        chart_time = driver.find_element(By.XPATH,ElementXpath['chart_Time'])
        chart_time_text = chart_time.text
        # Print the element's XPath and the extracted text
        logging.info(chart_time_text)
        
        
        search_input = driver.find_element(By.XPATH, ElementXpath['chart_search'])

        if search_input.is_displayed():
            logging.info("Search input is available on the page.")
        else:
            logging.error("Search input is not displayed.")
            pass_or_fail = "Fail"
            remark = "Search input is not displayed."

        if Element == "TPT":
            # Find the element with the specified class
            element = driver.find_element(By.XPATH, ElementXpath['chart_grp'])

            # Check if the element is displayed
            if element.is_displayed():
                logging.info("The 'chart_grp' element is available and displayed with 'showChart.svg'.")
            else:
                logging.error("The 'chart_grp' element is not available or not displayed.")
                pass_or_fail = "Fail"
                remark = "The 'chart_grp' element is not available or not displayed."
        else:
            logging.info("Chart is Not Support for TPR")
            pass

        # Find the element with the specified class
        element = driver.find_element(By.XPATH, ElementXpath['chart_packettime'])

        # Check if the element is displayed
        if element.is_displayed():
            logging.info("The 'packet timing sync' element is available and displayed with 'packetLeftAlign.svg'.")
        else:
            logging.error("The 'packet timing sync' element is available and displayed, but the src attribute does not contain 'packetLeftAlign.svg'.")
            pass_or_fail = "Fail"
            remark = "The 'packet timing sync' element is available and displayed, but the src attribute does not contain 'packetLeftAlign.svg'."

        element = driver.find_element(By.XPATH, ElementXpath['chart_filter'])

        # Check if the element is displayed
        if element.is_displayed():
            logging.info("The 'filter-packet-img' element is available and displayed with 'filter-packet.png'.")
        else:
            logging.error("The 'filter-packet-img' element is available and displayed, but the src attribute does not contain 'filter-packet.png'.")
            pass_or_fail = "Fail"
            remark = "The 'filter-packet-img' element is available and displayed, but the src attribute does not contain 'filter-packet.png'."

        element.click()
        time.sleep(2)
        elements = driver.find_elements(By.XPATH, ElementXpath['filterspacket'])
        filters=[]
        # Iterate through each element and print its text
        for element1 in elements:
            text = element1.text
            filters.append(text)

        if Element == "TPT":
            mpp = yaml_msg("tptfilter")
        else:
            mpp = yaml_msg("tprfilter")

        if filters == mpp:
            logging.info("Filters and mpp are identical.")
        else:
            logging.error("Filters and mpp are not identical.")
            pass_or_fail = "Fail"
            remark = "Filters and mpp are not identical."
        
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    test_description = "Please verify all the chart-headers and ensure that all the corresponding elements are present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_11_chart_toolbar(driver):
    logging.info("*************** TC-11 ***************")
    logging.info("Chart-Toolbar")
    pass_or_fail = "Pass"
    remark = " "
    try:
        chart_channeldropdown = driver.find_element(By.XPATH,ElementXpath['chart_channeldropdown'])
                
        #Click the dropdown if it exists
        if chart_channeldropdown:
            chart_channeldropdown.click()
            logging.info("Dropdown clicked successfully.")
            logging.info("Rectified Voltage is checked and it's visible in the yaxis.")
            logging.info("Rectified Current is checked and it's visible in the yaxis.")
        else:
            logging.error("Dropdown element not found")
            pass_or_fail = "Fail"

        checkbox_element = driver.find_element(By.XPATH, ElementXpath["chart_dpdw_IV"])

        #Check if the checkbox element is present
        if checkbox_element:
            checkbox_element.click()
            logging.info("Rectified Voltage is now unchecked.")

        checkbox_element1 = driver.find_element(By.XPATH, ElementXpath["chart_dpdw_IC"])

        #Check if the checkbox element is present
        if checkbox_element1:
            checkbox_element1.click()
            logging.info("Rectified Current is now unchecked.")
            time.sleep(5)
            checkbox_element.click()
            logging.info("Checkbox is now checked.")

            yaxis_labelIV= driver.find_element(By.XPATH, ElementXpath["yaxis_labelIV"])
            if yaxis_labelIV:
                logging.info("Inverter Voltage is checked and it's visible in the yaxis.")
            else:
                logging.info("Inverter Voltage is checked but it's disappered in the yaxis.")
                pass_or_fail = "Fail"
        else:
            logging.error("Checkbox element not found")
            pass_or_fail = "Fail"


        checkbox_element1 = driver.find_element(By.XPATH, ElementXpath["chart_dpdw_IC"])

        # Check if the checkbox element is present
        if checkbox_element1:
            checkbox_element1.click()
            logging.info("Checkbox is now unchecked.")

            time.sleep(5)
            checkbox_element1.click()
            logging.info("Checkbox is now checked.")

            yaxis_labelIC = driver.find_element(By.XPATH, ElementXpath["yaxis_labelIC"])
            if yaxis_labelIC:
                logging.info("Inverter Current is checked and it's visible in the yaxis.")
            else:
                logging.info("Inverter Current is checked but it's disappered in the yaxis.")
        else:
            logging.error("Checkbox element not found")
            pass_or_fail = "Fail"
            
        chart_save = driver.find_element(By.XPATH,ElementXpath['chart_save'])

        # Check if the element is displayed
        if chart_save.is_displayed():
            logging.info("The 'chart_save' element is available and displayed with 'save.png'.")
        else:
            logging.error("The 'chart_save' element is available and displayed, but the src attribute does not contain 'save.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_save' element is available and displayed, but the src attribute does not contain 'save.png'."

        chart_load = driver.find_element(By.XPATH,ElementXpath['chart_load'])

        # Check if the element is displayed
        if chart_load.is_displayed():
            logging.info("The 'chart_load' element is available and displayed with 'load.png'.")
        else:
            logging.error("The 'chart_load' element is available and displayed, but the src attribute does not contain 'load.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_load' element is available and displayed, but the src attribute does not contain 'load.png'."

        chart_clear = driver.find_element(By.XPATH,ElementXpath['chart_clear'])

        # Check if the element is displayed
        if chart_clear.is_displayed():
            logging.info("The 'chart_clear' element is available and displayed with 'recycle.png'.")
        else:
            logging.error("The 'chart_clear' element is available and displayed, but the src attribute does not contain 'recycle.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_clear' element is available and displayed, but the src attribute does not contain 'recycle.png'."

        chart_hand = driver.find_element(By.XPATH,ElementXpath['chart_hand'])

        # Check if the element is displayed
        if chart_hand.is_displayed():
            logging.info("The 'chart_hand' element is available and displayed with 'hand.png'.")
        else:
            logging.error("The 'chart_hand' element is available and displayed, but the src attribute does not contain 'hand.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_hand' element is available and displayed, but the src attribute does not contain 'hand.png'."

        chart_fit = driver.find_element(By.XPATH,ElementXpath['chart_fit'])

        # Check if the element is displayed
        if chart_fit.is_displayed():
            logging.info("The 'chart_fit' element is available and displayed with 'fit.png'.")
        else:
            logging.error("The 'chart_fit' element is available and displayed, but the src attribute does not contain 'fit.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_fit' element is available and displayed, but the src attribute does not contain 'fit.png'."

        chart_HZoomin = driver.find_element(By.XPATH,ElementXpath['chart_HZoomin'])

        # Check if the element is displayed
        if chart_HZoomin.is_displayed():
            logging.info("The 'chart_HZoomin' element is available and displayed with 'Zoomin.png'.")
        else:
            logging.error("The 'chart_HZoomin' element is available and displayed, but the src attribute does not contain 'Zoomin.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_HZoomin' element is available and displayed, but the src attribute does not contain 'Zoomin.png'."

        chart_HZoomout = driver.find_element(By.XPATH,ElementXpath['chart_HZoomout'])

        # Check if the element is displayed
        if chart_HZoomout.is_displayed():
            logging.info("The 'chart_HZoomout' element is available and displayed with 'Zoomout.png'.")
        else:
            logging.error("The 'chart_HZoomout' element is available and displayed, but the src attribute does not contain 'Zoomout.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_HZoomout' element is available and displayed, but the src attribute does not contain 'Zoomout.png'."
        
        chart_VZoomin = driver.find_element(By.XPATH,ElementXpath['chart_VZoomin'])

        # Check if the element is displayed
        if chart_VZoomin.is_displayed():
            logging.info("The 'chart_VZoomin' element is available and displayed with 'Zoomin.png'.")
        else:
            logging.error("The 'chart_VZoomin' element is available and displayed, but the src attribute does not contain 'Zoomin.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_VZoomin' element is available and displayed, but the src attribute does not contain 'Zoomin.png'."

        chart_VZoomout = driver.find_element(By.XPATH,ElementXpath['chart_VZoomout'])

        # Check if the element is displayed
        if chart_VZoomout.is_displayed():
            logging.info("The 'chart_VZoomout' element is available and displayed with 'Zoomout.png'.")
        else:
            logging.error("The 'chart_VZoomout' element is available and displayed, but the src attribute does not contain 'Zoomout.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_VZoomout' element is available and displayed, but the src attribute does not contain 'Zoomout.png'."

        chart_moveleft = driver.find_element(By.XPATH,ElementXpath['chart_moveleft'])
        # Check if the element is displayed
        if chart_moveleft.is_displayed():
            logging.info("The 'chart_moveleft' element is available and displayed with 'moveleft.png'.")
        else:
            logging.error("The 'chart_moveleft' element is available and displayed, but the src attribute does not contain 'moveleft.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_moveleft' element is available and displayed, but the src attribute does not contain 'moveleft.png'."

        chart_moveright = driver.find_element(By.XPATH,ElementXpath['chart_moveright'])

        # Check if the element is displayed
        if chart_moveright.is_displayed():
            logging.info("The 'chart_moveright' element is available and displayed with 'moveright.png'.")
        else:
            logging.error("The 'chart_moveright' element is available and displayed, but the src attribute does not contain 'moveright.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_moveright' element is available and displayed, but the src attribute does not contain 'moveright.png'."

        chart_top = driver.find_element(By.XPATH,ElementXpath['chart_top'])
        # Check if the element is displayed
        if chart_top.is_displayed():
            logging.info("The 'chart_top' element is available and displayed with 'top.png'.")
        else:
            logging.error("The 'chart_top' element is available and displayed, but the src attribute does not contain 'top.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_top' element is available and displayed, but the src attribute does not contain 'top.png'."

        chart_down = driver.find_element(By.XPATH,ElementXpath['chart_down'])

        # Check if the element is displayed
        if chart_down.is_displayed():
            logging.info("The 'chart_down' element is available and displayed with 'down.png'.")
        else:
            logging.error("The 'chart_down' element is available and displayed, but the src attribute does not contain 'down.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_down' element is available and displayed, but the src attribute does not contain 'down.png'."

        chart_vertical = driver.find_element(By.XPATH,ElementXpath['chart_vertical'])
        # Check if the element is displayed
        if chart_vertical.is_displayed():
            logging.info("The 'chart_vertical' element is available and displayed with 'vertical.png'.")
        else:
            logging.error("The 'chart_vertical' element is available and displayed, but the src attribute does not contain 'vertical.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_vertical' element is available and displayed, but the src attribute does not contain 'vertical.png'."

        chart_horizontal = driver.find_element(By.XPATH,ElementXpath['chart_horizontal'])

        # Check if the element is displayed
        if chart_horizontal.is_displayed():
            logging.info("The 'chart_horizontal' element is available and displayed with 'horizontal.png'.")
        else:
            logging.error("The 'chart_horizontal' element is available and displayed, but the src attribute does not contain 'horizontal.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_horizontal' element is available and displayed, but the src attribute does not contain 'horizontal.png'."

        chart_merge = driver.find_element(By.XPATH,ElementXpath['chart_merge'])
        # Check if the element is displayed
        if chart_merge.is_displayed():
            logging.info("The 'chart_merge' element is available and displayed with 'merge.png'.")
        else:
            logging.error("The 'chart_merge' element is available and displayed, but the src attribute does not contain 'merge.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_merge' element is available and displayed, but the src attribute does not contain 'merge.png'."

        chart_unmerge = driver.find_element(By.XPATH,ElementXpath['chart_unmerge'])

        # Check if the element is displayed
        if chart_unmerge.is_displayed():
            logging.info("The 'chart_unmerge' element is available and displayed with 'unmerge.png'.")
        else:
            logging.error("The 'chart_unmerge' element is available and displayed, but the src attribute does not contain 'unmerge.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_unmerge' element is available and displayed, but the src attribute does not contain 'unmerge.png'."

        chart_screenshot = driver.find_element(By.XPATH,ElementXpath['chart_screenshot'])

        # Check if the element is displayed
        if chart_screenshot.is_displayed():
            logging.info("The 'chart_screenshot' element is available and displayed with 'screenshot.png'.")
        else:
            logging.error("The 'chart_screenshot' element is available and displayed, but the src attribute does not contain 'screenshot.png'.")
            pass_or_fail = "Fail"
            remark = "The 'chart_screenshot' element is available and displayed, but the src attribute does not contain 'screenshot.png'."

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    test_description = "Please verify all the chart-Toolbar and ensure that all the corresponding elements are present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_12_loadtrace(driver):
    logging.info("--------------------TC-12--------------------")
    # Click on the image with the specified XPath
    pass_or_fail = "Pass"
    remark = ""
    image = driver.find_element(By.XPATH, '(//img[@src="./images/chartIcons/PNG/load.png"])[2]')
    image.click()
    time.sleep(3)
    try:
        popup = driver.find_element(By.XPATH, '//div[@class="custom-scroll-popup "]')
        if popup:
            logging.warning("Existing test results will be cleared from application if a new capture file is loaded.")
            logging.info("Click OK to load the capture file.")
            popup_ok = driver.find_element(By.XPATH, '//button[text()="Ok"]')
            popup_ok.click()
            time.sleep(3)
    except Exception as e:
        pass

    Keyboard = Controller()   

    # Extract the value of "grltrace_path" from the JSON data
    grltrace_path = setting.get("grltrace_path")  
    fixed_path = grltrace_path.replace('/', '\\')

    if Element == "TPT":
        path = fixed_path
    else:
        path = fixed_path
    # Provide the file path using send_keys
    Keyboard.type(path)
    time.sleep(3)
    Keyboard.press(Key.enter)
    Keyboard.release(Key.enter)
    time.sleep(7)
    logging.info("--------------------Trace load successfully--------------------")
    test_description = "Load the grl trace file and check that the file is uploaded successfully"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)