import json
import time
import yaml
from selenium.webdriver.common.by import By
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import WebDriverException

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
print(Element)
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

# result_colors = {
#     "Pass": "00FF00",  # Green
#     "Fail": "FF0000",  # Red
# }

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_Report__automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def tc2_reportpage(driver_setup):
    logging.info("*************** TC-2 ***************")
    pass_or_fail = "Pass"
    remark = ""

    try:
        if Element == 'TPT':
            logging.info("TPT Test validation is completed")
            path = tc3_reportpage(driver_setup)
            openbrowser(driver_setup, path)
            tc4_htmlreport(driver_setup)

        else:
            logging.info("TPR Test validation is completed")
            path = tc3_reportpage(driver_setup)
            openbrowser(driver_setup, path)
            tc4_htmlreport(driver_setup)

    except Exception as e:
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
        logging.error(remark)

    test_description = "Verify that the test completes the execution"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_reportpage(driver_setup):
    logging.info("*************** TC-3 ***************")

    pass_or_fail = "Pass"
    remark = ""

    try:
        if Element == 'TPT':
            path = r'C:\GRL\GRL-C3-MP-TPT\Report'
        else:
            path = r'C:\GRL\GRL-C3-MP-TPR\Report'

        new_folder_name = 'AAA'

        # Create the new folder if it doesn't exist
        new_folder_path = os.path.join(path, new_folder_name)
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)
        else:
            pass
        time.sleep(5)
        driver_setup.find_element(By.XPATH, ElementXpath['Report_tab']).click()
        time.sleep(8)

        detailed_result_element = driver_setup.find_element(By.XPATH, ElementXpath['viewreport'])
        if detailed_result_element.is_displayed():
            logging.info("ViewReport/Detailed Compliance Test Result element is present")
        else:
            logging.error("ViewReport/Detailed Compliance Test Result element is not currently visible on the page")
            pass_or_fail = "Fail"
            remark.append("ViewReport/Detailed Compliance Test Result element is not currently visible")

        detailed_result_element1 = driver_setup.find_element(By.XPATH, ElementXpath['currentHTML'])
        if detailed_result_element1.is_displayed():
            detailed_result_element1.click()
            driver_setup.find_element(By.XPATH, ElementXpath['currentdownload']).click()
            time.sleep(5)
            logging.info("Download Current HTML report is present and it's downloaded")
        else:
            logging.error("Download Current HTML report is not currently visible on the page")
            pass_or_fail = "Fail"
            remark = "Download Current HTML report is not currently visible"

        detailed_result_element4 = driver_setup.find_element(By.XPATH, ElementXpath['synthetic'])
        if detailed_result_element4.is_displayed():
            logging.info("Synthetic File is present")
        else:
            logging.error("Synthetic File is not currently visible on the page")
            pass_or_fail = "Fail"
            remark = "Synthetic File is not currently visible"

        detailed_result_element2 = driver_setup.find_element(By.XPATH, ElementXpath['BSUT'])
        if detailed_result_element2.is_displayed():
            detailed_result_element2.click()
            driver_setup.find_element(By.XPATH, ElementXpath['currentdownload']).click()
            time.sleep(5)
            logging.info("Download Current BSUT Report data is present and it's downloaded")
        else:
            logging.error("Download Current BSUT Report data is not currently visible on the page")
            pass_or_fail = "Fail"
            remark = "Download Current BSUT Report data is not currently visible"

        detailed_result_element3 = driver_setup.find_element(By.XPATH, ElementXpath['datamanagement'])
        if detailed_result_element3.is_displayed():
            logging.info("Report Data Management is present")
            logging.info("Delete Report and Download report option is present")
            detailed_result_element3.click()

            # Before Deleting the report
            logging.info("Before Deleting the report")
            pass_or_fail, remark = folder_check(driver_setup)

            # After Delete the report
            # Now Delete the report and check folder is deleted or not.
            #table = driver_setup.find_element(By.XPATH, ElementXpath['delete']).click()
            # Find the tr element containing 'AAA'
            row_with_aaa = driver_setup.find_element(By.XPATH, "//tr[td[text()='AAA']]")

            # Find the 'Delete Report' button within the identified row
            delete_button = row_with_aaa.find_element(By.XPATH, ".//button[text()='Delete Report']").click()

            time.sleep(5)
            driver_setup.find_element(By.XPATH, ElementXpath['currentdownload']).click()
            time.sleep(7)
            logging.info("After Deleting the report")
            pass_or_fail, remark = folder_check(driver_setup)

        else:
            logging.error("Report Data Management is present is not currently visible on the page")
            pass_or_fail = "Fail"
            remark = "Report Data Management is not currently visible"

    except Exception as e:
        pass_or_fail = "Fail"
        remark= f"An unexpected error occurred: {str(e)}"
        #logging.error(",".join(remark))

    test_description = "Verify the Report page and their functionality"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

    # Find the iframe element by its ID
    iframe_element = driver_setup.find_element(By.ID, "myId")

    # Get the value of the src attribute
    iframe_src = iframe_element.get_attribute("src")
    return iframe_src

def tc4_htmlreport(driver_setup):
    logging.info("*************** TC-4 ***************")
    pass_or_fail = "Pass"
    remark = ""
    if Element == 'TPT':
        html = ['html1', 'html2', 'html3', 'html4', 'html5']
        any_test_case_missing = False
    else:
        html = ['html1', 'html2', 'html3', 'html4', 'html5', 'html6', 'html7']
        any_test_case_missing = False

    for value in html:
        html_element = driver_setup.find_element(By.XPATH, ElementXpath[value])
        if not html_element.is_displayed():
            any_test_case_missing = True
            break  # No need to check further, we found one missing test case

    if any_test_case_missing:
        logging.error("One or more test cases which we ran are not present in the HTML Report")
        pass_or_fail = "Fail"
        remark = "One or more test cases which we ran are not present in the HTML Report"
    else:
        logging.info("All the test case which we ran are present in the HTML Report")
        
    test_description = "Verify that the HTML Report and it testcases are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def folder_check(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    # Fetch folder names using Selenium
    table = driver_setup.find_element(By.CLASS_NAME, "results-manager")
    time.sleep(3)
    folder_names_selenium = [row.find_elements(By.TAG_NAME, "td")[1].text for row in table.find_elements(By.TAG_NAME, "tr")[1:]]

    # Fetch folder names using os module
    if Element == 'TPT':
        path = r'C:\GRL\GRL-C3-MP-TPT\Report'
    else:
        path = r'C:\GRL\GRL-C3-MP-TPR\Report'
    folder_names_os = [folder for folder in os.listdir(path) if os.path.isdir(os.path.join(path, folder))]
    time.sleep(3)
    # Compare the two lists
    if set(folder_names_selenium) == set(folder_names_os):
        logging.info(f"Folder names obtained from GUI {len(folder_names_selenium)} match the folder names from the OSPath {len(folder_names_os)}.")
    else:
        logging.error(f"Folder names obtained from GUI {len(folder_names_selenium)} do not match the folder names from the OSPath {len(folder_names_os)}.")
        pass_or_fail = "Fail"
        remark = f"Folder names obtained from GUI {len(folder_names_selenium)} do not match the folder names from the OSPath {len(folder_names_os)}."
    
    return pass_or_fail, remark

def openbrowser(driver_setup,path):
    driver_setup.get(path)
    driver_setup.maximize_window()
    time.sleep(3)
