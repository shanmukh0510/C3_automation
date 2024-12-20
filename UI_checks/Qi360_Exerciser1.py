import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from pynput.keyboard import Key, Controller
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException
import random


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_Qi_Exerciser_Automation_360khz{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_verify_QiExercise(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("Verify the Qi_Exerciser")
    
    pass_or_fail = "Pass"
    remark = ""
    try:
        time.sleep(5)
        driver_setup.find_element(By.XPATH, ElementXpath['Qi_exe_tab']).click()
        logging.info("Qi-Exerciser is clicked ")
        time.sleep(5)
        driver_setup.find_element(By.XPATH, ElementXpath['toggle']).click()
        try:
            verify = driver_setup.find_element(By.XPATH, ElementXpath['Qi_exe_ver'])
            if verify:
                logging.info("Landed on the Qi_Exerciser page")
        except Exception as e:
            pass_or_fail = "Fail"
            logging.error("Not landed on the Qi_Exerciser page")
            remark = f"Not landed on the Qi_Exerciser page"
        time.sleep(5)
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"The attempt to click the Qi-Exerciser is failed: {str(e)}")
        remark = f"The attempt to click the Qi-Exerciser is failed: {str(e)}"

    test_description = "Verify the Qi_Exercise."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_360khz_remove_packet_seq(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("Verify the 360khz_remove_packet_seq")
    time.sleep(3)
    pass_or_fail, remark = removeall(driver_setup)

    test_description = "Verify the 360khz_remove_packet_seq."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_360khz_reset_packet_seq(driver_setup):
    logging.info("*************** TC-4 ***************")
    logging.info("Verify the 360khz_reset_packet_seq")
    
    time.sleep(2)
    pass_or_fail, remark = reset_packetSeq(driver_setup)
    driver_setup.find_element(By.XPATH, ElementXpath['toggle']).click()
    test_description = "Verify the 360khz_reset_packet_seq."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc5_360khz_reset_exerciser(driver_setup):
    logging.info("*************** TC-5 ***************")
    logging.info("Verify the 360khz_reset_exerciser")
    driver_setup.find_element(By.XPATH, ElementXpath['remove_all']).click()
    time.sleep(2)

    pass_or_fail, remark = reset_exerciser(driver_setup)
    driver_setup.find_element(By.XPATH, ElementXpath['toggle']).click()
    test_description = "Verify the 360khz_reset_exerciser."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc6_360khz_add_packets_info(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Verify the 360khz_add_packets_details")
    driver_setup.find_element(By.XPATH, ElementXpath['remove_all']).click()
    time.sleep(3)

    pass_or_fail, remark = add_packets_details(driver_setup)
    test_description = "Verify the 360khz_add_packets_details."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc7_360khz_add_packets_sequence(driver_setup):
    logging.info("*************** TC-7 ***************")
    logging.info("Verify the 360khz_add_packets_sequence")
    pass_or_fail, remark = add_packets_sequence(driver_setup)


    test_description = "Verify the 360khz_add_packets_sequence."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_360khz_add_packets_sequence_checks1_ID_Config(driver_setup):
    logging.info("*************** TC-8 ***************")
    logging.info("Verify the 360khz_add_packets_sequence_checks1_ID_Config")

    packet_xpath = 'ID_packet'
    config_xpath = 'srq_packet'
    config_xpath1 = 'srq_packet_prx'

    pass_or_fail, remark = add_packets_sequence_checks(driver_setup,packet_xpath,config_xpath,config_xpath1)

    test_description = "Verify the 360khz_add_packets_sequence_checks1_ID_Config."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc9_360khz_add_packets_sequence_checks2_PT_Phase(driver_setup):
    logging.info("*************** TC-9 ***************")
    logging.info("Verify the 360khz_add_packets_sequence_checks2_PT_Phase")

    packet_xpath = 'PT_packet'
    config_xpath = 'SRQ/egph'
    config_xpath1 = 'srq_packet_egph'

    driver_setup.find_element(By.XPATH, ElementXpath['Add']).click()
    time.sleep(3)
    pass_or_fail, remark = add_packets_sequence_checks(driver_setup,packet_xpath,config_xpath,config_xpath1)

    test_description = "Verify the 360khz_add_packets_sequence_checks2_PT_Phase."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc10_360khz_add_packets_sequence_checks3_Negotiation_Phase(driver_setup):
    logging.info("*************** TC-10 ***************")
    logging.info("Verify the 360khz_add_packets_sequence_checks3_Negotiation_Phase")

    packet_xpath = 'Nego_packet'
    config_xpath = 'SRQ/egpl'
    config_xpath1 = 'srq_packet_egpl'

    driver_setup.find_element(By.XPATH, ElementXpath['Add']).click()
    time.sleep(3)
    pass_or_fail, remark = add_packets_sequence_checks(driver_setup,packet_xpath,config_xpath,config_xpath1)

    test_description = "Verify the 360khz_add_packets_sequence_checks3_Negotiation_Phase."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc11_360khz_add_packets_sequence_checks4_Ping_Phase(driver_setup):
    logging.info("*************** TC-11 ***************")
    logging.info("Verify the 360khz_add_packets_sequence_checks4_Ping_Phase")

    packet_xpath = 'ping_packet'
    config_xpath = 'analogping'
    config_xpath1 = None
    Skip = 'Yes'

    driver_setup.find_element(By.XPATH, ElementXpath['Add']).click()
    time.sleep(3)
    pass_or_fail, remark = add_packets_sequence_checks(driver_setup,packet_xpath,config_xpath,config_xpath1, Skip)

    test_description = "Verify the 360khz_add_packets_sequence_checks4_Ping_Phase."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc12_360khz_edit_packet_sequence_1(driver_setup):
    logging.info("*************** TC-12 ***************")
    logging.info("Verify the 360khz_edit_packet_sequence 1")

    # edit 1st
    time.sleep(5)
    driver_setup.find_element(By.XPATH, ElementXpath['edit1']).click()

    try:
        qi_popup1 = driver_setup.find_element(By.XPATH, ElementXpath['qi_popup1'])
        if qi_popup1:
            logging.info("Edit Pop-up is present")
            pass_or_fail, remark = editsequence(driver_setup, ent='Yes')

    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Edit Pop-up is not present")
        remark="Edit Pop-up is not present"

    test_description = "Verify the 360khz_edit_packet_sequence 1."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc13_360khz_edit_packet_sequence_2(driver_setup):
    logging.info("*************** TC-13 ***************")
    logging.info("Verify the 360khz_edit_packet_sequence 2")

    # edit 2st
    time.sleep(5)
    driver_setup.find_element(By.XPATH, ElementXpath['edit2']).click()

    try:
        qi_popup1 = driver_setup.find_element(By.XPATH, ElementXpath['qi_popup1'])
        if qi_popup1:
            logging.info("Edit Pop-up is present")
            pass_or_fail, remark = editsequence(driver_setup, ent='Yes')

    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Edit Pop-up is not present")
        remark="Edit Pop-up is not present"

    test_description = "Verify the 360khz_edit_packet_sequence 2."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc14_360khz_edit_packet_sequence_3(driver_setup):
    logging.info("*************** TC-14 ***************")
    logging.info("Verify the 360khz_edit_packet_sequence 3")

    # edit 3st
    time.sleep(5)
    driver_setup.find_element(By.XPATH, ElementXpath['edit3']).click()
    try:
        qi_popup1 = driver_setup.find_element(By.XPATH, ElementXpath['qi_popup1'])
        if qi_popup1:
            logging.info("Edit Pop-up is present")
            pass_or_fail, remark = editsequence(driver_setup, ent='Yes')

    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Edit Pop-up is not present")
        remark="Edit Pop-up is not present"

    test_description = "Verify the 360khz_edit_packet_sequence 3."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc15_360Khz_set_packet_sequence(driver_setup):
    logging.info("*************** TC-15 ***************")
    logging.info("Verify the 360khz_set_packet_sequence")

    pass_or_fail, remark = set_packet_sequence(driver_setup)

    test_description = "Verify the 360khz_set_packet_sequence."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc16_360Khz_save_and_recall_sequence(driver_setup):
    logging.info("*************** TC-16 ***************")
    logging.info("Verify the 360khz_save_sequence")
    pass_or_fail, remark = save_and_recall(driver_setup)
    test_description = "Verify the 360khz_save_sequence."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc17_360khz_start_clear(driver_setup):
    logging.info("*************** TC-17 ***************")
    logging.info("Verify the 360khz_StartExerciser_Clearcapture")
    pass_or_fail, remark = start_and_recall(driver_setup)

    test_description = "Verify the 360khz_StartExerciser_Clearcapture."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc18_360khz_instant_packets(driver_setup):
    logging.info("*************** TC-18 ***************")
    logging.info("Verify the 360khz_instant_packets")
    sendinstant = driver_setup.find_element(By.XPATH, ElementXpath['sendinstant'])
    if sendinstant:
        sendinstant.click()
        time.sleep(15)
        logging.info("Send Instant Packets are present")
        pass_or_fail, remark = editsequence(driver_setup)

    test_description = "Verify the 360khz_instant_packets."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc19_360khz_sip_tableheader(driver_setup):
    logging.info("*************** TC-19 ***************")
    logging.info("Verify the 360khz_sip_tableheader")
    pass_or_fail, remark =  table_header(driver_setup)

    test_description = "Verify the 360khz_sip_tableheader."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc20_360khz_prx_initiator(driver_setup):
    logging.info("*************** TC-20 ***************")
    logging.info("Verify the 360khz_prx_initiator")
    pass_or_fail, remark =  prx_initiator(driver_setup)

    test_description = "Verify the 360khz_prx_initiator."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc21_360khz_ptx_initiator(driver_setup):
    logging.info("*************** TC-21 ***************")
    logging.info("Verify the 360khz_ptx_initiator")
    pass_or_fail, remark =  ptx_initiator(driver_setup)

    test_description = "Verify the 360khz_ptx_initiator."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc22_360khz_add_history(driver_setup):
    logging.info("*************** TC-22 ***************")
    logging.info("Verify the 360khz_add_history")
    add_history(driver_setup)
    
def add_history(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    time.sleep(3)
    add_in_history = driver_setup.find_element(By.XPATH, ElementXpath['Ath'])
    add_in_history.click()
    time.sleep(3)
    logging.info("Prx-Configuration and Ptx-MPP-ACK is created")

def prx_initiator(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    try:
        prx_init = driver_setup.find_element(By.XPATH, ElementXpath['prx_init'])
        if prx_init:
            logging.info("PRx Initiator is present")
            prx_init_drop = driver_setup.find_element(By.XPATH, ElementXpath['prx_init_drop'])
            prx_init_drop.click()
            # Find the dropdown menu element
            dropdown_menu = driver_setup.find_element(By.CLASS_NAME, "send-instant-qi-packet-type-dropdown-menu")

            # Find all the dropdown items within the menu
            dropdown_items = dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")

            # Print the text of each dropdown item
            for dropdown_item in dropdown_items:
                logging.info(dropdown_item.text)
            time.sleep(3)
            prx_init_drop1 = driver_setup.find_element(By.XPATH, ElementXpath['prx_init_drop1'])
            prx_init_drop1.click()
    except NoSuchElementException as e:
        pass_or_fail = "Fail"
        logging.error(f"Error occurred: {e}")
        remark = str(e)

    return pass_or_fail, remark

def ptx_initiator(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    try:
        time.sleep(2)
        ptx_init = driver_setup.find_element(By.XPATH, ElementXpath['ptx_init'])
        if ptx_init:
            logging.info("PTx Initiator is present")
            ptx_init_drop = driver_setup.find_element(By.XPATH, ElementXpath['ptx_init_drop'])
            ptx_init_drop.click()
            # Find the dropdown menu element
            dropdown_menu = driver_setup.find_elements(By.CLASS_NAME, "send-instant-qi-packet-type-dropdown-menu")
            
            second_dropdown_menu = dropdown_menu[1]
            # Find all the dropdown items within the menu
            dropdown_items = second_dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
            
            # Print the text of each dropdown item
            for dropdown_item in dropdown_items:
                logging.info(dropdown_item.text)
        time.sleep(5)
        prx_init_drop1 = driver_setup.find_element(By.XPATH, ElementXpath['prx_init_drop1'])
        prx_init_drop1.click()
    except NoSuchElementException as e:
        pass_or_fail = "Fail"
        logging.error(f"Error occurred: {e}")
        remark = str(e)

    return pass_or_fail, remark

def start_and_recall(driver_setup):
    pass_or_fail = "Pass"
    remark=""
    try:
        startexe = driver_setup.find_element(By.XPATH, ElementXpath['startexe'])
        clearcapture = driver_setup.find_element(By.XPATH, ElementXpath['clearcapture'])
        if startexe and clearcapture:
            logging.info("Start Exerciser and Clear Capture both are present")

    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error(f"Can't able to find the Start Exerciser and Clear Capture")
        remark = f"Can't able to find the Start Exerciser and Clear Capture"
    return pass_or_fail, remark

def save_and_recall(driver_setup):
    pass_or_fail = "Pass"
    remark=""
    random_filename_with_extension = generate()
    try:
        save_seq = driver_setup.find_element(By.XPATH, ElementXpath['save_seq'])
        if save_seq:
            logging.info("Save Sequence is present")
            save_seq.click()
            logging.info("Save sequence is clicked.")
            filename = driver_setup.find_element(By.XPATH, ElementXpath['filename'])
            if filename:
                filewrite = driver_setup.find_element(By.XPATH, ElementXpath['filewrite'])
                filewrite.click()
                filewrite.clear()
                filewrite.send_keys(random_filename_with_extension)
                time.sleep(3)
                driver_setup.find_element(By.XPATH, ElementXpath['save_ok']).click()
                time.sleep(6)
                driver_setup.find_element(By.XPATH, ElementXpath['delete_ok']).click()
                time.sleep(2)
                driver_setup.find_element(By.XPATH, ElementXpath['delete_ok']).click()
                logging.info("Few Sequence is deleted")
                try:
                    recall = driver_setup.find_element(By.XPATH, ElementXpath['recall'])
                    if recall:
                        logging.info("Recall Sequence is present")
                        recall.click()
                        time.sleep(4)
                        Keyboard = Controller()
                        path = "C:\\Users\\GRL\\Downloads\\" + random_filename_with_extension
                        print(path)
                        Keyboard.type(path)
                        time.sleep(3)
                        Keyboard.press(Key.enter)
                        Keyboard.release(Key.enter)
                        time.sleep(7)
                        logging.info("Recall Sequence is done and all the Packet Sequence are present.")
                except NoSuchElementException:
                    pass_or_fail = "Fail"
                    logging.error("Recall Sequence not found.")
                    remark = "Recall Sequence not found."
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Save Sequence not found.")
        remark = "Save Sequence not found."

    return pass_or_fail, remark

def generate():
    import uuid

    # Generate a random UUID and truncate it to 10 characters
    random_filename = str(uuid.uuid4())[:10]

    # Append the .json extension to the filename
    random_filename_with_extension = random_filename + ".json"

    print(random_filename_with_extension)
    return random_filename_with_extension

def set_packet_sequence(driver_setup):
    pass_or_fail = "Pass"
    remark=""
    try:
        set_packet = driver_setup.find_element(By.XPATH, ElementXpath['set_packet'])
        if set_packet:
            logging.info("Set Packet Sequence is present")
            set_packet.click()
            logging.info("Set packet sequence is clicked.")
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Set Packet Sequence not found.")
        remark = "Set Packet Sequence not found."

    return pass_or_fail, remark

def removeall(driver_setup):
    pass_or_fail = "Pass"
    remark=""
    try:
        check_PTx = driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
        if check_PTx:
            logging.info("Packet Sequence are present")
            try:
                driver_setup.find_element(By.XPATH, ElementXpath['remove_all']).click()
                try:
                    check_PTx = driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
                except Exception as e:
                    logging.info("Packet Sequence is removed from the configure packet sequence tab")
            except Exception as e:
                pass_or_fail = "Fail"
                logging.error(f"Can't able to find the remove all button: {str(e)}")
                remark = f"Can't able to find the remove all button: {str(e)}"
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"Can't able to find the packet sequence: {str(e)}")
        remark = f"Can't able to find the packet sequence: {str(e)}"
    
    return pass_or_fail, remark

def reset_packetSeq(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    try:
        check_PTx = driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
    except NoSuchElementException:
        logging.info("Packet Sequence is removed from the configure packet sequence tab")
        try:
            driver_setup.find_element(By.XPATH, ElementXpath['reset_packet']).click()
            try:
                driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
                logging.info("After the Reset packet, the configure packet sequence is visible.")
                time.sleep(2)
            except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("Can't find the packet sequence after the reset.")
                remark = "Can't find the packet sequence after the reset."
        except NoSuchElementException:
            pass_or_fail = "Fail"
            logging.error("Can't find the Reset Packet button.")
            remark = "Can't find the Reset Packet button."
    return pass_or_fail, remark

def reset_exerciser(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    try:
        check_PTx = driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
    except NoSuchElementException:
        logging.info("Packet Sequence is removed from the configure packet sequence tab")
        try:
            driver_setup.find_element(By.XPATH, ElementXpath['reset_exerciser']).click()
            try:
                driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
                logging.info("After the Reset exerciser, the configure packet sequence is visible.")
                time.sleep(2)
            except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("Can't find the packet sequence after the reset.")
                remark = "Can't find the packet sequence after the reset."
        except NoSuchElementException:
            pass_or_fail = "Fail"
            logging.error("Can't find the Reset exerciser button.")
            remark = "Can't find the Reset exerciser button."
    return pass_or_fail, remark

def add_packets_details(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    driver_setup.find_element(By.XPATH, ElementXpath['Add']).click()
    time.sleep(3)
    try:
        # Add packet info
        addinfo = driver_setup.find_element(By.XPATH, ElementXpath['addpacketinfo'])
        if addinfo:
            logging.info("Add packet info is available")
            try:
                # Add packet "Configure Packet Phase"
                add_config = driver_setup.find_element(By.XPATH, ElementXpath['add_config'])
                if add_config:
                    logging.info("Configure Packet is available")
                    try:
                        # Add packet "Configure Packet Phase dropdown"
                        add_config_drop_down = driver_setup.find_element(By.XPATH, ElementXpath['add_config_dropdown'])
                        if add_config_drop_down:
                            logging.info("Configure Packet dropdown is available")
                            try:
                                # Add packet "packet phase"
                                packet_phase = driver_setup.find_element(By.XPATH, ElementXpath['packet_phase'])
                                if packet_phase:
                                    logging.info("packet_phase is available")
                                    try:
                                        # Add packet "packet phase dropdown"
                                        packet_phase_dropdown = driver_setup.find_element(By.XPATH, ElementXpath['packet_phase_dropdown'])
                                        if packet_phase_dropdown:
                                            logging.info("packet_phase dropdown is available")
                                            try:
                                                # Add to sequence button
                                                add_sequence = driver_setup.find_element(By.XPATH, ElementXpath['addtoseq'])
                                                if add_sequence:
                                                    logging.info("Add to Sequence is present in the Add info tab")
                                                    try:
                                                        # Add to set Default button
                                                        setdefault = driver_setup.find_element(By.XPATH, ElementXpath['setdefault'])
                                                        if setdefault:
                                                            logging.info("Set Default is present in the Add info tab")
                                                            try:
                                                                # Add to set Default button
                                                                cancela = driver_setup.find_element(By.XPATH, ElementXpath['cancela'])
                                                                if cancela:
                                                                    logging.info("Cancel button is present in the Add info tab")
                                                            except NoSuchElementException:
                                                                pass_or_fail = "Fail"
                                                                logging.error("Can't able to find the Cancel button")
                                                                remark = "Can't able to find the Cancel button"
                                                    except NoSuchElementException:
                                                        pass_or_fail = "Fail"
                                                        logging.error("Can't able to find the Set Default button")
                                                        remark = "Can't able to find the Set Default button"
                                            except NoSuchElementException:
                                                pass_or_fail = "Fail"
                                                logging.error("Can't able to find the Add to sequence button")
                                                remark = "Can't able to find the Add to sequence button"
                                    except NoSuchElementException:
                                        pass_or_fail = "Fail"
                                        logging.error("Can't able to find the packet_phase dropdown")
                                        remark = "Can't able to find the packet_phase dropdown"
                            except NoSuchElementException:
                                pass_or_fail = "Fail"
                                logging.error("Can't able to find the packet_phase")
                                remark = "Can't able to find packet_phase"
                    except NoSuchElementException:
                        pass_or_fail = "Fail"
                        logging.error("Can't able to find the Configure Packet dropdown")
                        remark = "Can't able to find the Configure Packet dropdown"
            except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("Can't able to find the Configure Packet info")
                remark = "Can't able to find the Configure Packet info"
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Can't able to find the Add packet info")
        remark = "Can't able to find the Add packet info"

    return pass_or_fail, remark

def add_packets_sequence(driver_setup):
    pass_or_fail = "Pass"
    remark=""
    packet_phase_dropdown = driver_setup.find_element(By.XPATH, ElementXpath['packet_phase_dropdown'])
    packet_phase_dropdown.click()
    time.sleep(2)
    try:
        dropdown_menu = driver_setup.find_element(By.CLASS_NAME, "dropdown-menu")
        dropdown_items = dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
        
        dropdown_names = [item.text for item in dropdown_items]
        logging.info(dropdown_names)
        time.sleep(3)
        packet_phase_dropdown.click()
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("Dropdown menu not found.")
        remark = "Dropdown menu not found."

    for packetdrop in dropdown_names:
        logging.info(packetdrop)
        packet_phase_dropdown.click()
        time.sleep(3)
        xpath = ElementXpath['packetseq_value'].replace('$', packetdrop)
        driver_setup.find_element(By.XPATH, xpath).click()
        time.sleep(5)
        add_config_drop_down = driver_setup.find_element(By.XPATH, ElementXpath['add_config_dropdown'])
        add_config_drop_down.click()

        dropdown_menu = driver_setup.find_element(By.CLASS_NAME, "qi-packet-type-dropdown-menu")
        dropdown_items = dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
        dropdown_values = [item.text for item in dropdown_items]
        logging.info(dropdown_values)
        add_config_drop_down.click()
    return pass_or_fail, remark

    # for packetdrop in dropdown_names:
    #     logging.info(packetdrop)
    #     packet_phase_dropdown.click()
    #     time.sleep(3)
    #     xpath = ElementXpath['packetseq_value'].replace('$', packetdrop)
    #     driver_setup.find_element(By.XPATH, xpath).click()
    #     time.sleep(5)
    #     add_config_drop_down = driver_setup.find_element(By.XPATH, ElementXpath['add_config_dropdown'])
    #     add_config_drop_down.click()

def add_packets_sequence_checks(driver_setup,packet_xpath,config_xpath,config_xpath1, Skip='No'):
    pass_or_fail = "Pass"
    remark=""
    packet_phase_dropdown = driver_setup.find_element(By.XPATH, ElementXpath['packet_phase_dropdown'])
    add_config_drop_down = driver_setup.find_element(By.XPATH, ElementXpath['add_config_dropdown'])
    # Configure Packet Sequence checks
    
    # 1. Packetpahse drop down
    packet_phase_dropdown.click()
    time.sleep(2)
    driver_setup.find_element(By.XPATH, ElementXpath[packet_xpath]).click()

    # 1. Config phase dropdown
    add_config_drop_down.click()
    time.sleep(2)
    driver_setup.find_element(By.XPATH, ElementXpath[config_xpath]).click()

    # Add Sequence
    time.sleep(2)
    driver_setup.find_element(By.XPATH, ElementXpath['addtoseq']).click()
    time.sleep(8)

    logging.info(f"Packet Phase: {packet_xpath} is selected and in Configure Packet {config_xpath} is selected")

    # PTx(Tester)
    try:
        verify = driver_setup.find_element(By.XPATH, ElementXpath['PTx'])
        if verify:
            logging.info("PTx Tester is present")
            try:
                # PRx(DUT)
                verify = driver_setup.find_element(By.XPATH, ElementXpath['PRx'])
                if verify:
                    logging.info("PRx Tester is present")
                    try:
                        verify2 = driver_setup.find_element(By.XPATH, ElementXpath['PTxrepeat'])
                        if verify2:
                            logging.info("Repeat count is display in the PTx(Tester) side.")
                            try:
                                verify3 = driver_setup.find_element(By.XPATH, ElementXpath['PTxpreamble'])
                                if verify3:
                                    logging.info("Premble count is display in the PTx(Tester) side.")
                                    try:
                                        verify5 = driver_setup.find_element(By.XPATH, ElementXpath['PTx_X'])
                                        if verify5:
                                            logging.info("Delete button is display in the PTx(Tester) side.")
                                            try:
                                                verify6 = driver_setup.find_element(By.XPATH, ElementXpath['PTx_edit'])
                                                if verify6:
                                                    logging.info("Edit button is display in the PTx(Tester) side.")
                                            except NoSuchElementException:
                                                pass_or_fail = "Fail"
                                                logging.error("Edit button is not display in the PTx(Tester) side")
                                                remark = "Edit button count is not display in the PTx(Tester) side"
                                    except NoSuchElementException:
                                        pass_or_fail = "Fail"
                                        logging.error("Delete button is not display in the PTx(Tester) side")
                                        remark = "Delete button count is not display in the PTx(Tester) side"
                            except NoSuchElementException:
                                pass_or_fail = "Fail"
                                logging.error("Premble count is not display in the PTx(Tester) side")
                                remark = "Premble count is not display in the PTx(Tester) side"
                    except NoSuchElementException:
                        pass_or_fail = "Fail"
                        logging.error("Repeat count is not display in the PTx(Tester) side")
                        remark = "Repeat count is not display in the PTx(Tester) side"   
            except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("PRx Tester is present.")
                remark = "PRx Tester is not present."
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("PTx Tester is present.")
        remark = "PTx Tester is not present."

    if Skip == 'No':
        try:
            verify1 = driver_setup.find_element(By.XPATH, ElementXpath[config_xpath1])
            if verify1:
                logging.info(f"{config_xpath1} is display in the PRx(DUT) side")
        except NoSuchElementException:
            pass_or_fail = "Fail"
            logging.error(f"{config_xpath} is not display in the PRx(DUT) side.")
            remark = f"{config_xpath} is not display in the PRx(DUT) side."
    else:
        logging.info("Analog is present")

    if Skip == 'No':
        try:
            verify4 = driver_setup.find_element(By.XPATH, ElementXpath['PTx_ACK'])
            if verify4:
                logging.info("ACK is display in the PTx(Tester) side.")
        except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("ACK is not display in the PTx(Tester) side")
                remark = "ACk count is not display in the PTx(Tester) side"
    else:
        try:
            verify4 = driver_setup.find_element(By.XPATH, ElementXpath['analog'])
            if verify4:
                logging.info("analog is display in the PTx(Tester) side.")
        except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("analog is not display in the PTx(Tester) side")
                remark = "analog count is not display in the PTx(Tester) side"

    return pass_or_fail, remark

def editsequence(driver_setup, ent = 'No'):    
    pass_or_fail = "Pass"
    remark=""
    time.sleep(2)       
    try:
        tres = driver_setup.find_element(By.XPATH, ElementXpath['tres'])
        if tres:
            logging.info("T_response is present")
            input_box = driver_setup.find_element(By.XPATH, ElementXpath['tres_input'])
            # Click on the input box to make it active
            input_box.click()
            # Generate a random number between 1 and 10
            random_number = random.randint(1, 10)
            # Enter the generated random number into the input box
            time.sleep(2)
            input_box.clear()  # Clear any existing value
            time.sleep(2)
            input_box.send_keys(str(random_number))
            logging.info(f"T_response input value is changed into, {random_number}")
            try:
                rc = driver_setup.find_element(By.XPATH, ElementXpath['rc'])
                if rc:
                    logging.info("Repeat count is prsent")
                    rc_input = driver_setup.find_element(By.XPATH, ElementXpath['rc_input'])
                    rc_input.click()
                    # Generate a random number between 1 and 10
                    random_number = random.randint(1, 10)
                    # Enter the generated random number into the input box
                    time.sleep(2)
                    rc_input.clear()  # Clear any existing value
                    time.sleep(2)
                    rc_input.send_keys(str(random_number))
                    logging.info(f"repeat count input value is changed into, {random_number}")
                    try:
                        Preamble = driver_setup.find_element(By.XPATH, ElementXpath['Preamble'])
                        if Preamble:
                            logging.info("Preamble is present")
                            preamble_input = driver_setup.find_element(By.XPATH, ElementXpath['preamble_input'])
                            preamble_input.click()
                            # Generate a random number between 1 and 10
                            random_number = random.randint(1, 10)
                            # Enter the generated random number into the input box
                            time.sleep(2)
                            preamble_input.clear()  # Clear any existing value
                            time.sleep(2)
                            preamble_input.send_keys(str(random_number))
                            logging.info(f"Preamble count input value is changed into, {random_number}")
                            
                            try:
                                #Depth
                                depth = driver_setup.find_element(By.XPATH, ElementXpath['depth'])
                                if depth:
                                    logging.info("Depth and the dropdown is present")
                                    depth_dd = driver_setup.find_element(By.XPATH, ElementXpath['depth_dd']).click()
                                    time.sleep(2)
                                    # Find all dropdown menus on the page
                                    dropdown_menus = driver_setup.find_elements(By.CLASS_NAME, "dropdown-menu")

                                    # Select the first dropdown menu
                                    first_dropdown_menu = dropdown_menus[0]

                                    # Find all the dropdown items within the first menu
                                    dropdown_items = first_dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
                                    driver_setup.find_element(By.XPATH, ElementXpath['close']).click()
                                    time.sleep(5)
                                    # Loop through each dropdown item and select them one by one
                                    for item in dropdown_items:
                                        depth_dd = driver_setup.find_element(By.XPATH, ElementXpath['depth_dd']).click()
                                        print(item)
                                        # Click on the dropdown item
                                        item.click()
                                        # Wait for 2 seconds
                                        if ent == 'No':
                                            driver_setup.find_element(By.XPATH, ElementXpath['sendinstant']).click()
                                        time.sleep(2.5)
                                    logging.info("All the Depth Values are present.")
                                    try:
                                        #ScopeTrigger
                                        ScopeTrigger = driver_setup.find_element(By.XPATH, ElementXpath['ScopeTrigger'])
                                        if ScopeTrigger:
                                            logging.info("ScopeTrigger and the dropdown is present")
                                            scopetrigger_dd = driver_setup.find_element(By.XPATH, ElementXpath['scopetrigger_dd']).click()
                                            time.sleep(2)
                                            # Find all dropdown menus on the page
                                            dropdown_menus = driver_setup.find_elements(By.CLASS_NAME, "dropdown-menu")

                                            # Select the first dropdown menu
                                            first_dropdown_menu = dropdown_menus[1]

                                            # Find all the dropdown items within the first menu
                                            dropdown_items = first_dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
                                            driver_setup.find_element(By.XPATH, ElementXpath['close']).click()
                                            time.sleep(5)
                                            # Loop through each dropdown item and select them one by one
                                            if ent == 'Yes':
                                                for item in dropdown_items:
                                                    scopetrigger_dd = driver_setup.find_element(By.XPATH, ElementXpath['scopetrigger_dd']).click()
                                                    print(item)
                                                    # Click on the dropdown item
                                                    item.click()
                                                    # Wait for 2 seconds
                                                    time.sleep(2)
                                            logging.info("All the ScopeTrigger Values are present.")
                                            try:
                                                #Polarity
                                                Polarity = driver_setup.find_element(By.XPATH, ElementXpath['Polarity'])
                                                if Polarity:
                                                    logging.info("Polarity and the dropdown is present")
                                                    Polarity_dd = driver_setup.find_element(By.XPATH, ElementXpath['Polarity_dd']).click()
                                                    time.sleep(2)
                                                    # Find all dropdown menus on the page
                                                    dropdown_menus = driver_setup.find_elements(By.CLASS_NAME, "dropdown-menu")

                                                    # Select the first dropdown menu
                                                    first_dropdown_menu = dropdown_menus[2]

                                                    # Find all the dropdown items within the first menu
                                                    dropdown_items = first_dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
                                                    driver_setup.find_element(By.XPATH, ElementXpath['close']).click()
                                                    time.sleep(5)
                                                    # Loop through each dropdown item and select them one by one
                                                    for item in dropdown_items:
                                                        Polarity_dd = driver_setup.find_element(By.XPATH, ElementXpath['Polarity_dd']).click()
                                                        print(item)
                                                        # Click on the dropdown item
                                                        item.click()
                                                        # Wait for 2 seconds
                                                        time.sleep(2)
                                                    logging.info("All the Polarity Values are present.")
                                                    try:
                                                        #FSK Cycles
                                                        fsk = driver_setup.find_element(By.XPATH, ElementXpath['fsk'])
                                                        if fsk:
                                                            logging.info("Fsk and the dropdown is present")
                                                            fsk_dd = driver_setup.find_element(By.XPATH, ElementXpath['fsk_dd']).click()
                                                            time.sleep(2)
                                                            # Find all dropdown menus on the page
                                                            dropdown_menus = driver_setup.find_elements(By.CLASS_NAME, "dropdown-menu")

                                                            # Select the first dropdown menu
                                                            first_dropdown_menu = dropdown_menus[3]

                                                            # Find all the dropdown items within the first menu
                                                            dropdown_items = first_dropdown_menu.find_elements(By.CLASS_NAME, "dropdown-item")
                                                            driver_setup.find_element(By.XPATH, ElementXpath['close']).click()
                                                            time.sleep(5)
                                                            # Loop through each dropdown item and select them one by one
                                                            for item in dropdown_items:
                                                                fsk_dd = driver_setup.find_element(By.XPATH, ElementXpath['fsk_dd']).click()
                                                                print(item)
                                                                # Click on the dropdown item
                                                                item.click()
                                                                # Wait for 2 seconds
                                                                if ent == 'No':
                                                                    driver_setup.find_element(By.XPATH, ElementXpath['sendinstant']).click()
                                                                time.sleep(2)
                                                            logging.info("All the Fsk Values are present.")
                                                            if ent == 'Yes':
                                                                try:
                                                                    Parity_Inversion = driver_setup.find_element(By.XPATH, ElementXpath['Parity_Inversion'])
                                                                    if Parity_Inversion:
                                                                        logging.info("Parity_Inversion is present")
                                                                        try:
                                                                            cancel_pop = driver_setup.find_element(By.XPATH, ElementXpath['cancel_pop'])
                                                                            if cancel_pop:
                                                                                logging.info("Cancel button is present")
                                                                                try:
                                                                                    default_pop = driver_setup.find_element(By.XPATH, ElementXpath['default_pop'])
                                                                                    if default_pop:
                                                                                        logging.info("Default button is present")
                                                                                        try:
                                                                                            Checksum = driver_setup.find_element(By.XPATH, ElementXpath['Checksum'])
                                                                                            if Checksum:
                                                                                                logging.info("Checksum Inversion is present")
                                                                                            try:
                                                                                                mask = driver_setup.find_element(By.XPATH, ElementXpath['mask'])
                                                                                                if mask:
                                                                                                    logging.info("Mask response is present")
                                                                                                    mask_input = driver_setup.find_element(By.XPATH, ElementXpath['mask_input'])
                                                                                                    mask_input.click()
                                                                                                    # Generate a random number between 1 and 10
                                                                                                    random_number = random.randint(1, 10)
                                                                                                    # Enter the generated random number into the input box
                                                                                                    time.sleep(2)
                                                                                                    mask_input.clear()  # Clear any existing value
                                                                                                    time.sleep(2)
                                                                                                    mask_input.send_keys(str(random_number))
                                                                                                    logging.info(f"Mask count input value is changed into, {random_number}")
                                                                                                    time.sleep(2)
                                                                                            except NoSuchElementException:
                                                                                                pass_or_fail = "Fail"
                                                                                                logging.error("mask Response is not present")
                                                                                                remark='Preamble count is not present'
                                                                                        except NoSuchElementException:
                                                                                            pass_or_fail = "Fail"
                                                                                            logging.error("Checksum Inversion is not present")
                                                                                            remark='Checksum Inversion is not present'
                                                                                        try:
                                                                                            SS_pop = driver_setup.find_element(By.XPATH, ElementXpath['SS_pop'])
                                                                                            if SS_pop:
                                                                                                logging.info("Save sequence button is present")
                                                                                                SS_pop.click()
                                                                                        except NoSuchElementException:
                                                                                            pass_or_fail = "Fail"
                                                                                            logging.error("Save sequence button is not present")
                                                                                            remark='Save sequence button is not present'
                                                                                except NoSuchElementException:
                                                                                    pass_or_fail = "Fail"
                                                                                    logging.error("Default button is not present")
                                                                                    remark='Default button is not present'
                                                                        except NoSuchElementException:
                                                                            pass_or_fail = "Fail"
                                                                            logging.error("Cancel button is not present")
                                                                            remark='Cancel button is not present'
                                                                except NoSuchElementException:
                                                                    pass_or_fail = "Fail"
                                                                    logging.error("Parity_Inversion is not present")
                                                                    remark='Parity_Inversion is not present'
                                                            else:
                                                                try:
                                                                    Parity_Inversion = driver_setup.find_element(By.XPATH, ElementXpath['Parity_Inversion'])
                                                                    if Parity_Inversion:
                                                                        logging.info("Parity_Inversion is present")
                                                                        try:
                                                                            Checksum = driver_setup.find_element(By.XPATH, ElementXpath['Checksum1'])
                                                                            if Checksum:
                                                                                logging.info("Checksum Inversion is present")
                                                                                try:
                                                                                    Runtime = driver_setup.find_element(By.XPATH, ElementXpath['Runtime'])
                                                                                    if Runtime:
                                                                                        logging.info("RunTime info and Dropdown is present")
                                                                                        try:
                                                                                            add_in_history = driver_setup.find_element(By.XPATH, ElementXpath['add_in_history'])
                                                                                            if add_in_history:
                                                                                                add_in_history.click()
                                                                                                logging.info("Add in History and the checkbox is present")
                                                                                                try:
                                                                                                    sipsend = driver_setup.find_element(By.XPATH, ElementXpath['sipsend'])
                                                                                                    if sipsend:
                                                                                                        logging.info("Send key is present")
                                                                                                        try:
                                                                                                            default = driver_setup.find_element(By.XPATH, ElementXpath['default'])
                                                                                                            if default:
                                                                                                                logging.info("Set default is present")
                                                                                                                try:
                                                                                                                    mask1 = driver_setup.find_element(By.XPATH, ElementXpath['mask1'])
                                                                                                                    if mask1:
                                                                                                                        logging.info("Mask response is present")
                                                                                                                        mask_input = driver_setup.find_element(By.XPATH, ElementXpath['mask_input'])
                                                                                                                        mask_input.click()
                                                                                                                        # Generate a random number between 1 and 10
                                                                                                                        random_number = random.randint(1, 10)
                                                                                                                        # Enter the generated random number into the input box
                                                                                                                        time.sleep(2)
                                                                                                                        mask_input.clear()  # Clear any existing value
                                                                                                                        time.sleep(2)
                                                                                                                        mask_input.send_keys(str(random_number))
                                                                                                                        logging.info(f"Mask count input value is changed into, {random_number}")
                                                                                                                        time.sleep(2)
                                                                                                                except NoSuchElementException:
                                                                                                                    pass_or_fail = "Fail"
                                                                                                                    logging.error("mask Response is not present")
                                                                                                                    remark='Preamble count is not present'
                                                                                                        except NoSuchElementException:
                                                                                                            pass_or_fail = "Fail"
                                                                                                            logging.error("Set default not present")
                                                                                                            remark='Set default is not present'
                                                                                                except NoSuchElementException:
                                                                                                    pass_or_fail = "Fail"
                                                                                                    logging.error("Send key is not present")
                                                                                                    remark='Send key is not present'
                                                                                        except NoSuchElementException:
                                                                                            pass_or_fail = "Fail"
                                                                                            logging.error("Add in History and the checkbox is not present")
                                                                                            remark='Add in History and the checkbox is not present'
                                                                                except NoSuchElementException:
                                                                                    pass_or_fail = "Fail"
                                                                                    logging.error("RunTime info and Dropdown is not present")
                                                                                    remark='RunTime info and Dropdown is not present'
                                                                        except NoSuchElementException:
                                                                            pass_or_fail = "Fail"
                                                                            logging.error("Checksum Inversion is not present")
                                                                            remark='Checksum Inversion is not present'
                                                                except NoSuchElementException:
                                                                    pass_or_fail = "Fail"
                                                                    logging.error("Parity_Inversion is not present")
                                                                    remark='Parity_Inversion is not present'
                                                    except NoSuchElementException:
                                                        pass_or_fail = "Fail"
                                                        logging.error("Polarity is not present")
                                                        remark='Polarity is not present'
                                            except NoSuchElementException:
                                                pass_or_fail = "Fail"
                                                logging.error("Polarity is not present")
                                                remark='Polarity is not present'
                                    except NoSuchElementException:
                                        pass_or_fail = "Fail"
                                        logging.error("Scope Trigger is not present")
                                        remark='Scope Trigger is not present'
                            except NoSuchElementException:
                                pass_or_fail = "Fail"
                                logging.error("Depth is not present")
                                remark='Depth is not present'
                    except NoSuchElementException:
                        pass_or_fail = "Fail"
                        logging.error("Preamble count is not present")
                        remark='Preamble count is not present'
            except NoSuchElementException:
                pass_or_fail = "Fail"
                logging.error("repeat count is not present")
                remark='repeat count is not present'
    except NoSuchElementException:
        pass_or_fail = "Fail"
        logging.error("t_response is not present")
        remark='t_response is not present'
    return pass_or_fail, remark

def table_header(driver_setup):
    pass_or_fail = "Pass"
    remark = ""
    table = driver_setup.find_element(By.CLASS_NAME, "qi-packet-list-table")

    # Find all the header elements
    headers = table.find_elements(By.TAG_NAME, "th")

    # List of expected header text
    expected_headers = ["S.No", "Packet Header", "Packet Payload", "Actions"]

    # Check if all expected headers are present
    for expected_header in expected_headers:
        header_found = False
        for header in headers:
            if header.text.strip() == expected_header:
                header_found = True
                break
        if not header_found:
            logging.error(f"Header '{expected_header}' not found.")
            pass_or_fail = "Fail"
            remark = f"Header '{expected_header}' not found."
        else:
            logging.info(f"Header '{expected_header}' found.")
           

    # Check if the checkbox input element is present
    checkbox_present = False
    checkbox = table.find_element(By.CSS_SELECTOR, ElementXpath['SIP_Check'])
    if checkbox:
        checkbox_present = True
        logging.info("Checkbox input element found.")
    else:
        logging.error("Checkbox input element not found.")
        pass_or_fail = "Fail"
        remark = "Checkbox input element not found."
    return pass_or_fail, remark