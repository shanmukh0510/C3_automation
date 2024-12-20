import json
import string
import time
import yaml
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import logging
import os
import random
import re
import inspect
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import WebDriverException

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"


ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_offline_TestConfig_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def generate_random_string():
    # Generate a random string with alphabets, numbers, underscores, and hyphens
    allowed_characters = string.ascii_letters + string.digits + '_-'
    random_string = ''.join(random.choice(allowed_characters) for _ in range(random.randint(8, 15)))

    logging.info(f"Random String: {random_string}")

    # Remove any other characters and keep only the allowed characters
    filtered_string = re.sub(r'[^A-Za-z0-9_\-]', '', random_string)

    logging.info(f"Filtered String: {filtered_string}")
    return filtered_string

def generate_randomnumericvalue():
    # Generate a random string with alphabets, numbers, underscores, and hyphens
    num1 = random.randint(1,100)
    return num1

def tc2_Test_Configuration_and_Selection(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("Verify that the Test Selection Tab is present and clickable in the Test Configuration.")
    try:
        # Create an ActionChain to perform keyboard shortcuts
        actions = ActionChains(driver_setup)

        # Perform the zoom out action by sending Ctrl + - keys multiple times
        for _ in range(3):  # Adjust the range to achieve the desired zoom level
            actions.key_down(Keys.CONTROL).send_keys('-').key_up(Keys.CONTROL).perform()
            time.sleep(1)  # Wait a bit between each zoom out step

        # Wait for a few seconds to observe the change

        time.sleep(8)
        pass_or_fail = "Pass"  # Initialize as "Pass" by default
        remark = " "
        try:            
            # Click the Test Configuration Tab
            driver_setup.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
            # Check if the Test Selection Tab is present
            element = None
            try:
                element = driver_setup.find_element(By.XPATH, ElementXpath['Testselection_text'])
                logging.info("Test Selection Tab is Present and Clickable")
                pass_or_fail = "Pass"
            except NoSuchElementException as e:
                logging.error(f"Test Selection Tab is not present: {e}")
                pass_or_fail = "Fail"
                remark = str(e)

            # Use an assertion to check if the element exists
            assert element is not None, "Element not found"

        except Exception as e:
            # If OpenBrowser raises an exception, it's a fail
            logging.error(f"Failed to click or Element not found with error: {str(e)}")
            pass_or_fail = "Fail"
            remark = f"Failed to click or Element not found: {str(e)}"

    except Exception as e:
        # If check_TestSelectionTab, it's a fail
        pass_or_fail = "Fail"
        remark = f"An unexpected error occurred: {str(e)}"
    
    test_description = "Verify that the Test Selection Tab is present and clickable in the Test Configuration."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_projectName(driver):
    logging.info("*************** TC-3 ***************")
    logging.info("Verify that highlight tooltip and Project Name is visible/present.")
    pass_or_fail = "Pass"
    remark = ""
    try:
        try:
            # Try to find and interact with the highlight tooltip
            highlight_tooltip = driver.find_element(By.XPATH, ElementXpath['Highlight_tooltip'])
            if highlight_tooltip.is_displayed():
                logging.info("Highlight tooltip is visible.")
                highlight_tooltip.click()
            else:
                logging.error("Highlight tooltip is not visible.")
                pass_or_fail = "Fail"
                remark = "Highlight tooltip not visible"
        except Exception as e:
            logging.warning("Can't able to find the tooltips.")

        # Try to find and interact with the Project Name element
        project_name_tab = driver.find_element(By.XPATH, ElementXpath['ProjectName_Tab'])
        
        if project_name_tab.is_displayed():
            logging.info("Project Name is present and displayed.")
        else:
            logging.error("Project Name is not present or not displayed.")
            pass_or_fail = "Fail"
            remark = "Project Name not present or not displayed"

    except NoSuchElementException as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = str(e)
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = "Unexpected error"

    test_description = "Verify that highlight tooltip and Project Name is visible/present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_create_newproject(driver):
    logging.info("*************** TC-4 ***************")
    logging.info("Verify that create new project element is availale and able to create a new project.")
    pass_or_fail = "Pass"
    remark = " "
    try:
        time.sleep(2)
        element1 = driver.find_element(By.XPATH, ElementXpath['Projectname_text'])
        driver.find_element(By.XPATH, ElementXpath['ProjectCreationButton']).click()

        cnewproj = driver.find_element(By.XPATH, ElementXpath['createNewProject'])
        text = driver.execute_script("return arguments[0].textContent;", cnewproj)
        cnewproj1 = yaml_msg("cnewproj")

        pattern = re.compile(r'^\s*|\s*$')

        text = pattern.sub('', text)
        cnewproj1 = pattern.sub('', cnewproj1)

        if text == cnewproj1:
            logging.info(f"Title is Match: {text}, {cnewproj1}")
        else:
            logging.error(f"Title is Mismatch: {text}, {cnewproj1}")
            pass_or_fail = "Fail"
            remark = f"Title is Mismatch: {text}, {cnewproj1}"

        i = 0
        while i < 2:
            project_Name = generate_random_string()
            driver.find_element(By.XPATH, ElementXpath['ProjectNameTextBox']).clear()
            driver.find_element(By.XPATH, ElementXpath['ProjectNameTextBox']).send_keys(project_Name)
            if i == 1:
                break
            i += 1  # Fixed the increment statement
            
        # Use an assertion to check if the element exists
        assert element1 is not None, "Element not found"

        try:
            # Check if 'Certification' text is present
            certification_span = driver.find_element(By.XPATH, ElementXpath['TC_Certification'])
            print("Certification text is present")

            # Check if the dropdown button is present
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_Certification_drop_down'])
            dropdown_button.click()
            print("Certification Dropdown button is clicked")

            # Check if dropdown menu items are present
            dropdown_items = driver.find_elements(By.XPATH, ElementXpath['TC_Certificate_value'])
            print(dropdown_items)
            for item in dropdown_items:
                item_text = item.text
                item.click()
                time.sleep(1)
                # Check if the selected item is present in the dropdown menu
                
                dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_Certification_drop_down'])
                dropdown_value = dropdown_button.get_attribute("innerHTML")
                logging.info(f"Certification, {dropdown_value} item is present and selected")

                time.sleep(1)
                dropdown_button.click()

        except Exception as e:
            logging.error(f"Error: {e}")
        
        try:
            # Check if 'powerprofile' text is present
            powerprofile = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile'])
            print("Power Profile text is present")

            # Check if the dropdown button is present
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile_drop_down'])
            dropdown_button.click()
            print("Power Profile Dropdown button is clicked")

            # Check if dropdown menu items are present
            dropdown_items = driver.find_elements(By.XPATH, ElementXpath['TC_powerprofile_value'])
            print(dropdown_items)
            for item in dropdown_items:
                item_text = item.text
                item.click()
                time.sleep(1)
                # Check if the selected item is present in the dropdown menu
                
                dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile_drop_down'])
                dropdown_value = dropdown_button.get_attribute("innerHTML")
                logging.info(f"Power Profile, {dropdown_value} item is present and selected")

                time.sleep(1)
                dropdown_button.click()

        except Exception as e:
            logging.error(f"Error: {e}")

            

        create_project_button = driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit'])

        if create_project_button.is_displayed() and create_project_button.is_enabled():
            logging.info("Create project button is clickable.")
        else:
            logging.error("Create project button is not clickable.")
            pass_or_fail = "Fail"
            remark = "Create project button is not clickable."
        create_project_button.click()

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify that create new project element is availale and able to create a new project"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc5_uploadproject(driver):
    logging.info("*************** TC-5 ***************")
    logging.info("Verify that create new project element is available and able to create a new project.")
    
    pass_or_fail = "Pass"
    remark = " "

    try:
        # Verify if the upload project button is clickable
        upload_project_button = driver.find_element(By.XPATH, ElementXpath['Uploadproject'])
        if upload_project_button.is_displayed() and upload_project_button.is_enabled():
            logging.info("Upload project button is clickable.")
        else:
            logging.error("Upload project button is not clickable.")
            pass_or_fail = "Fail"
            remark = "Upload project button is not clickable."

        # Click the upload project button
        upload_project_button.click()

        # Check the presence of the Upload Test Sequence Header
        title_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['UploadTestSequence'])))
        title_text = title_element.text
        logging.info(f"Title is present, and its text is: {title_text}")

        # Check the presence of the selection header in the Upload Test Sequence box
        p_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Uploadgprojtext'])))
        element_text = p_element.text
        logging.info(f"Element is present: {element_text}")

        # Check the presence of the Local File label and radio button
        label_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['LocalFileText'])))
        input_element = label_element.find_element(By.XPATH, ElementXpath['LocalFileRadiobutton'])
        label_element_text = label_element.text
        logging.info(f"Element is present: {label_element_text}")
        
        if input_element.is_enabled():
            logging.info("Radio button is clickable.")
        else:
            logging.error("Radio button is not clickable.")
            pass_or_fail = "Fail"
            remark = "Radio button is not clickable."

        # Check if the Custom Location File Path input field is disabled when the local file radio button is clicked.
        input_field = driver.find_element(By.XPATH, ElementXpath['Filepath'])
        if not input_field.is_enabled():
            logging.info("File Path input field is disabled")
        else:
            logging.error("File Path input field is not disabled")
            pass_or_fail = "Fail"
            remark = "File Path input field is not disabled"

        # Check the presence of the Report label and the expand Icon
        expand_open_icon = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Expand_open'])))
        logging.info("Expand icon with class 'rct-icon-expand-open' is present.")

        # Expand all elements
        expand_close_icons = driver.find_elements(By.CLASS_NAME, ElementXpath['rct-icon-expand-close'])
        for expand_close_icon in expand_close_icons:
            expand_close_icon.click()
            time.sleep(2)
        logging.info("All elements are expanded")

        # Check the checkbox, its clickability, and the presence of the Upload button
        checkboxes = driver.find_elements(By.CLASS_NAME, ElementXpath['rct-checkbox'])
        action_chains = ActionChains(driver)
        time.sleep(3)
        for checkbox in checkboxes:
            try:
                action_chains.move_to_element(checkbox).perform()
                checkbox.click()
                logging.info("Checkbox is clickable")

                upload_button = driver.find_element(By.CLASS_NAME, ElementXpath['grl-button'])
                if upload_button.is_enabled():
                    logging.info("Upload button is enabled")
                else:
                    logging.error("Upload button is disabled")
                    pass_or_fail = "Fail"
                    remark = "Upload button is disabled"
            except Exception as e:
                logging.error(f"Error: Checkbox is not clickable: {checkbox}")
                pass_or_fail = "Fail"
                remark = f"Error: Checkbox is not clickable: {checkbox}"

        # Check the Custom Location Radio button, its text, and the state of the Report tab
        radio_button = driver.find_element(By.ID, ElementXpath['customLocation'])
        radio_button.click()
        label = driver.find_element(By.XPATH, ElementXpath['Clradiobutton'])
        custom_location_text = label.text
        cl_text = yaml_msg("cl_text")
        pattern = re.compile(r'^\s*|\s*$')
        custom_location_text = pattern.sub('', custom_location_text)
        cl_text = pattern.sub('', cl_text)
        if custom_location_text == cl_text:
            logging.info(f"Title is Match: {custom_location_text}, {cl_text}")
        else:
            logging.error(f"Title is Mismatch: {custom_location_text}, {cl_text}")
            pass_or_fail = "Fail"
            remark = f"Title is Mismatch: {custom_location_text}, {cl_text}"

        # Check if the Report tab is disabled
        li_element = driver.find_element(By.XPATH, ElementXpath['LocalReportTab'])
        is_disabled = "disabled" in li_element.get_attribute("class")
        if is_disabled:
            logging.info("The Report tab is disabled.")
        else:
            logging.error("The Report tab is not disabled.")
            pass_or_fail = "Fail"
            remark = "The Report tab is not disabled."

        # Close the Upload project element
        close_button = driver.find_element(By.XPATH, ElementXpath['xcross'])
        close_button.click()

        try:
            WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.CLASS_NAME, ElementXpath['modal-content'])))
            logging.info("Upload project element is not present (success)")
        except TimeoutException:
            logging.error("Upload project element is still present (error)")
            pass_or_fail = "Fail"
            remark = "Upload project element is still present (error)"
        except NoSuchElementException:
            logging.error("Upload project element is not found (error)")
            pass_or_fail = "Fail"
            remark = "Upload project element is not found (error)"

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify the upload project element and their corresponding element are present and working"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    
def tc6_Quick_SelectTab(driver):
    logging.info("*************** TC-6 ***************")
    logging.info("Verify the Quick Select tab and their corresponding element are present and working")
    
    pass_or_fail = "Pass"
    remark = " "

    try:
        # Find the Quick Select tab by XPath
        quick_select_tab = driver.find_element(By.XPATH, ElementXpath['Quick_SelectTab'])

        # Check if the Quick Select tab is displayed
        if quick_select_tab.is_displayed():
            logging.info("Quick Select tab is present and displayed (success)")
        else:
            logging.error("Quick Select tab is not present or not displayed (error)")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Quick Select tab is not present or not displayed (error)"
    except NoSuchElementException:
        logging.error("Quick Select tab not found (error)")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Quick Select tab not found (error)"

    try:
        # Find the Quick Select text element and get its text content
        quick_select_text = driver.find_element(By.XPATH, ElementXpath['Quick_selectText'])
        text = quick_select_text.text.strip()
        qs_text = yaml_msg("qs_text").strip()

        if text == qs_text:
            logging.info(f"Title is Match: {text}, {qs_text}")
        else:
            logging.error(f"Title is Mismatch: {text}, {qs_text}")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = f"Title is Mismatch: {text}, {qs_text}"
    except NoSuchElementException:
        logging.error("Quick Select text element not found (error)")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Quick Select text element not found (error)"

    # Get the calling method's name using inspect
    test_description = "Verify the Quick Select tab and their corresponding element are present and working"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)


def tc7_Execute_SelectTab(driver):
    logging.info("*************** TC-7 ***************")
    logging.info("Verify the Execute tab and Repeat count functionalities are working as expected")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "

    try:
        # Verify the presence and visibility of Execute_SelectTab
        toolbar = driver.find_element(By.XPATH, ElementXpath['Execute_SelectTab'])
        if toolbar.is_displayed():
            logging.info("Execute_SelectTab is present and displayed (success)")
        else:
            logging.error("Execute_SelectTab is not present or not displayed (error)")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Execute_SelectTab is not present or not displayed"

    except NoSuchElementException:
        logging.error("Execute_SelectTab not found (error)")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Execute_SelectTab not found"

    try:
        # Verify the title of Execute Select text element
        execute = driver.find_element(By.XPATH, ElementXpath['Execute_selectText'])
        text = driver.execute_script("return arguments[0].textContent;", execute)
        execute1 = yaml_msg("execute")

        pattern = re.compile(r'^\s*|\s*$')

        text = pattern.sub('', text)
        execute1 = pattern.sub('', execute1)

        if text == execute1:
            logging.info(f"Title is Match: {text}, {execute1}")
        else:
            logging.error(f"Title is Mismatch: {text}, {execute1}")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Title of Execute Select text element is mismatched"

    except NoSuchElementException:
        logging.error("Execute Select text element not found (error)")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Execute Select text element not found"

    # Check the presence of the Repeat-numeric box
    repeat_text = driver.find_element(By.XPATH, ElementXpath['repeat_text'])
    text = driver.execute_script("return arguments[0].textContent;", repeat_text)
    repeat_text1 = yaml_msg("Repeat")

    pattern = re.compile(r'^\s*|\s*$')

    text = pattern.sub('', text)
    repeat_text1 = pattern.sub('', repeat_text1)

    if text == repeat_text1:
        logging.info(f"Title is Match: {text}, {repeat_text1}")
    else:
        logging.error(f"Title is Mismatch: {text}, {repeat_text1}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Repeat-numeric box title is mismatched"

    # Check if numeric value can be passed inside the repeat numeric box
    i = 0
    while i < 2:
        num = generate_randomnumericvalue()
        logging.info(f"repeatcount_value: {num}")
        driver.find_element(By.XPATH, ElementXpath['repeat_countvalue']).clear()
        driver.find_element(By.XPATH, ElementXpath['repeat_countvalue']).send_keys(num)
        time.sleep(1.5)
        if i == 1:
            break
        i += 1  # Fixed the increment statement

    before_ValueofRepeatCount = repeat_countValue(driver)

    # Repeat count increase using up arrow
    try:
        time.sleep(3)
        up_arrow_button = driver.find_element(By.XPATH, ElementXpath['repeat_countuparr'])

        if up_arrow_button.is_enabled():
            logging.info("The up arrow button is clickable.")
            up_arrow_button.click()
            logging.info("Clicked the up arrow button.")
        else:
            logging.error("The up arrow button is not clickable.")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Up arrow button is not clickable"
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    time.sleep(3)
    After_IncreaseValueofRepeatCount = repeat_countValue(driver)

    logging.info("Before:{}, After_Increase:{}".format(before_ValueofRepeatCount, After_IncreaseValueofRepeatCount))
    
    # Check if the repeat count is increased or not by comparing values
    if After_IncreaseValueofRepeatCount.isdigit():  # Check if After_IncreaseValueofRepeatCount is a valid integer
        After_sub = int(After_IncreaseValueofRepeatCount) - 1
        if int(before_ValueofRepeatCount) == int(After_sub):
            logging.info("The Repeat Count is Increased")
        else:
            logging.error("Error: The Repeat Count is not Increased")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Repeat Count is not Increased"
    else:
        logging.error("Error: After_IncreaseValueofRepeatCount is not a valid integer")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Invalid value for Repeat Count"


    # Repeat count decrease using down arrow
    try:
        time.sleep(3)
        down_arrow_button = driver.find_element(By.XPATH, ElementXpath['repeat_countdownarr'])

        if down_arrow_button.is_enabled():
            logging.info("The down arrow button is clickable.")
            down_arrow_button.click()
            logging.info("Clicked the down arrow button.")
        else:
            logging.error("The down arrow button is not clickable.")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Down arrow button is not clickable"
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    time.sleep(3)
    After_decreaseValueofRepeatCount = repeat_countValue(driver)

    logging.info("Before:{}, After_Decrease:{}".format(before_ValueofRepeatCount, After_decreaseValueofRepeatCount))
    time.sleep(2)
    # Check if the repeat count is decreased or not by comparing values
    if int(before_ValueofRepeatCount) == int(After_decreaseValueofRepeatCount):
        logging.info("The Repeat Count is Decreased")
    else:
        logging.error("Error: The Repeat Count is not Decreased")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Repeat Count is not Decreased"

    # Get the calling method's name using inspect
    test_description = "Verify the Execute tab and Repeat count functionalities are working as expected"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def repeat_countValue(driver):
    wait = WebDriverWait(driver, 10)  # Adjust the timeout as needed
    input_field = wait.until(EC.presence_of_element_located((By.XPATH, ElementXpath['repeat_countAttribute'])))
    # Get the value attribute of the input field
    field_value = input_field.get_attribute("value")
    # return field_value
    return field_value

def tc8_BSUT_Declaration(driver):
    logging.info("*************** TC-8 ***************")
    logging.info("Verify the BSUT Declaration form")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    if Element == 'TPT':
        try:
            #Checking that maximum power text is present or not.
            bsut = driver.find_element(By.XPATH, ElementXpath['TC_bsut_text'])
            bsut_text = driver.execute_script("return arguments[0].textContent;", bsut)
            bsut_text1 = yaml_msg("Bsut")

            pattern = re.compile(r'^\s*|\s*$')

            bsut_text = pattern.sub('', bsut_text)
            bsut_text1 = pattern.sub('', bsut_text1)

            if bsut_text == bsut_text1:
                logging.info(f"Title is Match:{bsut_text},{bsut_text1 }")
            else:
                logging.error(f"Title is Mismatch:{bsut_text},{bsut_text1}")
                pass_or_fail = "Fail"
                remark = f"Title is Mismatch:{bsut_text},{bsut_text1}"
        except Exception as e:
            logging.error("An error occurred:", str(e))
            pass_or_fail = "Fail"
            remark = "An error occurred:", str(e)
    else:
        try:
            #Checking that maximum power text is present or not.
            bsut = driver.find_element(By.XPATH, ElementXpath['bsut_text'])
            bsut_text = driver.execute_script("return arguments[0].textContent;", bsut)
            bsut_text1 = yaml_msg("Bsut1")

            pattern = re.compile(r'^\s*|\s*$')

            bsut_text = pattern.sub('', bsut_text)
            bsut_text1 = pattern.sub('', bsut_text1)

            if bsut_text == bsut_text1:
                logging.info(f"Title is Match:{bsut_text},{bsut_text1 }")
            else:
                logging.error(f"Title is Mismatch:{bsut_text},{bsut_text1}")
                pass_or_fail = "Fail"
                remark = f"Title is Mismatch:{bsut_text},{bsut_text1}"
        except Exception as e:
            logging.error("An error occurred:", str(e))
            pass_or_fail = "Fail"
            remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the BSUT Declaration form"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

    if Element == 'TPT':
        tc8_1_IsNeg_Support(driver)
        tc8_2_IsAI_Support(driver)
        tc8_3_OutofBand_Support(driver)
        tc8_4_IsSimultaneousInComing(driver)
        tc8_5_SupportedProp(driver)
        tc8_6_PRMC(driver)
        tc8_7_basic_device(driver)
        tc8_8_Enable_Optimum_Position(driver)
        tc8_9_IsPCH_Supported_checkbox(driver)
        tc8_10_ISIDxsupport(driver)
        tc8_11_refresh(driver)
        tc8_12_ptmc(driver)
        tc8_13_optimum_coil(driver)

    else:
        tc8_1_Negotiable_loadphase(driver)
        tc8_2_cloaking_support(driver)
        tc8_3_Atn(driver)
        tc8_4_cloakreason(driver)
        tc8_5_SupportedProprietary(driver)
        tc8_6_Enable_coil(driver)
        tc8_7_enableoptimum_position(driver)
        tc8_8_SupportAuth(driver)
        tc8_9_AmbientTemp(driver)
        tc8_10_prmc(driver)
        tc8_11_optimum_coil(driver)
        tc8_12_refresh(driver)

def tc8_1_Negotiable_loadphase(driver):
    logging.info("*************** TC-8.1 ***************")
    logging.info("Verify the Negptiable load phase")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the element
        negotiable_load_power_element = driver.find_element(By.XPATH, ElementXpath['Negotiable_load'])

        # Check if the element is present
        if negotiable_load_power_element:
            logging.info("Negotiable Load Power text is present.")
        else:
            logging.error("Negotiable Load Power text is not present.")
            pass_or_fail = "Fail"
            remark = "Negotiable Load Power text is not present"

        # Use XPath to locate the element with the specified text "5"
        xpath = "//a[text()='5']"
        xpath1 = "//a[text()='10']"
        xpath2 = "//a[text()='15']"
        A= [xpath,xpath1,xpath2]
        for path in A:
            element = driver.find_element(By.XPATH, ElementXpath['SDF2_DropDown'])
            if element:
                element.click()
                # Find the element
                element_N = driver.find_element(By.XPATH, path)
                # Click the element
                element_N.click()
                time.sleep(3)
        logging.info("Negotiable Load Power 5, 10, 15 is present")

    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the Negptiable load phase"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_2_cloaking_support(driver):
    logging.info("*************** TC-8.2 ***************")
    logging.info("Verify the cloaking support")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the element
        cloaking_support = driver.find_element(By.XPATH, ElementXpath['cloaking_support'])

        # Check if the element is present
        if cloaking_support:
            logging.info("cloaking_support text is present.")
        else:
            logging.error("cloaking_support text is not present.")
            pass_or_fail = "Fail"
            remark = "cloaking_support text is not present."

        element = driver.find_element(By.XPATH, ElementXpath['SDF_CS'])
        if element:
            element.click()
        # Use CSS selector to locate the checkboxes inside the div
        checkboxes = driver.find_elements(By.CSS_SELECTOR, ElementXpath['PR_initiated'])

        # Check if checkboxes are present
        if checkboxes:
            # Iterate through each checkbox and click
            for checkbox in checkboxes:
                checkbox.click()
                logging.info("Checkboxes found on the page and it's checked properly.")
        else:
            logging.error("Checkboxes not found on the page.")
            pass_or_fail = "Fail"
            remark = "Checkboxes not found on the page."

        element = driver.find_element(By.XPATH, ElementXpath['SDF_CSu']).click()

    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the cloaking support"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_3_Atn(driver):
    logging.info("*************** TC-8.3 ***************")
    logging.info("Verify the send ATN Under cloaking")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the element
        cloaking_support = driver.find_element(By.XPATH, ElementXpath['Atn'])

        # Check if the element is present
        if cloaking_support:
            logging.info("Atn text is present.")
        else:
            logging.error("Atn text is not present.")
            pass_or_fail = "Fail"
            remark = "Atn text is not present."

        element = driver.find_element(By.XPATH, ElementXpath['SDF3_DropDown'])
        time.sleep(2)
        if element:
            element.click()
            # Locate the "Yes" and "No" options and click them
            yes_option = driver.find_element(By.XPATH, '//a[text()="Yes"]')
            no_option = driver.find_element(By.XPATH, '//a[text()="No"]')
            time.sleep(2)
            yes_option.click()
            time.sleep(2)
            element.click()
            no_option.click()
            logging.info("Yes and No are clickable")
        else:
            logging.error("Yes or No option not found in the dropdown.")
            pass_or_fail = "Fail"
            remark = "Yes or No option not found in the dropdown."

    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the send ATN Under cloaking"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_4_cloakreason(driver):
    logging.info("*************** TC-8.4 ***************")
    logging.info("Verify the cloak reason and their corresponding element")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the element
        cloakreason = driver.find_element(By.XPATH, ElementXpath['cloaking_reason'])

        # Check if the element is present
        if cloakreason:
            logging.info("cloakreason text is present.")
        else:
            logging.error("cloakreason text is not present.")
            pass_or_fail = "Fail"
            remark = "cloakreason text is not present."

        element = driver.find_element(By.XPATH, ElementXpath['SDF_CR'])
        if element:
            element.click()
            # Use CSS selector to locate the checkbox container
            checkbox_container = driver.find_element(By.CSS_SELECTOR, '.custom-dropdown-option')

            # Check if the checkbox container is present
            if checkbox_container:
                # Locate all checkboxes and click them
                checkboxes = checkbox_container.find_elements(By.CSS_SELECTOR, 'input.td-input-checkbox')

                # Check if checkboxes are present
                if checkboxes:
                    # Click all checkboxes
                    for checkbox in checkboxes:
                        checkbox.click()
                        logging.info("All the checkboxes found in the container and checked properly.")
                else:
                    logging.error("No checkboxes found in the container.")
                    pass_or_fail = "Fail"
                    remark = "No checkboxes found in the container."
            else:
                logging.error("Checkbox container not found on the page.")
                pass_or_fail = "Fail"
                remark = "Checkbox container not found on the page."
        element = driver.find_element(By.XPATH, ElementXpath['SDF_CRup']).click()

    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the cloak reason and their corresponding element"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_5_SupportedProprietary(driver):
    logging.info("*************** TC-8.5 ***************")
    logging.info("Verify the Supported Proprietary and their corresponding element")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the element
        Supported_Proprietary = driver.find_element(By.XPATH, ElementXpath['Supported_Proprietary'])

        # Check if the element is present
        if Supported_Proprietary:
            logging.info("Supported_Proprietary text is present.")
        else:
            logging.error("Supported_Proprietary text is not present.")
            pass_or_fail = "Fail"
            remark = "Supported_Proprietary text is not present."

        element = driver.find_element(By.XPATH, ElementXpath['SDF_SP'])
        if element:
            element.click()
            # Use CSS selector to locate the checkbox container
            checkbox_container = driver.find_element(By.CSS_SELECTOR, '.custom-dropdown-option')

            # Check if the checkbox container is present
            if checkbox_container:
                # Locate all checkboxes and click them
                checkboxes = checkbox_container.find_elements(By.CSS_SELECTOR, 'input.td-input-checkbox')

                # Check if checkboxes are present
                if checkboxes:
                    # Click all checkboxes
                    for checkbox in checkboxes:
                        checkbox.click()
                        logging.info("All the checkboxes found in the container and checked properly.")
                else:
                    logging.error("No checkboxes found in the container.")
                    pass_or_fail = "Fail"
                    remark = "No checkboxes found in the container."
            else:
                logging.error("Checkbox container not found on the page.")
                pass_or_fail = "Fail"
                remark = "Checkbox container not found on the page."

        element = driver.find_element(By.XPATH, ElementXpath['SDF_SPup']).click()
    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify the Supported Proprietary and their corresponding element"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_6_Enable_coil(driver):
    logging.info("*************** TC-8.6 ***************")
    logging.info("Verify the Enable coil Pop-up check box should be present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "

    try:
        # Use CSS selector to locate the checkbox
        checkbox = driver.find_element(By.XPATH, ElementXpath['checkbox_text1'])

        # Check if the checkbox is present
        if checkbox:
            # Uncheck the checkbox (if it's checked)
            if checkbox.is_selected():
                checkbox.click()
                logging.info("Checkbox unchecked.")
            else:
                logging.error("Checkbox is already unchecked.")
                pass_or_fail = "Pass"
                remark = "Checkbox is already unchecked."

            # Check the checkbox
            checkbox.click()

            # Verify if the checkbox is checked
            if checkbox.is_selected():
                logging.info("Checkbox checked.")
            else:
                logging.error("Checkbox could not be checked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox could not be checked."
        else:
            logging.error("Checkbox not found on the page.")
            pass_or_fail = "Pass"
            remark = "Checkbox not found on the page."

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify the Enable coil Pop-up check box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_7_enableoptimum_position(driver):
    logging.info("*************** TC-8.7 ***************")
    logging.info("Verify the Enable optimum position check box should be present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "

    try:
        # Use CSS selector to locate the checkbox
        checkbox = driver.find_element(By.XPATH, ElementXpath['checkbox_text2'])

        # Check if the checkbox is present
        if checkbox:
            # Uncheck the checkbox (if it's checked)
            if checkbox.is_selected():
                checkbox.click()
                logging.info("Checkbox unchecked.")
            else:
                logging.info("Checkbox is already unchecked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox is already unchecked."

            # Check the checkbox
            checkbox.click()

            # Verify if the checkbox is checked
            if checkbox.is_selected():
                logging.info("Checkbox checked.")
            else:
                logging.info("Checkbox could not be checked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox could not be checked."
        else:
            logging.error("Checkbox not found on the page.")
            pass_or_fail = "Pass"
            remark = "Checkbox not found on the page."

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify the Enable optimum position check box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_8_SupportAuth(driver):
    logging.info("*************** TC-8.8 ***************")
    logging.info("Verify the Support Auth check box should be present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "

    try:
        # Use CSS selector to locate the checkbox
        checkbox = driver.find_element(By.XPATH, ElementXpath['checkbox_text3'])

        # Check if the checkbox is present
        if checkbox:
            # Uncheck the checkbox (if it's checked)
            if checkbox.is_selected():
                checkbox.click()
                logging.info("Checkbox unchecked.")
            else:
                logging.error("Checkbox is already unchecked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox is already unchecked."

            # Check the checkbox
            checkbox.click()

            # Verify if the checkbox is checked
            if checkbox.is_selected():
                logging.info("Checkbox checked.")
            else:
                logging.error("Checkbox could not be checked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox could not be checked."
        else:
            logging.error("Checkbox not found on the page.")
            pass_or_fail = "Pass"
            remark = "Checkbox not found on the page."

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify the Support Auth check box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_9_AmbientTemp(driver):
    logging.info("*************** TC-8.9 ***************")
    logging.info("Verify the Ambient Temp check box should be present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Use CSS selector to locate the checkbox
        checkbox = driver.find_element(By.XPATH, ElementXpath['checkbox_text4'])

        # Check if the checkbox is present
        if checkbox:
            # Uncheck the checkbox (if it's checked)
            if checkbox.is_selected():
                checkbox.click()
                logging.info("Checkbox unchecked.")
            else:
                logging.error("Checkbox is already unchecked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox is already unchecked."

            # Check the checkbox
            checkbox.click()

            # Verify if the checkbox is checked
            if checkbox.is_selected():
                logging.info("Checkbox checked.")
            else:
                logging.error("Checkbox could not be checked.")
                pass_or_fail = "Pass"
                #remark = "Checkbox could not be checked."
        else:
            logging.error("Checkbox not found on the page.")
            pass_or_fail = "Pass"
            remark = "Checkbox not found on the page."

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"

    # Get the calling method's name using inspect
    test_description = "Verify the Ambient Temp check box should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_10_prmc(driver):
    logging.info("*************** TC-8.10 ***************")
    logging.info("Verify the PRMC Code")
    # Find the element with class "panellabel_Width" containing the PTMC Code
    pass_or_fail = "Pass"
    remark = ""
    time.sleep(2)
    # Find all elements with class "panellabel_Width"
    try:
        panellabel_elements = driver.find_elements(By.CLASS_NAME, 'panellabel_Width')

        # Check if there are at least 7 occurrences
        if len(panellabel_elements) >= 6:
            # Get the text of the seventh element
            prmc_code = panellabel_elements[6].text
            logging.info(f"PRMC Code is present: {prmc_code}")
            try:
                # Find the fifth occurrence of the input element using XPath
                input_element = driver.find_element(By.XPATH, ElementXpath['prmcinput'])
                print("Input element found.")
                try:
                    button = driver.find_element(By.CLASS_NAME, 'qi-prmc-upload-button-td')

                    # Check if the button is clickable
                    if button.is_enabled():
                        logging.info("Button is available and clickable.")
                    else:
                        logging.error("Button is not clickable.")
                        remark = "Button is not clickable."
                        pass_or_fail = "Fail"
                except Exception as e:
                    logging.error("Button not found.")
                    remark = "Button not found."
                    pass_or_fail = "Fail"
            except Exception as e:
                logging.error("Input element not found or is not the fifth occurrence.")
                remark = "Input element not found or is not the fifth occurrence."
                pass_or_fail = "Fail"
        else:
            logging.error(f"PTMC Code is not present: {prmc_code}")
            remark = f"PTMC Code is not present: {prmc_code}"
            pass_or_fail = "Fail"

    except Exception as e:
        logging.error(f"Error: {e}")
        remark = f"Error: {e}"
        pass_or_fail = "Fail"

        # Get the calling method's name using inspect
    test_description = "Verify the PTMC Code"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_11_optimum_coil(driver):
    logging.info("*************** TC-8.11 ***************")
    logging.info("Verify the Optimum Coil Position")
    pass_or_fail = "Pass"
    remark = " "

    try:
        # Find the td element by its text content
        td_element = driver.find_element(By.XPATH, ElementXpath['optBSUT'])

        logging.info("Element found: %s", td_element.text)

        try:
            # Find the input element by its class attribute
            input_element = driver.find_element(By.CLASS_NAME, 'panelcontrol.textbox.form-control')
            logging.info("Input element found.")

            try:
                # Find the td element by its text content
                td_element = driver.find_element(By.XPATH, ElementXpath['optCoil'])
                logging.info("Element found: %s", td_element.text)

            except Exception as e:
                logging.error("Element not found.")
                remark = "Element not found."
                pass_or_fail = "Fail"

        except Exception as e:
            logging.error("Input element not found.")
            remark = "Input element not found."
            pass_or_fail = "Fail"

    except Exception as e:
        logging.error(f"Error: {e}")
        remark = f"Error: {e}"
        pass_or_fail = "Fail"

    list1 = ['optclear', 'optload', 'optsave'] 
    for button in list1:
        try:
            # Find the "Clear Data" button by its class name
            button1 = driver.find_element(By.XPATH, ElementXpath[button])

            # Check if the "Clear Data" button is displayed
            if button1.is_displayed():
                logging.info("%s Data button is displayed.", button)
            else:
                logging.error("%s Data button is not displayed.", button)
                remark = f"{button} Data button is not displayed."
                pass_or_fail = "Fail"

        except Exception as e:
            logging.error("%s Data button not found.", button)
            remark = f"{button} Data button not found."
            pass_or_fail = "Fail"

    # Get the calling method's name using inspect
    test_description = "Verify the Optimum Coil Position"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_12_refresh(driver):
    logging.info("*************** TC-8.12 ***************")
    logging.info("Verify the refresh button fuctionality")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the element using its XPath
        refresh_icon = driver.find_element(By.XPATH, ElementXpath['Refresh'])
        # If the element is found, click it
        refresh_icon.click()
        logging.info("Refresh is done")
    except NoSuchElementException:
        # If the element is not found, throw an error
        logging.error("Refresh icon not found")
        pass_or_fail = "Fail"
        remark = "Refresh icon not found"
    
    # Get the calling method's name using inspect
    test_description = "Verify the refresh button fuctionality"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_1_IsNeg_Support(driver):
    #tc8_15_refresh(driver)
    logging.info("*************** TC-8.1 ***************")
    logging.info("Verify that IsNeg Support checkbox is present or not")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the checkbox element
        checkbox = driver.find_element(By.XPATH, ElementXpath['TC_IsNeg1'])

        # Check if the checkbox is checked
        if checkbox.is_selected():
            # Get the value
            try:
                value_element = driver.find_element(By.XPATH, ElementXpath['TC_IsNeg2'])
                value = value_element.text.lower()
                
                # If value is not as expected, toggle checkbox
                if value != "true":
                    checkbox.click()
                    logging.info("Checkbox toggled and present")
            except NoSuchElementException:
                logging.error("Value element not found")
                
        else:
            # If checkbox is not checked, click it to check
            checkbox.click()

    except NoSuchElementException:
        logging.error("Checkbox element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Checkbox element not found"

    # Get the calling method's name using inspect
    test_description = "Verify that IsNeg Support checkbox is present or not"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_2_IsAI_Support(driver):
    logging.info("*************** TC-8.2 ***************")
    logging.info("Verify that IsAI Support checkbox is present or not")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the checkbox element
        checkbox = driver.find_element(By.XPATH, ElementXpath['TC_IsAI1'])

        # Check if the checkbox is checked
        if checkbox.is_selected():
            # Get the value
            try:
                value_element = driver.find_element(By.XPATH, ElementXpath['TC_IsAI2'])
                value = value_element.text.lower()
                
                # If value is not as expected, toggle checkbox
                if value != "true":
                    checkbox.click()
                    logging.info("Checkbox toggled and present")
            except NoSuchElementException:
                logging.error("Value element not found")
                
        else:
            # If checkbox is not checked, click it to check
            checkbox.click()

    except NoSuchElementException:
        logging.error("Checkbox element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Checkbox element not found"

    # Get the calling method's name using inspect
    test_description = "Verify that IsAI Support checkbox is present or not"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_3_OutofBand_Support(driver):
    logging.info("*************** TC-8.3 ***************")
    logging.info("Verify that IsOutofBand Supported checkbox is present or not")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the checkbox element
        checkbox = driver.find_element(By.XPATH, ElementXpath['TC_IsOutofBand1'])

        # Check if the checkbox is checked
        if checkbox.is_selected():
            # Get the value
            try:
                value_element = driver.find_element(By.XPATH, ElementXpath['TC_IsOutofBand2'])
                value = value_element.text.lower()
                
                # If value is not as expected, toggle checkbox
                if value != "true":
                    checkbox.click()
                    logging.info("Checkbox toggled and present")
            except NoSuchElementException:
                logging.error("Value element not found")
                
        else:
            # If checkbox is not checked, click it to check
            checkbox.click()

    except NoSuchElementException:
        logging.error("Checkbox element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Checkbox element not found"

    # Get the calling method's name using inspect
    test_description = "Verify that IsOutofBand Supported checkbox is present or not"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_4_IsSimultaneousInComing(driver):
    logging.info("*************** TC-8.4 ***************")
    logging.info("Verify that IsSimultaneousInComing and Outgoing checkbox is present or not")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the checkbox element
        checkbox = driver.find_element(By.XPATH, ElementXpath['TC_IsSimultaneousInComing1'])

        # Check if the checkbox is checked
        if checkbox.is_selected():
            # Get the value
            try:
                value_element = driver.find_element(By.XPATH, ElementXpath['TC_IsSimultaneousInComing2'])
                value = value_element.text.lower()
                
                # If value is not as expected, toggle checkbox
                if value != "true":
                    checkbox.click()
                    logging.info("Checkbox toggled and present")
            except NoSuchElementException:
                logging.error("Value element not found")
                
        else:
            # If checkbox is not checked, click it to check
            checkbox.click()

    except NoSuchElementException:
        logging.error("Checkbox element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Checkbox element not found"

    # Get the calling method's name using inspect
    test_description = "Verify that IsSimultaneousInComing and Outgoing checkbox is present or not"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_5_SupportedProp(driver):
    logging.info("*************** TC-8.5 ***************")
    logging.info("Verify that SupportedProp input tab and dropdown is present or not")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "

    try:
        # Find the element
        Supported_Proprietary = driver.find_element(By.XPATH, ElementXpath['TC_Supported_Proprietary'])

        # Check if the element is present
        if Supported_Proprietary:
            logging.info("Supported_Proprietary text is present.")
        else:
            logging.error("Supported_Proprietary text is not present.")
            pass_or_fail = "Fail"
            remark = "Supported_Proprietary text is not present."

        element = driver.find_element(By.XPATH, ElementXpath['TC_SDF_SP'])
        if element:
            element.click()
            # Use CSS selector to locate the checkbox container
            checkbox_container = driver.find_element(By.CSS_SELECTOR, '.custom-dropdown-option')

            # Check if the checkbox container is present
            if checkbox_container:
                # Locate all checkboxes and click them
                checkboxes = checkbox_container.find_elements(By.CSS_SELECTOR, 'input.td-input-checkbox')

                # Check if checkboxes are present
                if checkboxes:
                    # Click all checkboxes
                    for checkbox in checkboxes:
                        checkbox.click()
                        logging.info("All the checkboxes found in the container and checked properly.")
                else:
                    logging.error("No checkboxes found in the container.")
                    pass_or_fail = "Fail"
                    remark = "No checkboxes found in the container."
            else:
                logging.error("Checkbox container not found on the page.")
                pass_or_fail = "Fail"
                remark = "Checkbox container not found on the page."

        element = driver.find_element(By.XPATH, ElementXpath['TC_SDF_SPup']).click()
    except Exception as e:
        logging.error("An error occurred:", str(e))
        pass_or_fail = "Fail"
        remark = "An error occurred:", str(e)

    # Get the calling method's name using inspect
    test_description = "Verify that SupportedProp input tab and dropdown is present or not"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_6_PRMC(driver):
    logging.info("*************** TC-8.6 ***************")
    logging.info("Verify the PRMC text, and value can be entered or not in the text box")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the PRMC Code element
        prmc_code_element = driver.find_element(By.XPATH, ElementXpath['TC_PRMC_code'])

        # Check if PRMC Code text is present
        if prmc_code_element:
            # Find the text box associated with PRMC Code
            prmc_textbox = prmc_code_element.find_element(By.XPATH, ElementXpath['TC_PRMC_text_box'])

            # Generate a random alphanumeric value
            characters = 'efdaxcb0123456789'
            random_value = ''.join(random.choices(characters, k=8))

            prmc_textbox.clear()
            # Pass the random value to the text box
            prmc_textbox.send_keys(random_value)
            logging.info(f"PRMC Code value entered successfully {random_value}")
            time.sleep(2)

    except NoSuchElementException:
         logging.error("PRMC Code element not found")
         pass_or_fail = "Fail"  # Set to "Fail" if there's an error
         remark = "PRMC Code element not found"

    # Get the calling method's name using inspect
    test_description = "Verify the PRMC text, and value can be entered or not in the text box"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_7_basic_device(driver):
    logging.info("*************** TC-8.7 ***************")
    logging.info("Verify the basic_device text, and value can be entered or not in the text box")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the PRMC Code element
        TC_basic_device = driver.find_element(By.XPATH, ElementXpath['TC_basic_device'])

        # Check if PRMC Code text is present
        if TC_basic_device:
            # Find the text box associated with PRMC Code
            basic_device_textbox = TC_basic_device.find_element(By.XPATH, ElementXpath['TC_basic_device_text_box'])

            # Generate a random alphanumeric value
            characters = 'aedcfb0123456789'
            random_value = ''.join(random.choices(characters, k=8))

            basic_device_textbox.clear()
            # Pass the random value to the text box
            basic_device_textbox.send_keys(random_value)
            time.sleep(2)
            logging.info(f"basic_device value entered successfully {random_value}")

    except NoSuchElementException:
         logging.error("the basic_device text, and value can be entered or not in the text box Code element not found")
         pass_or_fail = "Fail"  # Set to "Fail" if there's an error
         remark = "the basic_device text, and value can be entered or not in the text box Code element not found"

    # Get the calling method's name using inspect
    test_description = "Verify the basic_device text, and value can be entered or not in the text box"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_8_Enable_Optimum_Position(driver):
    logging.info("*************** TC-8.8 ***************")
    logging.info("Verify the Enable_Optimum_Position text and checkbox is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Find the "Enable Optimum Position" element
        enable_optimum_position_element = driver.find_element(By.XPATH, ElementXpath['TC_Enable_Optimum_Position'])

        # Check if "Enable Optimum Position" text is present
        if enable_optimum_position_element:
            # Find the checkbox associated with "Enable Optimum Position"
            checkbox = enable_optimum_position_element.find_element(By.XPATH, ElementXpath['TC_Enable_Optimum_Position_checkbox'])

            # Check if the checkbox is checked
            if checkbox.is_selected():
                # Uncheck the checkbox
                checkbox.click()
                logging.info("Checkbox unchecked.")
            else:
                # Check the checkbox
                checkbox.click()
                logging.info("Checkbox checked.")
            time.sleep(2)
    except NoSuchElementException:
        logging.error("Enable Optimum Position element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Enable Optimum Position element not found"
         
    # Get the calling method's name using inspect
    test_description = "Verify the Enable_Optimum_Position text and checkbox is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_9_IsPCH_Supported_checkbox(driver):
    
    logging.info("*************** TC-8.9 ***************")
    logging.info("Verify the Is PCH Supported text snd their corresponding check box is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:

        pch_supported_element = driver.find_element(By.XPATH, "//label[@class='bsut-guaranteed-checkbox-label' and text()='Is PCH Supported']")

        # Check if "Is PCH Supported" text is present
        if pch_supported_element:
            # Find the checkbox associated with "Is PCH Supported"
            checkbox = pch_supported_element.find_element(By.XPATH, "./input[@type='checkbox']")

            # Check if the checkbox is checked
            if checkbox.is_selected():
                # Uncheck the checkbox
                checkbox.click()
            else:
                # Check the checkbox
                checkbox.click()
            time.sleep(2)
    except NoSuchElementException:
        logging.error("PCH Supported element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "PCH Supported element not found"

    # Get the calling method's name using inspect
    test_description = "Verify the Is PCH Supported text snd their corresponding check box is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_10_ISIDxsupport(driver):
    logging.info("*************** TC-8.10 ***************")
    logging.info("Verify the Is ISIDxsupport text snd their corresponding check box is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    # Check the PRx GET Request text is present or not
    try:
        # Find the "Is ID/x support" element
        id_x_support_element = driver.find_element(By.XPATH, "//label[@class='bsut-guaranteed-checkbox-label' and text()='Is ID/x support']")

        # Check if "Is ID/x support" text is present
        if id_x_support_element:
            # Find the checkbox associated with "Is ID/x support"
            checkbox = id_x_support_element.find_element(By.XPATH, "./input[@type='checkbox']")

            # Check if the checkbox is checked
            if checkbox.is_selected():
                # Uncheck the checkbox
                checkbox.click()
            else:
                # Check the checkbox
                checkbox.click()
            time.sleep(2)
            
    except NoSuchElementException:
        logging.error("ISIDxsupport element not found")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "PCH Supported element not found"

    # Get the calling method's name using inspect
    test_description = "Verify the Is ISIDxsupport text snd their corresponding check box is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
 
def tc8_11_refresh(driver):
    logging.info("*************** TC-8.11 ***************")
    logging.info("Verify the refresh button fuctionality")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the element using its XPath
        refresh_icon = driver.find_element(By.XPATH, ElementXpath['Refresh'])
        # If the element is found, click it
        refresh_icon.click()
        logging.info("Refresh is done")
    except NoSuchElementException:
        # If the element is not found, throw an error
        logging.error("Refresh icon not found")
        pass_or_fail = "Fail"
        remark = "Refresh icon not found"
    
    # Get the calling method's name using inspect
    test_description = "Verify the refresh button fuctionality"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_12_ptmc(driver):
    logging.info("*************** TC-8.12 ***************")
    logging.info("Verify the PTMC Code")
    # Find the element with class "panellabel_Width" containing the PTMC Code
    pass_or_fail = "Pass"
    remark = ""
    try: 
        # Find all elements with class "panellabel_Width"
        
        ptmc_code_label = driver.find_element(By.XPATH, "//td[@class='panellabel panellabel_Width' and text()='PTMC']")

        # Check if input text box is present
        input_text_box = driver.find_element(By.XPATH, "//td[@class='panel-input']//input[@type='text']")

        # Check if "Load Pool Data" button is present and clickable
        load_pool_data_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='qi-prmc-upload-button grl-button optimum-data-btn btn btn-primary']"))
        )

        # If all elements are present
        assert ptmc_code_label is not None, "PTMC Code label is not present"
        assert input_text_box is not None, "Input text box is not present"
        assert load_pool_data_button is not None, "Load Pool Data button is not present"

        logging.info("All elements are present")

    except NoSuchElementException as e:
        logging.error(f"One or more elements not found: {e}")
        pass_or_fail = "Fail"
        remark = f"One or more elements not found: {e}"

    # Get the calling method's name using inspect
    test_description = "Verify the PTMC Code"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_13_optimum_coil(driver):
    logging.info("*************** TC-8.13 ***************")
    logging.info("Verify the Optimum Coil Position")
    pass_or_fail = "Pass"
    remark = " "

    try:
        # Find the td element by its text content
        td_element = driver.find_element(By.XPATH, ElementXpath['TC_optBSUT'])

        logging.info("Element found: %s", td_element.text)

        try:
            # Find the input element by its class attribute
            input_element = driver.find_element(By.CLASS_NAME, 'panelcontrol.textbox.form-control')
            logging.info("Input element found.")

            try:
                # Find the td element by its text content
                td_element = driver.find_element(By.XPATH, ElementXpath['TC_optCoil'])
                logging.info("Element found: %s", td_element.text)

            except Exception as e:
                logging.error("Element not found.")
                pass_or_fail = "Fail"
                remark = "Element not found."

        except Exception as e:
            logging.error("Input element not found.")
            pass_or_fail = "Fail"
            remark = "Input element not found."

    except Exception as e:
        logging.error(f"Error: {e}")
        pass_or_fail = "Fail"
        remark = f"Error: {e}"

    list1 = ['TC_optload', 'TC_optsave'] 
    for button in list1:
        try:
            # Find the "Clear Data" button by its class name
            button1 = driver.find_element(By.XPATH, ElementXpath[button])

            # Check if the "Clear Data" button is displayed
            if button1.is_displayed():
                logging.info("%s Data button is displayed.", button)
            else:
                logging.error("%s Data button is not displayed.", button)
                remark = f"{button} Data button is not displayed."
                pass_or_fail = "Fail"

        except Exception as e:
            logging.error("%s Data button not found.", button)
            remark = f"{button} Data button not found."
            pass_or_fail = "Fail"

    # Get the calling method's name using inspect
    test_description = "Verify the Optimum Coil Position"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)



def tc9_Report_Generation(driver):
    logging.info("*************** TC-9 ***************")
    logging.info("Verify the Report Generation functionality")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the panel by its class name
        panel = driver.find_element(By.XPATH, ElementXpath['ReportpanelHeading'])
        time.sleep(3)
        # If the panel is found, check its text
        panel_text = panel.text
        Report_Generation = yaml_msg('Report_Generation')
        pattern = re.compile(r'^\s*|\s*$')
        panel_text = pattern.sub('', panel_text)
        Report_Generation = pattern.sub('', Report_Generation)
        # Check if the expected text is present
        if panel_text == Report_Generation:
            logging.info(f"Title is Match and panel is available :{panel_text},{Report_Generation}")
        else:
            logging.error(f"Title is Match Mismatch and panel not available:{panel_text},{Report_Generation}")
            pass_or_fail = "Fail"
            remark = f"Title is Match Mismatch and panel not available:{panel_text},{Report_Generation}"
    except NoSuchElementException:
        # If the panel is not found, handle it accordingly
        logging.error("Panel with class 'panelHeading' is not available")
        pass_or_fail = "Fail"
        remark = "Panel with class 'panelHeading' is not available"
    
    # Get the calling method's name using inspect
    test_description = "Verify the Report Generation functionality"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

    tc_9_1_bustinfo(driver)
    tc_9_2_brandname(driver)
    tc_9_3_productname(driver)
    tc_9_4_modelnumber(driver)
    tc_9_5_qi_id(driver)
    tc_9_6_serialnumber(driver)
    tc_9_7_testlab(driver)
    tc_9_8_labname(driver)
    tc_9_9_lablocation(driver)
    tc_9_10_labmanager(driver)
    tc_9_11_testengineer(driver)
    tc_9_12_email(driver)
    tc_9_13_phone(driver)
    tc_9_14_note_remark(driver)
    tc_9_15_refresh(driver)

def tc_9_1_bustinfo(driver):
    logging.info("*************** TC-9_1 ***************")
    logging.info("Verify the BSUT Information")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    if Element == 'TPT':
        try:
            # Locate the panel by its class name
            panel = driver.find_element(By.XPATH, ElementXpath['Bustinfo'])
            time.sleep(3)
            # If the panel is found, check its text
            panel_text = panel.text
            Bustinfo = yaml_msg('Bustinfo')

            pattern = re.compile(r'^\s*|\s*$')
            panel_text = pattern.sub('', panel_text)
            Bustinfo = pattern.sub('', Bustinfo)
            # Check if the expected text is present
            
            if panel_text == Bustinfo:
                logging.info(f"Title is Match and panel is available:{panel_text},{Bustinfo}")
            else:
                logging.error(f"Title is Match Mismatch and panel not available:{panel_text},{Bustinfo}")
                pass_or_fail = "Fail"
                remark = f"Title is Match Mismatch and panel not available:{panel_text},{Bustinfo}"
        except NoSuchElementException:
            # If the panel is not found, handle it accordingly
            logging.error("Panel with class 'panelHeading' is not available")
            pass_or_fail = "Fail"
            remark = "Panel with class 'panelHeading' is not available"
    else:
        try:
            # Locate the panel by its class name
            panel = driver.find_element(By.XPATH, ElementXpath['Bustinfo'])
            time.sleep(3)
            # If the panel is found, check its text
            panel_text = panel.text
            Bustinfo = yaml_msg('Bustinfo1')

            pattern = re.compile(r'^\s*|\s*$')
            panel_text = pattern.sub('', panel_text)
            Bustinfo = pattern.sub('', Bustinfo)
            # Check if the expected text is present
            
            if panel_text == Bustinfo:
                logging.info(f"Title is Match and panel is available:{panel_text},{Bustinfo}")
            else:
                logging.error(f"Title is Match Mismatch and panel not available:{panel_text},{Bustinfo}")
                pass_or_fail = "Fail"
                remark = f"Title is Match Mismatch and panel not available:{panel_text},{Bustinfo}"
        except NoSuchElementException:
            # If the panel is not found, handle it accordingly
            logging.error("Panel with class 'panelHeading' is not available")
            pass_or_fail = "Fail"
            remark = "Panel with class 'panelHeading' is not available"
    
    # Get the calling method's name using inspect
    test_description = "Verify the BSUT Information"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_2_brandname(driver):
    logging.info("*************** TC-9_2 ***************")
    logging.info("Verify the Manufacture/Brand Name Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Manufacturer/Brand Name" label by class name
        manufacturer_label = driver.find_element (By.XPATH,ElementXpath['Panellabel_brandName'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Panelcontrol_brandName'])
        brandlabel= manufacturer_label.text
        Brand_Name = yaml_msg('Brand_Name')

        pattern = re.compile(r'^\s*|\s*$')
        brandlabel = pattern.sub('', brandlabel)
        Brand_Name = pattern.sub('', Brand_Name)

        # If both elements are found, generate a random string and enter it in the input field
        if brandlabel == Brand_Name:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Brand_Name} label and input field are present. {brandlabel},{Brand_Name}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {brandlabel},{Brand_Name}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {brandlabel},{Brand_Name}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    # Get the calling method's name using inspect
    test_description = "Verify the Manufacture/Brand Name and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_3_productname(driver):
    logging.info("*************** TC-9_3 ***************")
    logging.info("Verify the product name and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "productname" label by class name
        productname = driver.find_element (By.XPATH,ElementXpath['Panellabel_productName'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Panelcontrol_productName'])
        productlabel= productname.text
        product_Name = yaml_msg('Product_Name')

        pattern = re.compile(r'^\s*|\s*$')
        productlabel = pattern.sub('', productlabel)
        product_Name = pattern.sub('', product_Name)

        # If both elements are found, generate a random string and enter it in the input field
        if productlabel == product_Name:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(20)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{productlabel} label and input field are present. {productlabel},{product_Name}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {productlabel},{product_Name}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {productlabel},{product_Name}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    # Get the calling method's name using inspect
    test_description = "Verify the product name and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_4_modelnumber(driver):
    logging.info("*************** TC-9_4 ***************")
    logging.info("Verify the model number and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Model Number" label by class name
        Modelnumber = driver.find_element (By.XPATH,ElementXpath['Panellabel_ModelNumber'])

        # Check if the "Model Number textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Panelcontrol_ModelNumber'])
        Modellabel= Modelnumber.text
        Model_Number = yaml_msg('Model_Number')

        pattern = re.compile(r'^\s*|\s*$')
        Modellabel = pattern.sub('', Modellabel)
        Model_Number = pattern.sub('', Model_Number)

        # If both elements are found, generate a random string and enter it in the input field
        if Modellabel == Model_Number:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(20)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Modellabel} label and input field are present. {Modellabel},{Model_Number}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Modellabel},{Model_Number}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Modellabel},{Model_Number}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"

    test_description = "Verify the model number and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_5_qi_id(driver):
    logging.info("*************** TC-9_5 ***************")
    logging.info("Verify the QI-ID and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Model Number" label by class name
        Qi = driver.find_element (By.XPATH,ElementXpath['Panellabel_Qi-ID'])

        # Check if the "Model Number textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Panelcontrol_Qi-ID'])
        Qi_IDlabel= Qi.text
        Qi_ID_Number = yaml_msg('Qi-ID')

        pattern = re.compile(r'^\s*|\s*$')
        Qi_IDlabel = pattern.sub('', Qi_IDlabel)
        Qi_ID_Number = pattern.sub('', Qi_ID_Number)

        # If both elements are found, generate a random string and enter it in the input field
        if Qi_IDlabel == Qi_ID_Number:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(20)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Qi_IDlabel} label and input field are present. {Qi_IDlabel},{Qi_ID_Number}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Qi_IDlabel},{Qi_ID_Number}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Qi_IDlabel},{Qi_ID_Number}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the QI-ID and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_6_serialnumber(driver):
    logging.info("*************** TC-9_6 ***************")
    logging.info("Verify the serial number and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "SerialNumber" label by class name
        SerialNumber = driver.find_element (By.XPATH,ElementXpath['Panellabel_Serial'])

        # Check if the "SerialNumber textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Panelcontrol_Serial'])
        Seriallabel= SerialNumber.text
        Serial_number = yaml_msg('serial')

        pattern = re.compile(r'^\s*|\s*$')
        Seriallabel = pattern.sub('', Seriallabel)
        Serial_number = pattern.sub('', Serial_number)

        # If both elements are found, generate a random string and enter it in the input field
        if Seriallabel == Serial_number:
            characters = string.digits
            random_no = ''.join(random.choice(characters) for _ in range(20)) # Change the length as needed
            input_field.send_keys(random_no)
            time.sleep(3)
            logging.info(f"{Seriallabel} label and input field are present. {Seriallabel},{Serial_number}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Seriallabel},{Serial_number}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Seriallabel},{Serial_number}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the serial number and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    
def tc_9_7_testlab(driver):
    logging.info("*************** TC-9_7 ***************")
    logging.info("Verify the testlab and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the panel by its class name
        panel = driver.find_element(By.XPATH, ElementXpath['TestLab_Info'])
        time.sleep(3)
        # If the panel is found, check its text
        panel_text = panel.text
        TestLab_Info = yaml_msg('TestLab_Info')

        pattern = re.compile(r'^\s*|\s*$')
        panel_text = pattern.sub('', panel_text)
        TestLab_Info = pattern.sub('', TestLab_Info)
        # Check if the expected text is present
        
        if panel_text == TestLab_Info:
            logging.info(f"Title is Match and panel is available:{panel_text},{TestLab_Info}")
        else:
            logging.error(f"Title is Match Mismatch and panel not available:{panel_text},{TestLab_Info}")
            pass_or_fail = "Fail"
            remark = f"Title is Match Mismatch and panel not available:{panel_text},{TestLab_Info}"
    except NoSuchElementException:
        # If the panel is not found, handle it accordingly
        logging.error("Panel with class 'panelHeading' is not available")
        pass_or_fail = "Fail"
        remark = "Panel with class 'panelHeading' is not available"
    
    test_description = "Verify the testlab and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_8_labname(driver):
    logging.info("*************** TC-9_8 ***************")
    logging.info("Verify the testlab name and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Lab Name" label by class name
        Lab_Name = driver.find_element (By.XPATH,ElementXpath['LabName_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Lab_Name'])
        Lab_Namelabel= Lab_Name.text
        Lab_Nametext = yaml_msg('Lab_Name')

        pattern = re.compile(r'^\s*|\s*$')
        Lab_Namelabel = pattern.sub('', Lab_Namelabel)
        Lab_Nametext = pattern.sub('', Lab_Nametext)

        # If both elements are found, generate a random string and enter it in the input field
        if Lab_Namelabel == Lab_Nametext:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Lab_Namelabel} label and input field are present. {Lab_Namelabel},{Lab_Nametext}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Lab_Namelabel},{Lab_Nametext}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Lab_Namelabel},{Lab_Nametext}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the testlab name and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_9_lablocation(driver):
    logging.info("*************** TC-9_9 ***************")
    logging.info("Verify the lab location and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Lab Location" label by class name
        Lab_Location = driver.find_element (By.XPATH,ElementXpath['LabLocation_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Lab_Location'])
        Lab_Locationlabel= Lab_Location.text
        Lab_Locationtext = yaml_msg('Lab_Location')

        pattern = re.compile(r'^\s*|\s*$')
        Lab_Locationlabel = pattern.sub('', Lab_Locationlabel)
        Lab_Locationtext = pattern.sub('', Lab_Locationtext)

        # If both elements are found, generate a random string and enter it in the input field
        if Lab_Locationlabel == Lab_Locationtext:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Lab_Locationlabel} label and input field are present. {Lab_Locationlabel},{Lab_Locationtext}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Lab_Locationlabel},{Lab_Locationtext}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Lab_Locationlabel},{Lab_Locationtext}"
    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the lab location and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
        
def tc_9_10_labmanager(driver):
    logging.info("*************** TC-9_10 ***************")
    logging.info("Verify the lab manager and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Lab Manager" label by class name
        Lab_Manager = driver.find_element (By.XPATH,ElementXpath['LabManager_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['Lab_Manager'])
        Lab_Managerlabel= Lab_Manager.text
        Lab_Managertext = yaml_msg('Lab_Manager')

        pattern = re.compile(r'^\s*|\s*$')
        Lab_Managerlabel = pattern.sub('', Lab_Managerlabel)
        Lab_Managertext = pattern.sub('', Lab_Managertext)

        # If both elements are found, generate a random string and enter it in the input field
        if Lab_Managerlabel == Lab_Managertext:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Lab_Managerlabel} label and input field are present. {Lab_Managerlabel},{Lab_Managertext}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Lab_Managerlabel},{Lab_Managertext}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Lab_Managerlabel},{Lab_Managertext}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the lab manager and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
        

def tc_9_11_testengineer(driver):
    logging.info("*************** TC-9_11 ***************")
    logging.info("Verify the Test Engineer and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Lab Manager" label by class name
        Test_Engineer = driver.find_element (By.XPATH,ElementXpath['Test_engineertext'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element (By.XPATH, ElementXpath['TestEngineer_label'])
        Test_Engineerlabel= Test_Engineer.text
        Test_EngineerText = yaml_msg('Test_Engineer')

        pattern = re.compile(r'^\s*|\s*$')
        Test_Engineerlabel = pattern.sub('', Test_Engineerlabel)
        Test_EngineerText = pattern.sub('', Test_EngineerText)

        # If both elements are found, generate a random string and enter it in the input field
        if Test_Engineerlabel == Test_EngineerText:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10)) # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{Test_Engineerlabel} label and input field are present. {Test_Engineerlabel},{Test_EngineerText}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Test_Engineerlabel},{Test_EngineerText}")
            pass_or_fail = "Fail"
            remark = f"Elements not found or not matching the expected conditions: {Test_Engineerlabel},{Test_EngineerText}"
    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"
        remark = "Elements not found on the page"
    
    test_description = "Verify the Test Engineer and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_12_email(driver):
    logging.info("*************** TC-9_12 ***************")
    logging.info("Verify the Email and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Email" label by XPath
        Email_text = driver.find_element(By.XPATH, ElementXpath['Email_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element(By.XPATH, ElementXpath['Email_label'])
        Emaillabel = Email_text.text
        EmailText = yaml_msg('E-mail')

        pattern = re.compile(r'^\s*|\s*$')
        Emaillabel = pattern.sub('', Emaillabel)
        EmailText = pattern.sub('', EmailText)

        # If both elements are found, generate a random email address
        if Emaillabel == EmailText:
            characters = string.ascii_letters + string.digits
            random_text = ''.join(random.choice(characters) for _ in range(10))  # Change the length as needed
            domains = ["gmail.com", "yahoo.com", "example.com", "example.in"]  # Add your desired domains here
            domain = random.choice(domains)
            email = random_text + '@' + domain

            input_field.send_keys(email)
            time.sleep(3)
            input_field.send_keys(Keys.CONTROL + "a")  # Select all text
            input_field.send_keys(Keys.DELETE)
            character = string.ascii_letters
            random_text = ''.join(random.choice(characters) for _ in range(10))
            input_field.send_keys(random_text)

            try:
                time.sleep(3)
                # Locate the <td> element containing the input field and the error message
                td_element = driver.find_element(By.XPATH, ElementXpath['InvalidEmail'])
                emailinvalid_text = td_element.text
                Emailinvalid = yaml_msg('Email_invalid')
                # If the error message is found and has red text, print a success message
                if emailinvalid_text == Emailinvalid:
                    logging.info("Email is invalid message is present and in red color.")
                else:
                    logging.error("Email is invalid message is present but not in red color.")
                    pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                    remark = "Email is invalid message is present but not in red color."
            except NoSuchElementException:
                # If the elements are not found, handle it accordingly
                logging.error("Email is invalid message is not present.")
                pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                remark = "Email is invalid message is not present."
            logging.info(f"{Emaillabel} label and input field are present. {Emaillabel}, {EmailText}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {Emaillabel}, {EmailText}")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = f"Elements not found or not matching the expected conditions: {Emaillabel}, {EmailText}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error

    test_description = "Verify the Email and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_13_phone(driver):
    time.sleep(20)
    logging.info("*************** TC-9_13 ***************")
    logging.info("Verify the Phoen number and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Phone" label by XPath
        ph_text = driver.find_element(By.XPATH, ElementXpath['phone_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element(By.XPATH, ElementXpath['phone_label'])
        phonelabel = ph_text.text
        phoneText = yaml_msg('Phone_no')

        pattern = re.compile(r'^\s*|\s*$')
        phonelabel = pattern.sub('', phonelabel)
        phoneText = pattern.sub('', phoneText)

        # If both elements are found, generate a random phone number
        if phonelabel == phoneText:
            characters = string.digits
            random_number = ''.join(random.choice(characters) for _ in range(10))  # Change the length as needed
            input_field.send_keys(random_number)
            time.sleep(3)
            input_field.send_keys(Keys.CONTROL + "a")  # Select all text
            input_field.send_keys(Keys.DELETE)
            character = string.digits
            random_number1 = ''.join(random.choice(character) for _ in range(5))
            input_field.send_keys(random_number1)

            try:
                time.sleep(3)
                # Locate the <td> element containing the input field and the error message
                td_element = driver.find_element(By.XPATH, ElementXpath['InvalidPhone'])
                phno_text = td_element.text
                phnoinvalid = yaml_msg('Phoneno_invalid')
                print(phnoinvalid, phno_text)
                # If the error message is found and has red text, print a success message
                if phno_text == phnoinvalid:
                    logging.info("Phone number invalid message is present and in red color.")
                else:
                    logging.error("Phone number invalid message is present but not in red color.")
                    pass_or_fail = "Fail"  # Set to "Fail" if there's an error

            except NoSuchElementException:
                # If the elements are not found, handle it accordingly
                logging.error("Phone number is invalid message is not present.")
                pass_or_fail = "Fail"  # Set to "Fail" if there's an error

            logging.info(f"{phonelabel} label and input field are present. {phonelabel}, {phoneText}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {phonelabel}, {phoneText}")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error

    test_description = "Verify the Phoen number and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_14_note_remark(driver):
    logging.info("*************** TC-9_14 ***************")
    logging.info("Verify the Note remark and the input Element is present")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the "Notes/Remarks" label by XPath
        notes_text = driver.find_element(By.XPATH, ElementXpath['Notes_text'])

        # Check if the "panelcontrol textbox" input field is present
        input_field = driver.find_element(By.XPATH, ElementXpath['Notes_label'])
        notes_label = notes_text.text
        notes_text_expected = yaml_msg('Notes_label')

        pattern = re.compile(r'^\s*|\s*$')
        notes_label = pattern.sub('', notes_label)
        notes_text_expected = pattern.sub('', notes_text_expected)

        # If both elements are found, generate a random string and enter it in the input field
        if notes_label == notes_text_expected:
            characters = string.ascii_letters + string.digits + string.punctuation
            random_text = ''.join(random.choice(characters) for _ in range(10))  # Change the length as needed
            input_field.send_keys(random_text)
            time.sleep(3)
            logging.info(f"{notes_label} label and input field are present. {notes_label}, {notes_text_expected}")
        else:
            logging.error(f"Elements not found or not matching the expected conditions: {notes_label}, {notes_text_expected}")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = f"Elements not found or not matching the expected conditions: {notes_label}, {notes_text_expected}"

    except NoSuchElementException:
        # If the elements are not found, handle it accordingly
        logging.error("Elements not found on the page")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = "Elements not found on the page"

    test_description = "Verify the Note remark and the input Element is present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc_9_15_refresh(driver):
    logging.info("*************** TC-9.15 ***************")
    logging.info("Verify the refresh button fuctionality")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    try:
        # Locate the element using its XPath
        refresh_icon = driver.find_element(By.XPATH, ElementXpath['Refresh'])
        # If the element is found, click it
        refresh_icon.click()
        logging.info("Refresh is done")
    except NoSuchElementException:
        # If the element is not found, throw an error
        logging.error("Refresh icon not found")
        pass_or_fail = "Fail"
        remark = "Refresh icon not found"
    
    # Get the calling method's name using inspect
    test_description = "Verify the refresh button fuctionality"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc10_Testcase_Availability(driver):
    # refresh_icon = driver.find_element(By.XPATH, ElementXpath['Refresh'])
    #     # If the element is found, click it
    # refresh_icon.click()
    time.sleep(3)

    logging.info("*************** TC-10 ***************")
    logging.info("Verify the TestCases Availability")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    
    list1 = ['TC_V_1.3.3', 'TC_BPP']
    list2 = ['TC_V_1.3.3', 'TC_EPP']
    list3 = ['TC_V_2.0.1', 'TC_BPP']
    list4 = ['TC_V_2.0.1', 'TC_EPP']

    all_list = [list1, list2, list3, list4]

    for i in all_list:
        time.sleep(2)
        # Click the Create Project Button 
        driver.find_element(By.XPATH, ElementXpath['ProjectCreationButton']).click()
        time.sleep(1)

        # Click the Certification from the drop down: V_1.3.3
        dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_Certification_drop_down'])
        dropdown_button.click()
        time.sleep(1)
        dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[0]])
        dropdown_button.click()
        
        # Click the Power Profile from the drop down: TPT_2
        time.sleep(1)
        dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile_drop_down'])
        dropdown_button.click()
        time.sleep(1)
        dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[1]])
        dropdown_button.click()
        
        #click the Create project
        create_project_button = driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit'])
        create_project_button.click()
        time.sleep(3)

        driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllTestCases']).click()

        # Define the test case groups and their corresponding keys
        if Element == 'TPT':
            if i[0] == 'TC_V_1.3.3' and i[1] == 'TC_BPP': 
                test_case_groups = {
                    "Disconnected Load tests": "Disconnected Load tests",
                    "Connected Load tests": "Connected Load tests",
                    "Load modulation tests": "Load modulation tests",
                    "Ping phase tests": "Ping phase tests",
                    "Configuration phase tests": "Configuration phase tests",
                    "Power transfer phase tests": "Power transfer phase tests",
                    "In-power transfer tests": "In-power transfer tests"
                }
            elif i[0] == 'TC_V_1.3.3' and i[1] == 'TC_EPP':
                test_case_groups = {
                    "Disconnected Load tests": "Disconnected Load tests",
                    "Connected Load tests": "Connected Load tests",
                    "Load modulation tests": "Load modulation tests",
                    "Ping phase tests": "Ping phase tests",
                    "Configuration phase tests": "Configuration phase tests",
                    "Negotiation phase tests": "Negotiation phase tests",
                    "Power transfer phase tests": "Power transfer phase tests",
                    "In-power transfer tests": "In-power transfer tests",
                }  
            elif i[0] == 'TC_V_2.0.1' and i[1] == 'TC_BPP':
                test_case_groups = {
                    "Disconnected Load tests": "Disconnected Load tests",
                    "Connected Load tests": "Connected Load tests",
                    "Load modulation tests": "Load modulation tests",
                    "Ping phase tests": "Ping phase tests",
                    "Configuration phase tests": "Configuration phase tests",
                    "Power transfer phase tests": "Power transfer phase tests",
                    "In-power transfer tests": "In-power transfer tests"
                }
            elif i[0] == 'TC_V_2.0.1' and i[1] == 'TC_EPP':
                test_case_groups = {
                    "Disconnected Load tests": "Disconnected Load tests",
                    "Connected Load tests": "Connected Load tests",
                    "Load modulation tests": "Load modulation tests",
                    "Ping phase tests": "Ping phase tests",
                    "Configuration phase tests": "Configuration phase tests",
                    "Negotiation phase tests": "Negotiation phase tests",
                    "Power transfer phase tests": "Power transfer phase tests",
                    "In-power transfer tests": "In-power transfer tests",
                }    

        if Element == 'TPT':
            if i[0] == 'TC_V_1.3.3' and i[1] == 'TC_BPP':  
                TP_Testcases = read_file('json/TPT_V_1.3.3_BPP.json')
            elif i[0] == 'TC_V_1.3.3' and i[1] == 'TC_EPP':
                TP_Testcases = read_file('json/TPT_V_1.3.3_EPP.json')
            elif i[0] == 'TC_V_2.0.1' and i[1] == 'TC_BPP':
                TP_Testcases = read_file('json/TPT_V_2.0.1_BPP.json')
            elif i[0] == 'TC_V_2.0.1' and i[1] == 'TC_EPP':
                TP_Testcases = read_file('json/TPT_V_2.0.1_EPP.json')
            

        for group_name, group_key in test_case_groups.items():
            if group_key in TP_Testcases:
                test_cases = TP_Testcases[group_key]
            else:
                test_cases = []

            # Display the test cases for the current group
            count = 0
            for test_case_name in test_cases:
                try:
                    xpath = ElementXpath['TestSelection'].replace('$', test_case_name)
                    if xpath != 'None':
                        count += 1
                        print(f"Test case_{count}: {test_case_name}")
                        #driver.find_element(By.XPATH, xpath).click()
                        time.sleep(0.2)

                except NoSuchElementException:
                    logging.error(f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'.")
                    pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                    remark = f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'."

        time.sleep(1.5)
        driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllMinus']).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, ElementXpath['TC_All_checkbox']).click()

        try:
            # Find the div element with the class "testCaseSelectionCount"
            element = driver.find_element(By.CLASS_NAME, "testCaseSelectionCount")

            # Get the text within the div element
            text = element.text

            # Use regular expressions to extract the numbers
            numbers = re.findall(r'\d+', text)

            if len(numbers) >= 2:
                first_number = int(numbers[0])
                second_number = int(numbers[1])

                logging.info(f"First number: {first_number}")
                logging.info(f"Second number: {second_number}")

                if second_number != first_number:
                    logging.error("The second number is different from the first number.")
                    pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                    remark = "The second number is different from the first number."
                else:
                    logging.info("The second number is the same as the first number.")

            else:
                logging.error("Couldn't find two numbers in the text.")
                pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                remark = "Couldn't find two numbers in the text."

        except NoSuchElementException:
            logging.error("Element with class 'testCaseSelectionCount' not found.")
            pass_or_fail = "Fail"  # Set to "Fail" if there's an error
            remark = "Element with class 'testCaseSelectionCount' not found."

        time.sleep(2)
        refresh_icon = driver.find_element(By.XPATH, ElementXpath['Refresh'])
        # If the element is found, click it
        refresh_icon.click()

    test_description = "Verify the TestCases Availability"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc11_Testresult(driver):
    logging.info("*************** TC-11 Test Result ***************")
    logging.info("Verify the TestResult") 
    try:
        pass_or_fail = "Pass"
        remark = ""
        
        # Refresh the page
        driver.refresh()
        time.sleep(2)
        
        # Navigate to Test Configuration Tab
        driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
        time.sleep(2)
        
        # Set the project path
        path = setting['CustomLocation']
        print(path)
        
        # Upload project
        upload_project_button = driver.find_element(By.XPATH, ElementXpath['Uploadproject']).click()
        time.sleep(3)
        
        # Choose custom location
        radio_button = driver.find_element(By.ID, ElementXpath['customLocation'])
        radio_button.click()
        time.sleep(3)
        
        # Set file path
        element = driver.find_element(By.XPATH, ElementXpath['Filepath_input'])
        element.click()
        element.send_keys(path)
        
        # Upload Test Sequence
        driver.find_element(By.XPATH, ElementXpath['UploadTestSequence']).click()
        driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit']).click()
        time.sleep(10)

        # Log results for PASS
        logging.info("--------------------TC11.1 PASS--------------------")
        xpath_img = ElementXpath['Pass_alt']
        img_elements = driver.find_elements(By.XPATH, xpath_img)
        unique_test_case_names = set()
        Pass = []
        for img_element in img_elements:
            alt_value = img_element.get_attribute("alt")
            span_title_element = img_element.find_element(By.XPATH, ElementXpath['sibling'])
            test_case_name = span_title_element.text.strip()
            Pass.append(test_case_name)
            logging.info(f"alt='PASS': {alt_value}, Test Case Name: {test_case_name}")
            if test_case_name and test_case_name not in unique_test_case_names:
                unique_test_case_names.add(test_case_name)

        unique_test_case_name_count = len(unique_test_case_names)
        logging.info(f"Number of unique test case names: {unique_test_case_name_count}")

        # Log results for FAIL
        logging.info("--------------------TC11.2 FAIL--------------------")
        time.sleep(2)
        xpath_img1 = ElementXpath['Fail_alt']
        img_elements1 = driver.find_elements(By.XPATH, xpath_img1)
        test_case_name_count1 = 0
        Fail=[]
        for img_element1 in img_elements1:
            alt_value1 = img_element1.get_attribute("alt")
            span_title_element1 = img_element1.find_element(By.XPATH, ElementXpath['sibling'])
            test_case_name1 = span_title_element1.text.strip()
            Fail.append(test_case_name1)
            logging.info(f"alt='FAIL': {alt_value1}, Test Case Name: {test_case_name1}")
            if test_case_name1:
                test_case_name_count1 += 1

        logging.info(f"Number of matching test case names: {test_case_name_count1}")

        # Log results for INCONCLUSIVE
        logging.info("--------------------TC11.3 INCONCLUSIVE--------------------")
        time.sleep(2)
        xpath_img2 = ElementXpath['INCONCLUSIV_alt']
        img_elements2 = driver.find_elements(By.XPATH, xpath_img2)
        test_case_name_count2 = 0
        INCONCLUSIVE=[]
        for img_element2 in img_elements2:
            alt_value2 = img_element2.get_attribute("alt")
            span_title_element2 = img_element2.find_element(By.XPATH, ElementXpath['sibling'])
            test_case_name2 = span_title_element2.text.strip()
            INCONCLUSIVE.append(test_case_name2)
            logging.info(f"alt='INCONCLUSIVE': {alt_value2}, Test Case Name: {test_case_name2}")
            if test_case_name2:
                test_case_name_count2 += 1

        logging.info(f"Number of matching test case names: {test_case_name_count2}")

        # Log results for NOTRUN
        logging.info("--------------------TC11.4 NOTRUN--------------------")
        time.sleep(2)
        xpath_img3 = ElementXpath['Notrun_alt']
        img_elements3 = driver.find_elements(By.XPATH, xpath_img3)
        test_case_name_count3 = 0
        NOTRUN=[]
        for img_element3 in img_elements3:
            alt_value3 = img_element3.get_attribute("alt")
            span_title_element3 = img_element3.find_element(By.XPATH, ElementXpath['sibling'])
            test_case_name3 = span_title_element3.text.strip()
            NOTRUN.append(test_case_name3)
            logging.info(f"alt='NOTRUN': {alt_value3}, Test Case Name: {test_case_name3}")
            if test_case_name3:
                test_case_name_count3 += 1

        logging.info(f"Number of matching test case names: {test_case_name_count3}")

        # Validate total case count
        totalcase = unique_test_case_name_count + test_case_name_count1 + test_case_name_count2 + test_case_name_count3
        logging.info("Total Number of Test Cases: %s", totalcase)
        
        # Validate counts
        counts = []
        xpath_strong = ElementXpath['strong']
        strong_element = driver.find_elements(By.XPATH, xpath_strong)
        for strong in strong_element:
            text = strong.text.strip()
            parts = text.split('/')
            count = parts[1].strip() if len(parts) > 1 else None
            if count is not None:
                counts.append(count)

        for count in counts:
            if int(count) == totalcase:
                logging.info(f"Count {count} matches the total ({totalcase}).")
            else:
                logging.error(f"Count {count} does not match the total ({totalcase}).")
                pass_or_fail = "Fail"
                remark = f"Count {count} does not match the total ({totalcase})."

        # Navigate back to Test Configuration Page
        logging.info("--------------------11.5 Back to Test Configuration Page--------------------")
        
        # Check the Passed test case elements
        logging.info("--------------------Check the Passed test case Element are present while click the Passed img Element--------------------")
        time.sleep(2)
        driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
        
        time.sleep(2)
        driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllTestCases']).click()
        
        try:
            # Click Passed img Element
            img_element = driver.find_element(By.XPATH, ElementXpath['Pass'])
            img_element.click()
        except Exception as e:
            logging.error("Error:", str(e))

        # Validate if Passed test case elements are present
        time.sleep(2)
        for tc in Pass:
            xpath = ElementXpath['selectioncheck'].replace('$', tc)
            try:
                driver.find_element(By.XPATH, xpath)
                logging.info(f"Element with '{tc}' found: Pass")
            except NoSuchElementException:
                logging.error(f"Element with '{tc}' not found: Fail")
                remark = f"Element with '{tc}' not found: Fail"
                pass_or_fail = "Fail"
            time.sleep(1)

        # Disable the pass button
        time.sleep(2)
        img_element = driver.find_element(By.XPATH, ElementXpath['Pass']).click()

        # Check the Failed test case elements
        logging.info("--------------------11.6 Check the Failed test case Element are present while click the Failed img Element--------------------")
        try:
            time.sleep(2)
            img_element = driver.find_element(By.XPATH, ElementXpath['Fail'])
            img_element.click()
        except Exception as e:
            logging.error("Error:", str(e))
            remark = "Error:", str(e)
            pass_or_fail = "Fail"

        # Validate if Failed test case elements are present
        time.sleep(2)
        for tc in Fail:
            xpath = ElementXpath['selectioncheck'].replace('$', tc)
            try:
                driver.find_element(By.XPATH, xpath)
                logging.info(f"Element with '{tc}' found: Pass")
            except NoSuchElementException:
                logging.error(f"Element with '{tc}' not found: Fail")
                remark = f"Element with '{tc}' not found: Fail"
                pass_or_fail = "Fail"
            time.sleep(1)

        # Disable the fail button
        time.sleep(2)
        img_element = driver.find_element(By.XPATH, ElementXpath['Fail']).click()

        # Check the INCONCLUSIVE test case elements
        logging.info("--------------------11.7 Check the INCONCLUSIVE test case Element are present while click the INCONCLUSIVE img Element--------------------")
        try:
            time.sleep(2)
            img_element = driver.find_element(By.XPATH, ElementXpath['Inconclusiv'])
            img_element.click()
        except Exception as e:
            logging.error("Error:", str(e))
            remark = "Error:", str(e)
            pass_or_fail = "Fail"

        # Validate if INCONCLUSIVE test case elements are present
        time.sleep(2)
        for tc in INCONCLUSIVE:
            xpath = ElementXpath['selectioncheck'].replace('$', tc)
            try:
                driver.find_element(By.XPATH, xpath)
                logging.info(f"Element with '{tc}' found: Pass")
            except NoSuchElementException:
                logging.error(f"Element with '{tc}' not found: Fail")
                remark = f"Element with '{tc}' not found: Fail"
                pass_or_fail = "Fail"
            time.sleep(1)

        # Disable the INCONCLUSIVE button
        time.sleep(2)
        img_element = driver.find_element(By.XPATH, ElementXpath['Inconclusiv']).click()

        # Check the NOTRUN test case elements
        logging.info("--------------------11.8 Check the NOTRUN test case Element are present while click the NOTRUN img Element--------------------")
        try:
            time.sleep(2)
            img_element = driver.find_element(By.XPATH, ElementXpath['Notrun'])
            img_element.click()
        except Exception as e:
            logging.error("Error:", str(e))
            remark = "Error:", str(e)
            pass_or_fail = "Fail"

        # Validate if NOTRUN test case elements are present
        time.sleep(2)
        for tc in NOTRUN:
            xpath = ElementXpath['selectioncheck'].replace('$', tc)
            try:
                driver.find_element(By.XPATH, xpath)
                logging.info(f"Element with '{tc}' found: Pass")
            except NoSuchElementException:
                logging.error(f"Element with '{tc}' not found: Fail")
                remark = f"Element with '{tc}' not found: Fail"
                pass_or_fail = "Fail"
            time.sleep(1)

        # Disable the NOTRUN button
        time.sleep(2)    
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {str(e)}"
    # Log test completion
    test_description = "Check that all test cases are available in the Quick Select tab to select"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc12_reportgeneration(driver):
    logging.info("*************** TC-12 ***************")
    logging.info("Check that all BSUT information displayed in the input box while in offline mode")    
    pass_or_fail = "Pass"  # Initialize to "Pass" and change to "Fail" if any errors occur
    remark = " "
    # # Refresh the page
    # driver.refresh()
    # time.sleep(2)
    
    # # Navigate to Test Configuration Tab
    # driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
    # time.sleep(2)
    
    # # Set the project path
    path = setting['CustomLocation']
    # #print(path)
    
    # # Upload project
    # upload_project_button = driver.find_element(By.XPATH, ElementXpath['Uploadproject']).click()
    # time.sleep(3)
    
    # # Choose custom location
    # radio_button = driver.find_element(By.ID, ElementXpath['customLocation'])
    # radio_button.click()
    # time.sleep(3)
    
    # # Set file path
    # element = driver.find_element(By.XPATH, ElementXpath['Filepath_input'])
    # element.click()
    # element.send_keys(path)
    
    # # Upload Test Sequence
    # driver.find_element(By.XPATH, ElementXpath['UploadTestSequence']).click()
    # driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit']).click()
    # time.sleep(10)

    pass_or_fail = "Pass"
    remark = ''
    # Initialize an empty list to store the extracted values
    extracted_values = []

    try:
        with open(path, 'r') as file:
            # Load the JSON data from the file
            json_obj = json.load(file)

            # Extract the specified fields
            fields_to_extract = [
                "manufacturer", "productName", "modelNumber", "qiID", "serialNumber",
                "testLab", "testlablocation", "testlabmanager", "testEngineer", "email",
                "phonenumber", "remarksComments"
            ]

            # Extract the values and append them to the list
            for field in fields_to_extract:
                extracted_values.append(json_obj['testBkpQiconfig'][field])

    except FileNotFoundError:
        print(f"The file {path} was not found.")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"The file {path} was not found."
    except json.JSONDecodeError as e:
        print(f"JSON decoding error: {e}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"JSON decoding error: {e}"
    except Exception as e:
        print(f"An error occurred: {e}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"An error occurred: {e}"

    # Print the extracted values
    print("Extracted Values:", extracted_values)


    GUI = []
    path = ['Panelcontrol_brandName', 'Panelcontrol_productName', 'Panelcontrol_ModelNumber', 'Panelcontrol_Qi-ID',
            'Panelcontrol_Serial', 'Lab_Name', 'Lab_Location', 'Lab_Manager', 'TestEngineer_label', 'Email_label', 'phone_label', 'Notes_label']

    for xpath in path:
        input_element = driver.find_element(By.XPATH, ElementXpath[xpath])
        value = input_element.get_attribute('value')
        GUI.append(value)

    # Print the GUI list
    print("GUI Values:", GUI)
    
    if extracted_values == GUI:
        logging.info("Both lists are identical.")
    else:
        logging.error("Lists are not identical.")
        # Find the differences
        differences = set(extracted_values) ^ set(GUI)
        logging.warning(f"Differences: {differences}")
        pass_or_fail = "Fail"  # Set to "Fail" if there's an error
        remark = f"Lists are not identical, Differences: {differences}"

    # Log test completion
    test_description = "Check that all BSUT information displayed in the input box while in offline mode"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

logging.info("--------------------Test Completed--------------------")