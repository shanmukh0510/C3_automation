import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
import logging
import os
import inspect
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill
from dateutil import parser
import re
import random
import string
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
APIs = Xpath['API'][setting['Mode']]
Element = setting['Mode']

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_ReportAnalyser_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_Report_AnalyserTab(driver_setup):
    logging.info("*************** TC-2 ***************")
    logging.info("Verify that Report Analyser Tab should be present and landed.")

    pass_or_fail = "Pass"
    remark = ""
    time.sleep(5)
    try:
        report_analyser_tab = driver_setup.find_element(By.XPATH, ElementXpath['Report_analyser'])
        time.sleep
        if report_analyser_tab.is_displayed():
            report_analyser_tab.click()
            logging.info("Report Analyser Tab is visible and clickable")
        else:
            logging.error("Report Analyser Tab is not visible")
            pass_or_fail = "Fail"
            remark = "Report Analyser Tab is not visible"

        report_analyser_text = driver_setup.find_element(By.XPATH, ElementXpath['Report_analyser_text'])
        if report_analyser_text.is_displayed():
            logging.info("Report Analyser Tab Page is loaded and landed")
        else:
            logging.error("Can't able to land on the Report Analyser Tab")
            pass_or_fail = "Fail"
            remark = "Report Analyser Tab Page is not loaded"

    except NoSuchElementException:
        logging.error("Report Analyser Tab not found.")
        pass_or_fail = "Fail"
        remark = "Report Analyser Tab not found."

    test_description = "Verify that Report Analyser Tab should be present and landed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_RA_uploadjson(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("Verify JSON was uploaded successfully and check their corresponding element")
    pass_or_fail = "Pass"
    remark = ""
    tpt_value = ""
    try:
        time.sleep(2)
        ra_uploadjson = driver_setup.find_element(By.XPATH, ElementXpath['RA_uploadbutton'])
        if ra_uploadjson.is_displayed():
            time.sleep(2)
            ra_uploadjson.click()
            logging.info("Json upload Button is visible and clicked")
            time.sleep(5)
        else:
            logging.error("Json upload Button is not visible.")
            remark = "Json upload Button is not visible."
            pass_or_fail = "Fail"

        RA_uploadtext = driver_setup.find_element(By.XPATH, ElementXpath['RA_uploadtext'])
        # Check if the element is present (found) and visible
        if RA_uploadtext.is_displayed():
            logging.info("Text 'Upload JSON' is present and visible.")
        else:
            logging.error("Text 'Upload JSON' is not visible.")
            remark = "Text 'Upload JSON' is not visible."
            pass_or_fail = "Fail"

        RA_Localtext = driver_setup.find_element(By.XPATH, ElementXpath['RA_Localtext'])
        # Check if the element is present (found) and visible
        if RA_Localtext.is_displayed():
            logging.info("Text 'Local File' is present and visible.")
        else:
            logging.error("Text 'Local File' is not visible.")
            remark = "Text 'Local File' is not visible."
            pass_or_fail = "Fail"

        RA_Report = driver_setup.find_element(By.XPATH, ElementXpath['RA_Report'])
        # Check if the element is present (found) and visible
        if RA_Report.is_displayed():
            logging.info("Text 'Report' is present and visible.")
        else:
            logging.error("Text 'Report' is not visible.")
            remark = "Text 'Report' is not visible."
            pass_or_fail = "Fail"

        # Extracting the TPT value from the CustomLocation using string manipulation
        custom_location_parts = setting['CustomLocation'].split('/Report/')
        print("custom_location_parts",custom_location_parts)
        print(len(custom_location_parts))
        if len(custom_location_parts) >= 1:
            tpt_value = custom_location_parts[1].split('/')[0]
            logging.info(f"TPT/TPR Value is selected: {tpt_value}")
        else:
            logging.error("TPT/TPR Value not found in CustomLocation.")
            remark = "TPT/TPR Value not found in CustomLocation."
            pass_or_fail = "Fail"

        xpath = ElementXpath['RA_Jsonselect'].replace('$', tpt_value)
        RA_Jsonselect = driver_setup.find_element(By.XPATH, xpath)
        RA_Jsonselect.click()
        
        RA_Drag_drop_text = driver_setup.find_element(By.XPATH, ElementXpath['RA_Drag_drop_text'])
        if RA_Drag_drop_text.is_displayed():
            logging.info("Text 'Drag_drop' is present and visible.")
        else:
            logging.error("Text 'Drag_drop' is not visible.")
            pass_or_fail = "Fail"
            remark = "Text 'Drag_drop' is not visible."

        RA_Drag_drop_box = driver_setup.find_element(By.XPATH, ElementXpath['RA_Drag_drop_box'])
        if RA_Drag_drop_box.is_displayed():
            logging.info("Element 'Drag_drop_box' is present and visible.")
        else:
            logging.error("Element 'Drag_drop_box' is not visible.")
            pass_or_fail = "Fail"
            remark = "Element 'Drag_drop_box' is not visible."

        RA_upload = driver_setup.find_element(By.XPATH, ElementXpath['RA_upload'])
        if RA_upload.is_displayed():
            logging.info("Element 'upload button' is present and visible.")
            RA_upload.click()
            logging.info("Json File is uploaded successfully")
        else:
            logging.error("Element 'upload button' is not visible.")
            logging.error("Can't able to loaded the Json file")
            pass_or_fail = "Fail"
            remark = "Element 'upload button' is not visible /Can't able to loaded the Json file"
        time.sleep(5)

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Verify JSON was uploaded successfully and check their corresponding element"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_RA_table_header(driver_setup):
    logging.info("*************** TC-4 ***************")
    logging.info("Verify that all the table headers should be present") 
    pass_or_fail = "Pass"
    remark = ""
    try:
        
        div_elements = driver_setup.find_elements(By.XPATH, ElementXpath['RA_table_header'])
        # Extract text values from each div element
        text_values = [div.text for div in div_elements]
        ra_table_hd = yaml_msg('RA_table_header')

        for label_text in ra_table_hd:
            if label_text in text_values:
                logging.info(f"{label_text} is present.")
            else:
                logging.error(f"{label_text} is not present.")
                pass_or_fail = "Fail"  # Set to "Fail" if there's an error
                remark = f"{label_text} is not present."

    except Exception as e:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = f"An error occurred: {str(e)}"

    test_description = "Verify that all the table headers should be present"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc5_Test_lab(driver_setup):
    logging.info("*************** TC-5 ***************")
    logging.info("Please verify the Test Lab table and ensure that all elements are present.")
    pass_or_fail = "Pass"
    remark = " "
    try:
        #logging.info("_______________GUI DATA________________")
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_tldropdown'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test lab is expanded")
        else:
            logging.error("Test lab is not expanded")
            pass_or_fail = "Fail"
            remark = 'Test lab is not expanded'

        A = []
        B = []

        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_TB_1'])
        for j in div_elements1:
            k = j.text
            A.append(k if k is not None and k.strip() != '' else '-')
        A = [value for value in A if value != '-']

        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['2'])
        for l in div_elements2:  
            m = l.text
            B.append(m if m is not None and m.strip() != '' else '-')

        result_dict = dict(zip(A, B))
        After_rm_tl_text = {key.replace(' ', ''): value for key, value in result_dict.items()}
        #logging.info(f"GUI header value: {After_rm_tl_text}")

        #logging.info("_______________JSON DATA_______________")
        json1 = setting['CustomLocation']
        directory_path = os.path.dirname(json1)
        files = os.listdir(directory_path)

        # Iterate through the files and find the first '.json' file
        json_file_path = None
        for file in files:
            if file.lower().endswith('.json'):
                json_file_path = os.path.join(directory_path, file)
                break  # Stop when the first JSON file is found

        # Print the full path to the first JSON file (or None if none is found)
        #logging.info(json_file_path)
        with open(json_file_path, "r") as file:
            data = json.load(file)

        json_tl_text = {key: value if value.strip() != "" else "-" for key, value in data["TestLab"].items()}
        #logging.info(f"Json header value: {json_tl_text}")

        #logging.info("_______________Compare GUI & Json Data_______________")

        if After_rm_tl_text == json_tl_text:
            logging.info("Both GUI and Json values are similar as expected")
        else:
            logging.error("GUI and Json Value are not similar.")
            pass_or_fail = "Fail"
            remark = "GUI and Json Value are not similar."

        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_tlup'])
        if div_elements2:
            for ih in div_elements2:
                ih.click()
                logging.info("Test lab is minimized")
        else:
            logging.error("Test lab is not minimized")
            pass_or_fail = "Fail"
            remark = "Test lab is not minimized"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    # Get the calling method's name using inspect
    test_description = "Please verify the Test Lab table and ensure that all elements are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc6_Test_Execution(driver_setup):
    logging.info("*************** TC-6 ***************")
    logging.info("Please verify the Test_Execution table and ensure that all elements are present")
    pass_or_fail = "Pass"
    remark = " "
    try:
        #logging.info("_______________GUI DATA________________")
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_tedropdown'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_Execution is expanded")
        else:
            logging.error("Test_Execution is not expanded")
            pass_or_fail = "Fail"
            remark = 'Test_Execution is not expanded'

        A = []
        B = []
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_TB_1'])
        for j in div_elements1:
            k = j.text
            A.append(k if k is not None and k.strip() != '' else '-')
        A = [value for value in A if value != '-']

        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['2'])
        for l in div_elements2:  
            m = l.text
            B.append(m if m is not None and m.strip() != '' else '-')

        data_dict_GUI = {
        A[0]: B[0],
        A[1]: B[1],
        A[2]: [test for test in B[2:] if test.startswith('TD')],
        A[3]: B[-3],
        A[4]: B[-2],
        A[5]: B[-1]
        }

        # Print the resulting dictionary
        data = {key.replace(' ', ''): value for key, value in data_dict_GUI.items()}
        
        pass_count = int(data['TestResult'].split('(')[1].split(')')[0])
        fail_count = int(data['TestResult'].split('(')[2].split(')')[0])
        inconclusive_count = int(data['TestResult'].split('(')[3].split(')')[0])
        not_run_count = int(data['TestResult'].split('(')[4].split(')')[0])

        formatted_test_result = f'Pass: {pass_count}, Fail: {fail_count}, Inconclusive: {inconclusive_count}, Not Run: {not_run_count}'

        # Update the dictionary with the new formatted 'TestResult'
        data['TestResult'] = formatted_test_result

        test_scope = data.get('TestScope', [])

        # Remove the 'TD_' prefix from each item in TestScope
        test_scope_updated = [item.replace('TD_', '') for item in test_scope]   

        # Update GUI_DATA with the modified TestScope
        data['TestScope'] = test_scope_updated

        # Remove whitespace, underscores, and dots from TestScope
        data['TestScope'] = [test.replace(" ", "").replace("_", "").replace(".", "").replace("-", "") for test in data['TestScope']]

        report_sequence = data.get('ReportSequence', '')

        # Remove spaces and convert to an integer
        report_sequence = int(report_sequence.replace(' ', ''))

        # Update GUI_DATA with the modified ReportSequence
        data['ReportSequence'] = report_sequence

        data['TestScope'] = [test.replace("kHz", "KHz") for test in data['TestScope']]

        #logging.info(f"GUI_DATA :{data}")

        test_scope_count = len(data['TestScope'])

        #logging.info("_______________JSON DATA_______________")
        json1 = setting['CustomLocation']
        directory_path = os.path.dirname(json1)
        files = os.listdir(directory_path)

        # Iterate through the files and find the first '.json' file
        json_file_path = None
        for file in files:
            if file.lower().endswith('.json'):
                json_file_path = os.path.join(directory_path, file)
                break  # Stop when the first JSON file is found

        # Print the full path to the first JSON file (or None if none is found)
        #logging.info(json_file_path)
        with open(json_file_path, "r") as file:
            json_data = json.load(file)

        data1 = json_data['TestExecutionDetails']

        #Parse the original timestamp using dateutil.parser.parse
        original_timestamp = parser.parse(data1['CreationTime'])

        # Format the timestamp in the desired format
        formatted_timestamp = original_timestamp.strftime("%d/%m/%Y, %H:%M:%S")

        # Update the dictionary with the new formatted 'CreationTime'
        data1['CreationTime'] = formatted_timestamp
        
        # Remove whitespace, underscores, and dots from TestScope
        data1['TestScope'] = [test.replace(" ", "").replace("_", "").replace(".", "").replace("-", "") for test in data1['TestScope']]

        #logging.info(f"JsonData :{data1}")

        test_scope_count = len(data1['TestScope'])

        #logging.info("_______________Compare GUI & Json Data_______________")
        print("Data",data)
        print("Data1",data1)
        result, remarks = compare_dicts(data, data1)
        if result == "Pass":
            pass_or_fail = result
        else:
            pass_or_fail = result
            remark = remarks
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_teup'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_Execution is minimized")
        else:
            logging.error("Test_Execution is not minimized")
            pass_or_fail = "Fail"
            remark = 'Test_Execution is not minimized'

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    # Get the calling method's name using inspect
    test_description = "Please verify the Test_Execution table and ensure that all elements are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc7_Test_toolinfo(driver_setup):
    logging.info("*************** TC-7 ***************")
    logging.info("Please verify the Test Tool info table and ensure that all elements are present.")
    pass_or_fail = "Pass"
    remark = ""
    try:
        #logging.info("_______________GUI DATA________________")
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_ttidropdown'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_Tool_Info is expanded")
        else:
            logging.error("Test_Tool_Info is not expanded")
            pass_or_fail = "Fail"
            remark = "Test_Tool_Info is not expanded"

        A = []
        B = []
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_TB_1'])
        for j in div_elements1:
            k = j.text
            A.append(k if k is not None and k.strip() != '' else '-')
        A = [value for value in A if value != '-']
        
        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['2'])
        for l in div_elements2:  
            m = l.text
            B.append(m if m is not None and m.strip() != '' else '-')
        B = [True if item == '-' else item for item in B]

        result_dict = dict(zip(A, B))
        After_rm_tti_text = {key.replace(' ', ''): value for key, value in result_dict.items()}
        #logging.info(f"GUI header value: {After_rm_tti_text}")

        #logging.info("_______________JSON DATA_______________")
        json1 = setting['CustomLocation']
        directory_path = os.path.dirname(json1)
        files = os.listdir(directory_path)

        # Iterate through the files and find the first '.json' file
        json_file_path = None
        for file in files:
            if file.lower().endswith('.json'):
                json_file_path = os.path.join(directory_path, file)
                break  # Stop when the first JSON file is found

        # Print the full path to the first JSON file (or None if none is found)
        #logging.info(json_file_path)
        with open(json_file_path, "r") as file:
            data = json.load(file)

        json_tl_text =  data["TestToolInfo"]
        #logging.info(f"Json header value: {json_tl_text}")

        #logging.info("_______________Compare GUI & Json Data_______________")
        compare_dicts(After_rm_tti_text, json_tl_text)

        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_ttiup'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_Tool_Info is minimized")
        else:
            logging.error("Test_Tool_Info is not minimized")
            pass_or_fail = "Fail"
            remark = "Test_Tool_Info is not minimized"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    # Get the calling method's name using inspect
    test_description = "Please verify the Test Tool info table and ensure that all elements are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc8_Test_dutinfo(driver_setup):
    logging.info("*************** TC-8 ***************")
    logging.info("Please verify the Test dust info table and ensure that all elements are present")
    pass_or_fail = "Pass"
    remark = ""
    try:
        #logging.info("_______________GUI DATA________________")
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_didropdown'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_dutinfo is expanded")
        else:
            logging.error("Test_dutinfo is not expanded")
            pass_or_fail = "Fail"

        A = []
        B = []

        """
        return result_list_key: Fetch the Keys from the GUI page 
        """
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_TB_1'])
        for j in div_elements1:
            k = j.text
            A.append(k if k is not None and k.strip() != '' else '-')
        A = [value for value in A if value != '-']
        indices=[7,8]
 
        for i in sorted(indices, reverse=True):
            del A[i]
        # Store the final list A in result_list
        result_list_key = A 

        """
        return result_list_key: Fetch the values from the GUI page 
        """
        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['2'])
        for l in div_elements2:  
            m = l.text
            B.append(m if m is not None and m.strip() != '' else '-')
            result_list_values = B
        formatted_list = []

        # Find the index of the first occurrence of a line starting with 'Manufacturing Code'
        start_index = next((i for i, s in enumerate(result_list_values) if isinstance(s, str) and s.startswith('Manufacturing Code')), None)

        if start_index is not None:
            for item in result_list_values[start_index:]:
                if isinstance(item, str):
                    # Remove specific patterns and replace with ','
                    item = re.sub(r'\n-\n|\n\d+\n|\n, \n|\n-\n|\n0x[0-9A-Fa-f]+\n|\n\d+|\n', ',', item)

                    # Replace remaining '\n' with ' '
                    item = item.replace('-', ' ')
                    item = item.replace(' ', '')
                    # Add to the formatted list
                    formatted_list.append(item)

        # Add the initial items to the beginning of the formatted list
        initial_items = result_list_values[:7]
        formatted_list = initial_items + formatted_list

        # Split the last two items into separate lists
        formatted_list1 = formatted_list[-1].split(',')
        formatted_list2 = formatted_list[-2].split(',')

        # Remove empty strings and extra commas from the lists
        formatted_list1 = [item for item in formatted_list1 if item]
        formatted_list2 = [item for item in formatted_list2 if item]

        #logging.info("_______________JSON DATA_______________")
        json1 = setting['CustomLocation']
        directory_path = os.path.dirname(json1)
        files = os.listdir(directory_path)

        # Iterate through the files and find the first '.json' file
        json_file_path = None
        for file in files:
            if file.lower().endswith('.json'):
                json_file_path = os.path.join(directory_path, file)
                break  # Stop when the first JSON file is found

        # Print the full path to the first JSON file (or None if none is found)
        #logging.info(json_file_path)
        with open(json_file_path, "r") as file:
            data = json.load(file)

        json_tl_text =  data["DutInfo"]
        #logging.info(f"Json header value: {json_tl_text}")

        """
        Fetch the Keys from the json File 
        """
        json_data_keys_list = list(json_tl_text.keys())

        indices=[7,8]
 
        for i in sorted(indices, reverse=True):
            del json_data_keys_list[i]
                    
        #print(json_data_keys_list)
        formatted_list = [format_element(value) for value in json_data_keys_list]
        formatted_list3 = [value.replace('Du T', 'DUT').replace('I D', 'ID').title() for value in formatted_list]
        
        # Fetching PTx and PRx keys and converting to lists
        PTx_keys = list(data["DutInfo"]["PTx"].keys())
        PRx_keys = list(data["DutInfo"]["PRx"].keys())
        
        #logging.info("_______________Compare GUI & Json Data_______________")
        """
        Compare: Keys from the json File and Keys from the GUI
        """
        
        if all(x.lower() == y.lower() for x, y in zip(result_list_key, formatted_list3)):
            print("1st if")
            pass
            #logging.info(f"Dut info keys are similar {result_list_key}{formatted_list3}")
        else:
            print("1st else")
            differences1 = set(result_list_key) ^ set(formatted_list3)
            # Convert the set to a sorted list
            sorted_list = sorted(differences1)

            # Convert the sorted list back to a set
            sorted_set = set(sorted_list)
            logging.error(f"Dut info keys are not similar 1{sorted_set}")
            pass_or_fail = "Fail" 
            remark = f"Dut info keys are not similar {sorted_set}"        

        if all(x.lower() == y.lower() for x, y in zip(formatted_list2, PTx_keys)):
            print("2nd if")
            pass
            #logging.info(f"Dut info keys are similar {formatted_list2}{PTx_keys}")
        else:

            print("2 else")
            differences2 = set(formatted_list2) ^ set(PTx_keys)
            # Convert the set to a sorted list
            sorted_list = sorted(differences2)

            # Convert the sorted list back to a set
            sorted_set = set(sorted_list)
            logging.error(f"Dut info keys are not similar 2{sorted_set}")
            pass_or_fail = "Fail" 
            remark = f"Dut info keys are not similar {sorted_set}"   

        if all(x.lower() == y.lower() for x, y in zip(formatted_list1, PRx_keys)):
            print("3rd if")
            pass
            #logging.info(f"Dut info keys are similar {formatted_list1}{PRx_keys}")
        else:
            print("3 else")
            differences3 = set(formatted_list1) ^ set(PRx_keys)
            # Convert the set to a sorted list
            sorted_list = sorted(differences3)

            # Convert the sorted list back to a set
            sorted_set = set(sorted_list)
            logging.error(f"Dut info keys are not similar 3{sorted_set}")
            pass_or_fail = "Fail" 
            remark = f"Dut info keys are not similar {sorted_set}"   
        
        """
        Value from the json File
        """
        # Create a list to store values
        values_list = []
        # Fetching values
        values_list.append(data["DutInfo"]["DuTType"])
        values_list.append(data["DutInfo"]["BrandName"])
        values_list.append(data["DutInfo"]["ProductName"])
        values_list.append(data["DutInfo"]["QiID"])
        values_list.append(data["DutInfo"]["SerialNumber"])
        values_list.append(data["DutInfo"]["PowerProfile"])
        values_list.append(data["DutInfo"]["SpecificationSupported"])

        values_list = [value if value != '' else '-' for value in values_list]


        """
        Compare: Values from the json File and Values from the GUI
        """
        if all(x.lower() == y.lower() for x, y in zip(initial_items, values_list)):
            pass
            #logging.info(f"Dut info values are similar {initial_items}{values_list}")
        else:
            logging.error(f"Dut info values are similar {initial_items}{values_list}")
            pass_or_fail = "Fail"
            remark = f"Dut info values are similar {initial_items}{values_list}"

        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_diup'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Test_dutinfo is minimized")
        else:
            logging.error("Test_dutinfo is not minimized")
            pass_or_fail = "Fail"
            remark = "Test_dutinfo is not minimized"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Please verify the Test dust info table and ensure that all elements are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc9_Test_scope(driver_setup):
    logging.info("*************** TC-9 ***************")
    logging.info("Testing Scopes")
    pass_or_fail = "Pass"
    remark = " "
    try:
        #logging.info("_______________GUI DATA________________")
        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_tsdropdown'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Testing_Scopes is expanded")
        else:
            logging.error("Testing_Scopes is not expanded")
            pass_or_fail = "Fail"
            remark = "Testing_Scopes is not expanded"

        A = []
        B = []

        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['1'])
        for j in div_elements1:
            k = j.text
            A.append(k if k is not None and k.strip() != '' else '-')
        A = [value for value in A if value != '-']

        div_elements2 = driver_setup.find_elements(By.XPATH, ElementXpath['Testcriteria'])
        for l in div_elements2:  
            m = l.text
            B.append(m if m is not None and m.strip() != '' else '-')


        result_dict = dict(zip(A, B))
        #logging.info(result_dict)

        #logging.info("_______________JSON DATA_______________")
        json1 = setting['CustomLocation']
        directory_path = os.path.dirname(json1)
        files = os.listdir(directory_path)

        # Iterate through the files and find the first '.json' file
        json_file_path = None
        for file in files:
            if file.lower().endswith('.json'):
                json_file_path = os.path.join(directory_path, file)
                break  # Stop when the first JSON file is found

        # Print the full path to the first JSON file (or None if none is found)
        #logging.info(json_file_path)
        with open(json_file_path, "r") as file:
            data = json.load(file)

            test_ids = []
            test_results = []
            # Process each test in the 'TestingScope' data
            for test in data.get("TestingScope", []):
                test_id = test.get("TestId")
                test_result = test.get("TestResult")
                # Append test_id to the list
                test_ids.append(test_id)
                test_results.append(test_result)

        result_dict1 = dict(zip(test_ids, test_results))
        #logging.info(result_dict1)

        #logging.info("_______________Compare GUI & Json Data_______________")
        compare_dicts(result_dict1,result_dict1)

        div_elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['RA_tsup'])
        if div_elements1:
            for ih in div_elements1:
                ih.click()
                logging.info("Testing_Scopes is minimized")
        else:
            logging.error("Testing_Scopes is not minimized")
            remarks = "Testing_Scopes is not minimized"
            pass_or_fail = "Fail"

    except NoSuchElementException:
        logging.error("Element not found")
        remark = "Element not found"
        pass_or_fail = "Fail"

    test_description = "Please verify the Test Scope table and ensure that all elements are present."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc12_boarder_header(driver_setup):
    logging.info("*************** TC-12 ***************")
    logging.info("Header")
    pass_or_fail = "Pass"
    remark = ""
    try:
        # Define expected values for first and last
        expected_first = yaml_msg('expected_first')
        expected_last = yaml_msg('expected_last')

        # Find all elements using XPath
        elements = driver_setup.find_elements(By.XPATH, ElementXpath['xpandall_first'])
        first = elements[0].text
        if expected_first == first:
            logging.info(f"Expand All Tab is visible")
        else:
            logging.error(f"Expand All Tab is not visible: {expected_first},{first}")
            pass_or_fail = "Fail"
            remark = f"Expand All Tab is not visible: {expected_first},{first}"
      
        # Find all elements using XPath
        elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['finalreportlast'])
        middle0 = elements1[0].text
        middle1= elements1[1].text
        last= elements1[2].text

        if expected_last == last:
            logging.info(f"Final report is visible")
        else:
            logging.error(f"Final report is not visible: {expected_last},{last}")
            pass_or_fail = "Fail"
            remark = f"Final report is not visible: {expected_last},{last}"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Please verify the boarder headerand ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc10_expandall(driver_setup):
    logging.info("*************** TC-10 ***************")
    logging.info(" Expand All")
    pass_or_fail = "Pass"
    remark = " "
    try:
        expand = driver_setup.find_element(By.XPATH, ElementXpath['ra_expandall'])
        expand.click()
        chech_up = driver_setup.find_elements(By.XPATH, ElementXpath['chech_up'])
        if chech_up:
            #logging.info(f"Found {len(chech_up)} arrow buttons pointing up.")
            logging.info("All the Tabs are expanded")
        else:
            #logging.error("No arrow buttons pointing up found.")
            logging.error("Error: Tabs are not expanded ")
            pass_or_fail = "Fail"
            remark = "Error: Tabs are not expanded "
        
        time.sleep(2)
        expand.click()
        chech_upq = driver_setup.find_elements(By.XPATH, ElementXpath['chech_down'])
        if chech_upq:
            #logging.info(f"Found {len(chech_upq)} arrow buttons pointing down.")
            logging.info("All the Tabs are minimized")
        else:
            #logging.error("No arrow buttons pointing up found.")
            logging.error("Error: Tabs are not minimized")
            pass_or_fail = "Fail"
            remark = "Error: Tabs are not minimized"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Please verify that the expand all button is present and ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc13_edit_reportRemark(driver_setup):
    logging.info("*************** TC-13 ***************")
    logging.info("Edit report remarks")
    pass_or_fail = "Pass"
    remark = " "
    try:
        rr_expand = driver_setup.find_elements(By.XPATH, ElementXpath['RA_rrdropdown'])
        if rr_expand:
            for i in rr_expand:
                i.click()
                logging.info("Edit button is clicked and present")

        edit_click = driver_setup.find_element(By.XPATH, ElementXpath['edit_click'])
        edit_click.click()
        # XPath expression for input boxes
        xpath_expression = ElementXpath['edit_input']

        # Find all input elements using XPath
        input_elements = driver_setup.find_elements(By.XPATH, xpath_expression)

        # Generate random 10 characters
        random_text = ''.join(random.choices(string.ascii_letters + string.digits, k=10))

        # Pass the random text to each input box
        for input_element in input_elements:
            input_element.send_keys(random_text)
            logging.info("Text pass successfully")

        time.sleep(5)
    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"

    test_description = "Please ensure that the Edit button is functional and can be clicked on to pass the value."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc14_merge(driver_setup):
    logging.info("*************** TC-14 ***************")
    logging.info("Merge_report")
    pass_or_fail = "Pass"
    remark = ""
    try:
        mr_merge = driver_setup.find_elements(By.XPATH, ElementXpath['merge_button'])
        if mr_merge:
            for i in mr_merge:
                i.click()
                logging.info("Merge button is clicked and present")
                check_mergebox = driver_setup.find_elements(By.XPATH, ElementXpath['check_mergebox'])
                RA_removeall = driver_setup.find_elements(By.XPATH, ElementXpath['RA_removeall'])
                if check_mergebox and RA_removeall:
                    logging.info("Report Merge is done")
                else:
                    logging.error("Report Merge is not done")
                    pass_or_fail = "Fail"
                    remark = "Report Merge is not done"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark ="Element not found"

    test_description = "Check that Report Merge is done"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc15_summary(driver_setup):
    logging.info("*************** TC-15 ***************")
    logging.info("Summary")
    pass_or_fail = "Pass"
    remark = " "
    try:
        Ra_summary = driver_setup.find_elements(By.XPATH, ElementXpath['Ra_summary'])
        if Ra_summary:
            for i in Ra_summary:
                i.click()
                Ra_summary_chart = driver_setup.find_elements(By.XPATH, ElementXpath['Ra_summary_chart'])
                if Ra_summary_chart:
                    logging.info("Summary Chart is displayed and it's clickable")
                    Ra_chart_summary = driver_setup.find_elements(By.XPATH, ElementXpath['Ra_chart_summary'])
                    if Ra_chart_summary:
                        logging.info("Report Summary text is present")
                        chart_testresult = driver_setup.find_elements(By.XPATH, ElementXpath['chart_testresult'])
                        if chart_testresult:
                            logging.info("Chart Test result text is present")
                            chart_testresult_button = driver_setup.find_elements(By.XPATH, ElementXpath['chart_testresult_button'])
                            if chart_testresult_button:
                                for i in chart_testresult_button:
                                    i.click()
                                    time.sleep(2)
                                    i.click()
                                    logging.info("Chart Test result button  is present")
                                    chart_Menu = driver_setup.find_element(By.XPATH, ElementXpath['chart_Menu'])
                                    if chart_Menu:
                                        chart_Menu.click()
                                        logging.info("Chart Menu is displayed and it's clickable")
                                        # Find all elements with the class "apexcharts-menu-item"
                                        menu_item_elements = driver_setup.find_elements(By.CLASS_NAME, 'apexcharts-menu-item')

                                        # Extract titles from the elements
                                        titles = [element.get_attribute('title') for element in menu_item_elements]

                                        print(titles)
                                        titles1 = yaml_msg('titles1')
                                        if titles == titles1:
                                            logging.info(f"All the titles are present as expected,{titles},{titles1}")
                                            chart_close = driver_setup.find_element(By.XPATH, ElementXpath['chart_close'])
                                            if chart_close:
                                                chart_close.click()
                                                logging.info("Summary page is close")
                                            else:
                                                logging.error("Summary page can't able to close and close button is not present")
                                                pass_or_fail = "Fail"
                                                remark = "Summary page can't able to close and close button is not present"
                                        else:
                                            logging.info(f"The titles are not present as expected,{titles},{titles1}")
                                            pass_or_fail = "Fail"
                                            remark = f"The titles are not present as expected,{titles},{titles1}"
                                    else:
                                        logging.error("chartMenu is not  present")
                                        pass_or_fail = "Fail"
                                        remark = "chartMenu is not  present"
                            else:
                                logging.info("chart_testresult_button text is not present")
                                pass_or_fail = "Fail"
                                remark = "chart_testresult_button text is not present"
                        else:
                            logging.error("chart_testresult text is not present")
                            pass_or_fail = "Fail"
                            remark = "chart_testresult text is not present"
                    else:
                        logging.error("Report Summary text is not present")
                        pass_or_fail = "Fail"
                        remark = "Report Summary text is not present"
                else:
                    logging.error("Summary Chart is not displayed")
                    pass_or_fail = "Fail"
                    remark = "Summary Chart is not displayed"
        else:
            logging.error("Summary tab is not displayed")
            pass_or_fail = "Fail"
            remark = "Summary tab is not displayed"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Check that summary is present and ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc17_delete(driver_setup):
    logging.info("*************** TC-17 ***************")
    logging.info("delete")
    pass_or_fail = "Pass"
    remark = " "
    try:
        elements1 = driver_setup.find_elements(By.XPATH, ElementXpath['finalreportlast'])
        middle0 = elements1[0].text
        middle1= elements1[1].text
        print(middle0,middle1)
        trash = driver_setup.find_element(By.XPATH, ElementXpath['trash'])
        if trash:
            trash.click()
            logging.info("Trash button is present and clicked")
            trashpopup = driver_setup.find_element(By.XPATH, ElementXpath['trashpopup'])
            if trashpopup:
                logging.info("Clear Reports pop-up is displayed")
                text = driver_setup.find_element(By.XPATH, ElementXpath['text'])
                if text:
                    #logging.info("Text is present")
                    clear_report = driver_setup.find_element(By.XPATH, ElementXpath['clear_report'])
                    if clear_report:
                        clear_report.click()
                        logging.info("Report is cleared")
                    else:
                        logging.error("Report is not cleared")
                        pass_or_fail = "Fail"
                        remark = "Report is not cleared"
                else:
                    logging.error("Text is not present")
                    pass_or_fail = "Fail"
                    remark = "Text is not present"
            else:
                logging.error("Clear Reports pop-up is not displayed")
                pass_or_fail = "Fail"
                remark = "Clear Reports pop-up is not displayed"
        else:
            logging.error("Trash button is not present")
            pass_or_fail = "Fail"
            remark = "Trash button is not present"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Check that delete is present and ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc16_export_file(driver_setup):
    logging.info("*************** TC-16 ***************")
    logging.info("export_file")
    remark = ""
    pass_or_fail = "Pass"
    try:
        export = driver_setup.find_element(By.XPATH, ElementXpath['export'])
        if export:
            export.click()
            #logging.info("Export file is present and clicked")
            export_all = driver_setup.find_element(By.XPATH, ElementXpath['export_all'])
            all = export_all.text
            export_HTML = driver_setup.find_element(By.XPATH, ElementXpath['export_HTML'])
            HTML = export_HTML.text
            export_PDF = driver_setup.find_element(By.XPATH, ElementXpath['export_PDF'])
            PDF = export_PDF.text
            export_json = driver_setup.find_element(By.XPATH, ElementXpath['export_json'])
            json = export_json.text
            lis = [all, HTML, PDF, json]
            title2 = yaml_msg('titles2')
            if lis == title2:
                logging.info(f"All the export items are present: {lis} {title2}")
                check_items = driver_setup.find_elements(By.XPATH, ElementXpath['export_checkbox'])
                for i in check_items:
                    i.click()
                    #logging.info("All the export items are checked")    
                export_cancel = driver_setup.find_element(By.XPATH, ElementXpath['export_cancel'])
                if export_cancel:
                    logging.info("Export Cancel is present")
                else:
                    logging.error("Export Cancel is not present")
                    remark = "Export Cancel is not present"
                export_download = driver_setup.find_element(By.XPATH, ElementXpath['export_download'])
                if export_download:
                    logging.info("Export Download is present")
                else:
                    logging.error("Export Download is not present")      
                    remark = "Export Download is not present"          
            else:
                logging.error(f"All the export items are not present: {lis} {title2}")
                remark = f"All the export items are not present: {lis} {title2}"
        else:
            logging.error("Export file is not present nor clicked")
            pass_or_fail = "Fail"
            remark = "Export file is not present nor clicked"

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    # Get the calling method's name using inspect
    test_description = "Check that export button is present and ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc18_digitalsignature(driver_setup):
    logging.info("*************** TC-18 ***************")
    logging.info("Digital Signature")
    remark = ""
    pass_or_fail = "Pass"

    try:
        element = driver_setup.find_element(By.XPATH, ElementXpath['DSI'])
        logging.info("Digital Signature Element is present.")
        element.click()
        try:
            #time.sleep(2)
            element1 = driver_setup.find_element(By.XPATH, ElementXpath['encrypt'])
            logging.info("Encrypted Hash 256 Bits is present")
            try:
                #time.sleep(2)
                element2 = driver_setup.find_element(By.XPATH, ElementXpath['TTPK'])
                logging.info("Test Tool Public Key is present")
            except NoSuchElementException:
                logging.error("Test Tool Public Key is not present")
                remark = "Test Tool Public Key is not present"
                pass_or_fail = "Fail"
        except NoSuchElementException:
            logging.error("Encrypted Hash 256 Bits is not present")
            remark = "Encrypted Hash 256 Bits is not present"
            pass_or_fail = "Fail"
    except NoSuchElementException:
        logging.error("Digital Signature Element is not present.")
        remark = "Digital Signature Element is not present."
        pass_or_fail = "Fail"

    # Get the calling method's name using inspect
    test_description = "Check that Digital Signature Info."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def compare_dicts(dict1, dict2):
    mismatch_found = False  # Flag to check if any mismatch is found

    for key in dict1:
        if key not in dict2:
            return log_and_return_error("JsonData", key)
        else:
            value1 = dict1[key]
            value2 = dict2[key]

            if value1 != value2:
                return log_and_return_mismatch(key, value1, value2)

    # Check for keys in JsonData that are not in GUI_DATA
    for key in dict2:
        if key not in dict1:
            return log_and_return_error("GUI_DATA", key)

    if not mismatch_found:
        logging.info("Comparison successful: GUI_DATA and JsonData match.")
        return "Pass", "Comparison successful: GUI_DATA and JsonData match."

def log_and_return_error(data_type, key):
    logging.error(f"Key '{key}' is missing in {data_type}.")
    return "Fail", f"Key '{key}' is missing in {data_type}."

def log_and_return_mismatch(key, value1, value2):
    differences = set(value1) ^ set(value2)
    logging.error(f"Mismatch for key '{key}': {differences}")
    return "Fail", f"Mismatch for key '{key}': {differences}"
    # mismatch_found = False  # Flag to check if any mismatch is found

    # for key in dict1:
    #     if key not in dict2:
    #         logging.error(f"Key '{key}' is missing in JsonData.")
    #         mismatch_found = True
    #         pass_or_fail = "Fail"
    #         remark = f"Key '{key}' is missing in JsonData."
    #         return pass_or_fail,remark
    #     else:
    #         value1 = dict1[key]
    #         value2 = dict2[key]

    #         if value1 != value2:
    #             differences1 = set(value1) ^ set(value2)
    #             logging.error(f"Mismatch for key{differences1}")
    #             mismatch_found = True
    #             pass_or_fail = "Fail"
    #             remark = f"Mismatch for key{differences1}"
    #             return pass_or_fail,remark

    # # Check for keys in JsonData that are not in GUI_DATA
    # for key in dict2:
    #     if key not in dict1:
    #         logging.error(f"Key '{key}' is missing in GUI_DATA.")
    #         mismatch_found = True
    #         pass_or_fail = "Fail"
    #         remark = f"Key '{key}' is missing in GUI_DATA."
    #         return pass_or_fail,remark

    # if not mismatch_found:
    #     logging.info("Comparison successful: GUI_DATA and JsonData match.")
    
# def compare_dicts(dict1, dict2):
#     mismatch_found = False  # Flag to check if any mismatch is found

#     for item in dict1:
#         if item not in dict2:
#             logging.error(f"Item '{item}' is missing in JsonData.")
#             mismatch_found = True

#     # Check for items in JsonData that are not in GUI_DATA
#     for item in dict2:
#         if item not in dict1:
#             logging.error(f"Item '{item}' is missing in GUI_DATA.")
#             mismatch_found = True

#     if not mismatch_found:
#         logging.info("Comparison successful: GUI_DATA and JsonData match.")

def format_element(value):
    result = ''
    for char in value:
        if char.isupper():
            result += ' ' + char
        else:
            result += char
    return result.strip()

def tc11_RA_uploadsecondjson(driver_setup):
    logging.info("*************** TC-11 ***************")
    logging.info("Upload second Json file")
    pass_or_fail = "Pass"
    remark = ""
    try:
        ra_uploadjson = driver_setup.find_element(By.XPATH, ElementXpath['RA_uploadbutton'])
        if ra_uploadjson.is_displayed():
            ra_uploadjson.click()
            logging.info("Json upload Button is visible and clicked")
            time.sleep(5)
        else:
            logging.error("Json upload Button is not visible.")
            pass_or_fail = "Fail"
            remark = "Json upload Button is not visible."

        # Extracting the TPT value from the CustomLocation using string manipulation
        custom_location_parts = setting['CustomLocation'].split('/Report/')

        if len(custom_location_parts) > 1:
            tpt_value = custom_location_parts[1].split('/')[0]
            logging.info(f"TPT/TPR Value is selected:{tpt_value}")
        else:
            logging.error("TPT/TPR Value not found in CustomLocation.")
            pass_or_fail = "Fail"
            remark = "TPT/TPR Value not found in CustomLocation."

        xpath = ElementXpath['RA_Jsonselect'].replace('$', tpt_value)
        print(xpath)
        RA_Jsonselect = driver_setup.find_element(By.XPATH, xpath)
        RA_Jsonselect.click()

        RA_upload = driver_setup.find_element(By.XPATH, ElementXpath['RA_upload'])
        if RA_upload.is_displayed():
            logging.info("Element 'upload button' is present and visible.")
            RA_upload.click()
            logging.info("Json File is uploaded successfully")
        else:
            logging.error("Element 'upload button' is not visible.")
            logging.error("Can't able to loaded the Json file")
            pass_or_fail = "Fail"
            remark = "Element 'upload button' is not visible\Can't able to loaded the Json file"
        time.sleep(5)

    except NoSuchElementException:
        logging.error("Element not found")
        pass_or_fail = "Fail"
        remark = "Element not found"

    test_description = "Please verify that the expand all button is present and ensure that the functionality is working."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)