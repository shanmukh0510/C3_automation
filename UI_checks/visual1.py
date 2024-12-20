import json
import time
from typing import KeysView
import yaml
from selenium.webdriver.common.by import By
import logging
import os
import cv2
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
import inspect
import image_diff
from PIL import Image, ImageChops
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

# Configure logging
log_file = 'Visual_reg_automation.log'

logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

Xpath = read_file('Xpath.json')
setting = read_file('setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None

timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# Function to create or open the Excel file
def create_or_open_excel_file(self):
        if not hasattr(self, 'excel_file') or self.excel_file is None:
            try:
                current_path = os.path.abspath(os.path.dirname(__file__))
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                self.excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_Visual_reg_automation_{timestamp}.xlsx")

                if not os.path.exists(self.excel_file):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(["Method Name", "Result"])
                    header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                    header_font = Font(bold=True)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    wb.save(self.excel_file)
                else:
                    wb = openpyxl.load_workbook(self.excel_file)
                self.ws = wb.active

                for row in self.ws.iter_rows(min_row=2, min_col=2, max_row=self.ws.max_row, max_col=self.ws.max_column):
                    for cell in row:
                        cell_value = str(cell.value)
                        if cell_value in ["Pass", "Fail"]:
                            fill_color = "00FF00" if cell_value == "Pass" else "FF0000"
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                        else:
                            logging.error(f"Cell with unexpected value: {cell_value}")
            except Exception as e:
                logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, result):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, result])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def OpenBrowser(driver_setup, mode):
    logging.info("--------------------1. OpenBrowser--------------------")
    try:
        pass_or_fail = "Pass"  # Initialize as "Pass" by default
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        logging.error(f"OpenBrowser failed with error: {str(e)}")

    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def connectionsetup_page(driver_setup):
    logging.info("*************** connectionsetup_page ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_ConnectionSetup.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_ConnectionSetup.png"
    connectionsetup_xpath = (ElementXpath['connectionsetup_xpath'])
    compare = ImageCompare(driver_setup,connectionsetup_xpath, reference_image_path, name='connectionsetup_page')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def Qi_exe(driver_setup):
    logging.info("*************** Qi_exe ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_Qi_exe_tab.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_Qi_exe_tab.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Qi_exe_tab']).click()
    connectionsetup_xpath = (ElementXpath['Qi_exe'])
    compare = ImageCompare(driver_setup,connectionsetup_xpath, reference_image_path, name='Qi_exe')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def testconfig_page(driver_setup):
    logging.info("*************** testconfig_page ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_testconfig_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_testconfig_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
    highlight_tooltip = driver_setup.find_element(By.XPATH, ElementXpath['Highlight_tooltip'])
    if highlight_tooltip:
        #logging.info("highlight_tooltip is visible.")
        highlight_tooltip.click()
    time.sleep(3)
    testconfig_page = (ElementXpath['testconfig_page'])
    compare = ImageCompare(driver_setup,testconfig_page, reference_image_path, name='testconfig_page')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def Results(driver_setup):
    logging.info("*************** Results ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_result_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_result_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Result_tab']).click()
    
    result_page = (ElementXpath['result_page'])
    print(result_page)
    compare = ImageCompare(driver_setup,result_page, reference_image_path, name='Results')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def Report(driver_setup):
    logging.info("*************** Report ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_report_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_report_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Report_tab']).click()
    
    report_page = (ElementXpath['report_page'])
    print(report_page)
    compare = ImageCompare(driver_setup,report_page, reference_image_path, name='Report')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def Report_Analyser(driver_setup):
    logging.info("*************** Report_Analyser ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_report_analyser_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_report_analyser_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Report_analyser']).click()
    time.sleep(5)
    report_page = (ElementXpath['report_analyser'])
    print(report_page)
    compare = ImageCompare(driver_setup,report_page, reference_image_path, name='Report_Analyser')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)


def Qi_Page(driver_setup):
    logging.info("*************** Qi_Page ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_Qi_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_Qi_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Qi']).click()
    report_page = (ElementXpath['Qi_page'])
    print(report_page)
    compare = ImageCompare(driver_setup,report_page, reference_image_path, name='Qi_Page')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

def Help_page(driver_setup):
    logging.info("*************** Help_page ***************")
    pass_or_fail = "Pass"
    if Element == 'TPT':
        reference_image_path = "img\\MPP\\TPT\\TPT_Help_page.png"
    else:
        reference_image_path = "img\\MPP\\TPR\\TPR_Help_page.png"
    driver_setup.find_element(By.XPATH, ElementXpath['Help']).click()
    report_page = (ElementXpath['Help_page'])
    print(report_page)
    compare = ImageCompare(driver_setup,report_page, reference_image_path, name='Help_page')

    if compare == True:
        logging.info("Images are similar.")
    else:
        logging.error("Images are different.")
        pass_or_fail = "Fail"
    # Get the calling method's name using inspect
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, pass_or_fail)

# def visualize_image_difference(reference_image_path, screenshot_image_path, diff_image_path):
#     reference_image = Image.open(reference_image_path)
#     screenshot_image = Image.open(screenshot_image_path)

#     # Ensure the images have the same size
#     reference_image = reference_image.resize(screenshot_image.size)

#     # Compute the image difference
#     diff_image = ImageChops.difference(reference_image, screenshot_image)

#     # Save the difference image
#     diff_image.save(diff_image_path)


def ImageCompare(driver_setup, screenshot_xpath,reference_image_path,name): 
    logging.info("*************** ImageCompare ***************")
    threshold = 100  # Define your threshold here
    screenshot_path = f"img\\screenshot\\{name}_{timestamp}.png"
    pass_or_fail = "Pass"
    try:
        # driver_setup.find_element(By.XPATH, ElementXpath['SetupDiagram_button'] ).click()
        time.sleep(5)
        image_element = driver_setup.find_element(By.XPATH, screenshot_xpath)

        image_element.screenshot(screenshot_path)

        # # Load the reference image from the provided file path
        # reference_image = Image.open(reference_image_path)

        # # Load the captured screenshot and convert it to the same size as the reference image
        # screenshot_image = Image.open(screenshot_path).resize(reference_image.size)

        # # Visualize the differences and save the image
        # visualize_image_difference(reference_image_path, screenshot_path, f"img\\difference\\{name}_{timestamp}_diff.png")

        # # Compare the two images
        # difference_value = ImageChops.difference(reference_image, screenshot_image).getbbox()

        # # Check if the images are similar based on threshold
        # if cv2.countNonZero(difference_value) < threshold:
        #     logging.info(f"Images are similar. Difference Value: {difference_value}")
        #     return True
        # else:
        #     logging.error(f"Images are different. Difference Value: {difference_value}")
        #     return False

        # Load the reference image from the provided file path
        reference_image = cv2.imread(reference_image_path, cv2.IMREAD_GRAYSCALE)

        # Load the captured screenshot and convert it to grayscale
        screenshot_image = cv2.imread(screenshot_path, cv2.IMREAD_GRAYSCALE)

        # Compare the two images
        difference = cv2.absdiff(reference_image, screenshot_image)

        # Check if the images are similar
        time.sleep(5)
        if cv2.countNonZero(difference) < threshold:
            logging.info(f"Images are similar.")
            return True
        else:
            logging.error(f"Images are different. Difference Value")
            # Optionally, save the difference image for further analysis
            cv2.imwrite(f"img\\difference\\{name}_{timestamp}.png", difference)
            return False

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        

