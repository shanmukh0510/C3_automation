import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException 

def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
print(Element)
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

# result_colors = {
#     "Pass": "00FF00",  # Green
#     "Fail": "FF0000",  # Red
# }

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_online_testconfiguration_automation_{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_connect_to_the_ethernet(driver_setup):
    logging.info("*************** TC-2 ***************")
    static_dyanmic = setting['static_dynamic']
    logging.info(static_dyanmic)
    # pass_or_fail = "Pass"  # Initialize as "Pass" by default
    # remark = ""
    # check that start button is disable state.
    time.sleep(5)
    if driver_setup.find_element(By.XPATH, "//img[@class='connection-status-icon']"):
        logging.info("Start button is disable state")
    if static_dyanmic == 'Static':
        pass_or_fail, remark = scanNetworkstate(driver_setup)
        pass_or_fail, remark = defaultIp(driver_setup)
    else:
        pass_or_fail, remark = dynamicIp(driver_setup)

    test_description = "Please ensure that the MPP setup should be properly connected to the Ethernet."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_runtest(driver):
    logging.info("*************** TC-3 ***************")
    logging.info("Verify that, the testcase is selected and the run is started.")
    
    pass_or_fail = "Pass"
    remark = ""
    time.sleep(2)
    
    try:
        driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
    except NoSuchElementException:
        logging.error("Error: Test Configuration Tab not found.")
        pass_or_fail = "Fail"
        remark = "Error: Test Configuration Tab not found."
        return

    list1 = ['TC_V_2.0.1', 'TC_EPP']
    all_list = [list1]

    for i in all_list:
        time.sleep(2)
        
        try:
            driver.find_element(By.XPATH, ElementXpath['ProjectCreationButton']).click()
            time.sleep(1)
            
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_Certification_drop_down'])
            dropdown_button.click()
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[0]])
            dropdown_button.click()
            
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile_drop_down'])
            dropdown_button.click()
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[1]])
            dropdown_button.click()
            
            create_project_button = driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit'])
            create_project_button.click()
            time.sleep(3)
            driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllTestCases']).click()
            time.sleep(1.5)
            
            logging.info(f"Certification {i[0]} is selected")
            logging.info(f"Power Profile {i[1]} is selected")
            
            if Element == 'TPT':
                if i[0] == 'TC_V_2.0.1' and i[1] == 'TC_EPP':
                    test_case_groups = {
                        
                        "Ping phase tests": "Ping phase tests",
                    }    
                TP_Testcases = read_file('json/TPT_V_2.0.1_EPP.json')
                
            for group_name, group_key in test_case_groups.items():
                test_cases = TP_Testcases.get(group_key, [])
                
                for test_case_name in test_cases:
                    try:
                        test_selected = []
                        xpath = ElementXpath['TestSelection'].replace('$', test_case_name)
                        if xpath != 'None':
                            test_element = driver.find_element(By.XPATH, xpath)
                            test_element.click()
                            time.sleep(0.2)
                            test_selected.append(test_case_name)
                        logging.info(f"Selected test cases: {test_selected} to run in online mode")
                    except NoSuchElementException:
                        logging.error(f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'.")
                        pass_or_fail = "Fail"
                        remark = f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'."
                        break

            time.sleep(1.5)
            driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllMinus']).click()
            
            driver.find_element(By.XPATH, ElementXpath['TC_TestStart']).click()
            logging.info("Start button is Enable state")
            logging.info("Test Execution Started..!")
            time.sleep(10)

            for _ in range(2):
                handlePopup(driver)
                time.sleep(7)

            while True:
                try:
                    WebDriverWait(driver, 5).until_not(
                        EC.presence_of_element_located((By.XPATH, '//*[@class="CircularProgressbar-text"]'))
                    )
                    break  # Break the loop once the CircularProgressbar-text disappears
                except TimeoutException:
                    pass_or_fail = "Fail"
                    time.sleep(5)
            time.sleep(5)

            # Check the Test Result
            tc3_5_no_testcase(driver)
            time.sleep(5)

        except NoSuchElementException as e:
            logging.error(f"Error: {str(e)}")
            pass_or_fail = "Fail"
            remark = f"Error: {str(e)}"
            break

    test_description = "Verify that, the testcase is selected and the run is started.."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc4_stop_the_testrun(driver):
    logging.info("*************** TC-4 ***************")
    logging.info("Skip and Stop the test run while Executing")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        driver.find_element(By.XPATH, ElementXpath['TestConfigurationTab']).click()
    except NoSuchElementException:
        logging.error("Error: Test Configuration Tab not found.")
        pass_or_fail = "Fail"
        remark = "Error: Test Configuration Tab not found."
        return

    list4 = ['TC_V_2.0.1', 'TC_EPP']
    all_list = [list4]

    for i in all_list:
        time.sleep(2)
        
        try:
            driver.find_element(By.XPATH, ElementXpath['ProjectCreationButton']).click()
            time.sleep(1)
            
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_Certification_drop_down'])
            dropdown_button.click()
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[0]])
            dropdown_button.click()
            
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath['TC_powerprofile_drop_down'])
            dropdown_button.click()
            time.sleep(1)
            dropdown_button = driver.find_element(By.XPATH, ElementXpath[i[1]])
            dropdown_button.click()
            
            create_project_button = driver.find_element(By.XPATH, ElementXpath['ProjectCreationSubmit'])
            create_project_button.click()
            time.sleep(3)
            driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllTestCases']).click()
            time.sleep(1.5)
            
            logging.info(f"Certification {i[0]} is selected")
            logging.info(f"Power Profile {i[1]} is selected")
            
            if Element == 'TPT':
                if i[0] == 'TC_V_2.0.1' and i[1] == 'TC_EPP':
                    test_case_groups = {
                        "Ping phase tests": "Ping phase tests",
                    }    
                TP_Testcases = read_file('json/TPT_V_2.0.1_EPP.json')
                
            for group_name, group_key in test_case_groups.items():
                test_cases = TP_Testcases.get(group_key, [])
                
                for test_case_name in test_cases:
                    try:
                        test_selected = []
                        xpath = ElementXpath['TestSelection'].replace('$', test_case_name)
                        if xpath != 'None':
                            test_element = driver.find_element(By.XPATH, xpath)
                            test_element.click()
                            time.sleep(0.2)
                            test_selected.append(test_case_name)
                        logging.info(f"Selected test cases: {test_selected} to run in online mode")
                    except NoSuchElementException:
                        logging.error(f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'.")
                        pass_or_fail = "Fail"
                        remark = f"Error: Test case '{test_case_name}' not found or cannot be selected in group '{group_name}'."
                        break

            time.sleep(1.5)
            driver.find_element(By.XPATH, ElementXpath['TC_ExpandAllMinus']).click()
            
            driver.find_element(By.XPATH, ElementXpath['TC_TestStart']).click()
            logging.info("Test Execution Started..!")
            time.sleep(10)

            for _ in range(2):
                handlePopup(driver)
                time.sleep(7)

            time.sleep(2)
            try:
                skip = driver.find_element(By.XPATH, ElementXpath['TC_Skip']).click()
                logging.info("Skip Button is clicked..!")
                logging.info("Test Execution Skipped..!")
            except NoSuchElementException:
                logging.error("Test Execution is not Skipped or Skipped button is not found")
                pass_or_fail = "Fail"
                remark = "Test Execution is not Skipped or Skipped button is not found"
            time.sleep(10)
            try:
                stop = driver.find_element(By.XPATH, ElementXpath['TC_Stop']).click()  
                logging.info("Stop Button is clicked..!")     
                logging.info("Test Execution Stopped..!") 
            except NoSuchElementException:
                logging.error("Test Execution is not Stopped or Stopped button is not found")
                pass_or_fail = "Fail"
                remark = "Test Execution is not Stopped or Stopped button is not found"
            time.sleep(10)

        except NoSuchElementException as e:
            logging.error(f"Error: {str(e)}")
            pass_or_fail = "Fail"
            remark = f"Error: {str(e)}"
            break

    test_description = "Skip and Stop the test run while Executing"
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_1_verfiy_Passcount(driver):
    time.sleep(20)
    logging.info("*************** TC-3.1: Verify Pass Count ***************")
    logging.info("Pass Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = " "
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_Pass'])))
    # Check if the element is visible
    if img_element.is_displayed():
                
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'PASS' element is visible and it's clicked")
        else:
            logging.error("The 'PASS' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'PASS' element is not clickable."
    else:
        logging.error("The 'PASS' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'PASS' element is not visible."

    xpath_img = ElementXpath['Pass_alt']

    # Find matching 'img' elements
    img_elements = driver.find_elements(By.XPATH, xpath_img)

    # Initialize a set to store the unique test case names
    unique_test_case_names = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element in img_elements:
        alt_value = img_element.get_attribute("alt")
        span_title_element = img_element.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name = span_title_element.text.strip()
        
        # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name and test_case_name not in unique_test_case_names:
            unique_test_case_names.add(test_case_name)

    # Print the count of unique test case names
    unique_test_case_name_count = len(unique_test_case_names)
    logging.info(f"Number of Pass test case names: {unique_test_case_name_count}")

    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Pass element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count

def tc3_2_verfiy_Failcount(driver):
    logging.info("*************** TC-3.2: Verify Fail Count ***************")
    logging.info("Fail Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = " "
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_Fail'])))
    # Check if the element is visible
    if img_element.is_displayed():        
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'FAIL' element is visible and it's clicked")
        else:
            logging.error("The 'FAIL' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'FAIL' element is not clickable."
    else:
        logging.error("The 'FAIL' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'FAIL' element is not visible."
    time.sleep(2)
    xpath_img1 = ElementXpath['Fail_alt']

    # Find matching 'img' elements
    img_elements1 = driver.find_elements(By.XPATH, xpath_img1)

    # Initialize a set to store the unique test case names
    unique_test_case_names1 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element1 in img_elements1:
        #time.sleep(1)
        alt_value1 = img_element1.get_attribute("alt")
        span_title_element1 = img_element1.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name1 = span_title_element1.text.strip()

        # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name1 and test_case_name1 not in unique_test_case_names1:
            unique_test_case_names1.add(test_case_name1)

    # Print the count of unique test case names
    unique_test_case_name_count1 = len(unique_test_case_names1)
    logging.info(f"Number of FAIL test case names: {unique_test_case_name_count1}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Fail element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count1

def tc3_3_verfiy_Inconclusivecount(driver):
    logging.info("*************** TC-3.3: Verify Inconclusive Count ***************")
    logging.info("Inconclusive count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_inconclusive'])))
    # Check if the element is visible
    if img_element.is_displayed():       
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'Inconclusive' element is visible and it's clicked")
        else:
            logging.error("The 'INCONCLUSIVE' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'INCONCLUSIVE' element is not clickable."
    else:
        logging.error("The 'INCONCLUSIVE' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'INCONCLUSIVE' element is not visible."
    time.sleep(2)
    xpath_img2 = ElementXpath['INCONCLUSIV_alt']

    # Find matching 'img' elements
    img_elements2 = driver.find_elements(By.XPATH, xpath_img2)

    # Initialize a set to store the unique test case names
    unique_test_case_names2 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element2 in img_elements2:
        #time.sleep(1)
        alt_value2 = img_element2.get_attribute("alt")
        span_title_element2 = img_element2.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name2 = span_title_element2.text.strip()

    # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name2 and test_case_name2 not in unique_test_case_names2:
            unique_test_case_names2.add(test_case_name2)

    # Print the count of unique test case names
    unique_test_case_name_count2 = len(unique_test_case_names2)
    logging.info(f"Number of INCONCLUSIVE test case names: {unique_test_case_name_count2}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Inconclusive element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count2

def tc3_4_verfiy_Notruncount(driver):          
    logging.info("*************** TC-3.4: Verify NotRun Count ***************")
    logging.info("NotRun Count")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    img_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, ElementXpath['Tg_notrun'])))
    # Check if the element is visible
    if img_element.is_displayed():
        # Check if the element is clickable
        if img_element.is_enabled():
            logging.info("The 'NotRun' element is visible and it's clicked")
        else:
            logging.error("The 'NOTRUN' element is not clickable.")
            pass_or_fail = "Fail"
            remark = "The 'NOTRUN' element is not clickable."
    else:
        logging.error("The 'NOTRUN' element is not visible.")
        pass_or_fail = "Fail"
        remark = "The 'NOTRUN' element is not visible."
    time.sleep(2)
    xpath_img3 = ElementXpath['Notrun_alt']

    # Find matching 'img' elements
    img_elements3 = driver.find_elements(By.XPATH, xpath_img3)

    # Initialize a set to store the unique test case names
    unique_test_case_names3 = set()

    # Iterate through the 'img' elements and fetch the data
    for img_element3 in img_elements3:
        #time.sleep(1)
        alt_value3 = img_element3.get_attribute("alt")
        span_title_element3 = img_element3.find_element(By.XPATH, ElementXpath['sibling'])
        test_case_name3 = span_title_element3.text.strip()

    # Check if the test case name is not in the set of unique names, and then add it
        if test_case_name3 and test_case_name3 not in unique_test_case_names3:
            unique_test_case_names3.add(test_case_name3)

    # Print the count of unique test case names
    unique_test_case_name_count3 = len(unique_test_case_names3)
    logging.info(f"Number of NOTRUN test case names: {unique_test_case_name_count3}")
    # Get the calling method's name using inspect
    test_description = "Please verify that the test progress has been checked and that the Notrun element and its corresponding count are displayed."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)
    return unique_test_case_name_count3

def tc3_5_no_testcase(driver):
    unique_test_case_name_count= tc3_1_verfiy_Passcount(driver)
    unique_test_case_name_count1 = tc3_2_verfiy_Failcount(driver)
    unique_test_case_name_count2= tc3_3_verfiy_Inconclusivecount(driver)
    unique_test_case_name_count3 = tc3_4_verfiy_Notruncount(driver)
    logging.info("*************** TC-3.5: Verify Total Count ***************")
    pass_or_fail = "Pass"  # Initialize as "Pass" by default
    remark = ""
    totalcase = unique_test_case_name_count + unique_test_case_name_count1 + unique_test_case_name_count2 + unique_test_case_name_count3
    logging.info("Total Number of Test Cases: %s", totalcase)
    
    # Initialize a list to store the counts
    counts = []
    # Define the XPath expression to select the 'strong' element
    xpath_strong = ElementXpath['strong']
    strong_element = driver.find_elements(By.XPATH, xpath_strong)
    # Iterate through the 'strong' elements and extract the counts
    for strong in strong_element:
        # Extract the text from the 'strong' element
        text = strong.text.strip()
        # Split the text to get the count
        parts = text.split('/')
        count = parts[1].strip() if len(parts) > 1 else None
        # Append the count to the list if it's not None
        if count is not None:
            counts.append(count)
    # Print the list of counts
    TP = ['Pass','Fail', 'Inconclusive','Notrun']
    # Compare each index with the total
    for count, TPs in zip(counts, TP):
            if int(count) == totalcase:
                pass
                #logging.info(f"Count {count} matches the total {TPs} count ({totalcase}).")
            else:
                logging.error(f"Count {count} does not match the total {TPs} count ({totalcase}).")
                pass_or_fail = "Fail"
                remark = f"Count {count} does not match the total {TPs} count ({totalcase})."

    counts1 = []
    xpath_strong = ElementXpath['strong']
    strong_element = driver.find_elements(By.XPATH, xpath_strong)
    # Iterate through the 'strong' elements and extract the counts
    for strong in strong_element:
        # Extract the text from the 'strong' element
        text = strong.text.strip()
        # Extract just the "28" from the text
        result = text.split('/')[0].strip()
        # Print the result
        counts1.append(result)

    my_list = counts1[:-1]  # Remove the last element
    #logging.info(my_list)

    my_list1 = [unique_test_case_name_count, unique_test_case_name_count1, unique_test_case_name_count2, unique_test_case_name_count3]
    #logging.info(my_list1)

    labels = ['PASS', 'FAIL', 'INCONCLUSIVE', 'NOTRUN']

    # Iterate through the lists and generate the desired output
    for label, result, string_result in zip(labels, my_list1, my_list):
        if result == int(string_result):
            logging.info(f"The Count of {string_result} matches the total number of {label}, which is also ({string_result}).")
        else:
            logging.error(f"The Count of {string_result} doesn't matches the total number of {label}, which is({string_result}).")
            pass_or_fail = "Fail"
            remark = f"The Count of {string_result} doesn't matches the total number of {label}, which is({string_result})."

    # Get the calling method's name using inspect
    test_description = "Please verify whether the total number of executed test cases matches the expected value."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def process_image(driver, image_src, result_type):
    try:
        # Locate the count element and extract the text
        count_xpath = f"//img[@src='{image_src}']/following-sibling::strong[@class='test-results-notify-count']"
        count_element = driver.find_element(By.XPATH, count_xpath)
        count_text = count_element.text.strip()

        # Extract the numerical count values
        current_count, total_count = map(int, count_text.split('/'))

        print(f"{result_type} - Current Count: {current_count}")
        print(f"{result_type} - Total Count: {total_count}")

        # # Locate and click the image
        # image_xpath = f"(//img[@src='{image_src}'])[2]"
        # image_element = driver.find_element(By.XPATH, image_xpath)
        # image_element.click()

        # Return current and total counts
        return current_count, total_count
    
    except NoSuchElementException as e:
        print(f"{result_type} - Element not found: {e}")
    except Exception as e:
        print(f"{result_type} - An error occurred: {e}")

    

def tc5_run_fail_inconclusive_notRun_testcases(driver):
    # try:
    #     Pass_current, Pass_total  = process_image(driver, './images/PassUnselect.png', 'Pass')
    #     Fail_current, Fail_total = process_image(driver, './images/FailUnselect.png', 'Fail')
    #     Incon_current, Incon_total = process_image(driver, './images/InConclusiveUnselect.png', 'Inconclusive')
    #     Notrun_current, Notrun_total = process_image(driver, './images/NotRunUnselect.png', 'Not Run')

    #     if Pass_current == Pass_total:
    #         print("All test cases are passed. No failures, inconclusive, or not run test cases.")
    #     else:
    #         if Fail_current > 0:
    #             print(f"{Fail_current} test cases are failed.")
    #             # # Locate and click the image
    #             image_xpath = f"(//img[@src='./images/FailUnselect.png'])[2]"
    #             image_element = driver.find_element(By.XPATH, image_xpath)
    #             image_element.click()
    #         if Incon_current > 0:
    #             print(f"{Incon_current} test cases are inconclusive.")
    #             # Locate and click the image
    #             image_xpath = f"(//img[@src='./images/InConclusiveUnselect.png'])[2]"
    #             image_element = driver.find_element(By.XPATH, image_xpath)
    #             image_element.click()
    #         if Notrun_current > 0:
    #             print(f"{Notrun_current} test cases are not run.")
    #             # Locate and click the image
    #             image_xpath = f"(//img[@src='./images/NotRunUnselect.png'])[2]"
    #             image_element = driver.find_element(By.XPATH, image_xpath)
    #             image_element.click()

    # except Exception as e:
    #     print(f"An error occurred: {e}")
    try:
        results = {
            "Pass": "./images/PassUnselect.png",
            "Fail": "./images/FailUnselect.png",
            "Inconclusive": "./images/InConclusiveUnselect.png",
            "Not Run": "./images/NotRunUnselect.png"
        }

        counts ={}

        for result_type, image_src in results.items():
            counts[result_type] = process_image(driver, image_src, result_type)
        
        pass_current, pass_total = counts["Pass"]
        fail_current, _ = counts["Fail"]
        incon_current, _ = counts["Inconclusive"]
        notrun_current, _ = counts["Not Run"]

        if pass_current == pass_total:
            print("All test cases are passed. No failures, inconclusive, or not run test cases.")
        else:
            for result_type, (current_count, total_count) in counts.items():
                if result_type != "Pass" and current_count > 0:
                    print(f"{current_count} test cases are {result_type}.")
                    # Locate and click the image
                    image_xpath = f"(//img[@src='{results[result_type]}'])[2]"
                    image_element = driver.find_element(By.XPATH, image_xpath)
                    image_element.click()
            time.sleep(80)
    
    except Exception as e:
        print(f"An error occurred: {e}")


def scanNetworkstate(driver_setup):
    remark = ""
    input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])
    # Click on the input element to focus it
    input_element.click()
    time.sleep(3)
    # Select all text in the input field and delete it
    input_element.send_keys(Keys.CONTROL + "a")  # Select all text
    input_element.send_keys(Keys.DELETE)  # Delete the selected text
    # Enter the new IP 
    in_valid = "192.168.255.2"
    input_element.send_keys(in_valid)
    time.sleep(3)
    button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
    button_element.click()
    time.sleep(5)
    try:
        pass_or_fail = "Pass"
        # Find the "Scan Network" and "Connect" buttons
        scan_button = driver_setup.find_element(By.XPATH,  ElementXpath['scanbutton'])
        connect_button = driver_setup.find_element(By.XPATH, ElementXpath['connect_button'])

        # Click the "Scan Network" button
        scan_button.click()

        # Define a wait for the loading icon to appear
        wait = WebDriverWait(driver_setup, 40)  # Adjust the timeout as needed

        # Wait for the loading icon to appear
        try:
            loading_icon = wait.until(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            
            # Check if the "Connect" button is disabled
            if not connect_button.is_enabled():
                #logging.info("'Connect' button is disabled as expected")
                pass
            else:
                #logging.error("Error: 'Connect' button is not disabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not disabled as expected."
            
        except Exception as e:
            logging.error("Loading icon did not appear within the specified time or is not visible.")
            pass_or_fail = "Fail"
            remark= "Loading icon did not appear within the specified time or is not visible."
        # Wait for the loading icon to disappear
        try:
            wait.until_not(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            logging.info("Loading icon has disappeared")
            
            # Check if the "Connect" button is enabled
            if connect_button.is_enabled():
                #logging.info("'Connect' button is enabled as expected.")
                pass
            else:
                #logging.error("Error: 'Connect' button is not enabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not enabled as expected."

        except Exception as e:
            logging.error("Loading icon did not disappear within the specified time or is still visible.")
            pass_or_fail = "Fail"
            remark = "Loading icon did not disappear within the specified time or is still visible." 

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        logging.error(f" The Scan Network button not visible on the UI: {str(e)}")
        remark = " The Scan Network button not visible on the UI"

    return pass_or_fail,remark

def defaultIp(driver_setup):
    check = setting["tester_connected"]
    remark = ""
    if check == True:
        try:
            pass_or_fail = "Pass"
            ip = 'DefaultIP'
            elements = driver_setup.find_elements(By.CSS_SELECTOR, ElementXpath['Ipdropdown'])
            element_req = None
            for element in elements:
                ip_element = element.find_element(By.TAG_NAME,ElementXpath['option'])
                ip_address = ip_element.get_attribute('value')
                Defaultip = yaml_msg("Defaultip")
                if ip_address == Defaultip:
                    element_req = element
            logging.info(f"The Default IP is :{Defaultip}")
            if element_req is not None:  # Check if element_req is assigned
                time.sleep(2)
                element_req.click()   
            connected = False
            while not connected:
                try:
                    # Try to find the image element using the provided XPath
                    image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
                    
                    # If the image is found, set connected to True and break out of the loop
                    connected = True
                    logging.info("Tester is connected with Default IP")

                except Exception as e:
                    try:
                        # If the image is not found, try to click the button
                        button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                        button_element.click()
                        #print("Clicked the connection setup button")
                        time.sleep(7)
                        # You can add additional logic here if needed after clicking the button
                        
                    except Exception as e:
                        # If both the image and button are not found, print a message and wait before trying again
                        logging.info("c3 is not connected, retrying...")
                        time.sleep(5)  # Adjust the sleep duration as needed

        except Exception as e:
            # If OpenBrowser raises an exception, it's a fail
            pass_or_fail = "Fail"
            logging.error(f"DefaultIp failed with error: {str(e)}")
            remark = f"DefaultIp failed with error: {str(e)}"
            #time.sleep(180)
        # Get the calling method's name using inspect
    else:
        logging.info("Tester is not connected so the test is get skipped")
        pass_or_fail = "Skip"
        remark = "Tester is not connected so the test is get skipped"

    return pass_or_fail, remark
       
def dynamicIp(driver_setup):
    remark=""
    input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])
    # Click on the input element to focus it
    input_element.click()
    time.sleep(3)
    # Select all text in the input field and delete it
    input_element.send_keys(Keys.CONTROL + "a")  # Select all text
    input_element.send_keys(Keys.DELETE)  # Delete the selected text
    # Enter the new IP 
    in_valid = "192.168.255.2"
    input_element.send_keys(in_valid)
    time.sleep(3)
    button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
    button_element.click()
    check = setting["tester_connected"]
    if check == True:
        try:
            pass_or_fail = "Pass"
            ip = 'DynamicIP'
            # Find the input element using its class name
            input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])

            # Click on the input element to focus it
            input_element.click()
            time.sleep(3)
            # Select all text in the input field and delete it
            input_element.send_keys(Keys.CONTROL + "a")  # Select all text
            input_element.send_keys(Keys.DELETE)  # Delete the selected text
            # Enter the new IP 
            in_valid = setting["DynamicIP"]
            input_element.send_keys(in_valid)
            logging.info(in_valid)
            time.sleep(5)
            connected = False
            while not connected:
                try:
                    # Try to find the image element using the provided XPath
                    image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
                    
                    # If the image is found, set connected to True and break out of the loop
                    connected = True
                    logging.info("Tester is connected Dynamically")

                except Exception as e:
                    try:
                        # If the image is not found, try to click the button
                        button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                        button_element.click()
                        #print("Clicked the connection setup button")
                        time.sleep(7)
                        # You can add additional logic here if needed after clicking the button
                        
                    except Exception as e:
                        # If both the image and button are not found, print a message and wait before trying again
                        logging.info("c3 is not connected, retrying...")
                        time.sleep(5)  # Adjust the sleep duration as needed

        except Exception as e:
            # If OpenBrowser raises an exception, it's a fail
            pass_or_fail = "Fail"
            logging.error(f"Can't able to find the IP text box or can't able to pass the IP address: {str(e)}")
            remark = f"Can't able to find the IP text box or can't able to pass the IP address: {str(e)}"

    else:
        logging.info("Tester is not connected so the test is get skipped")
        pass_or_fail = "Skip"
        remark = "Tester is not connected so the test is get skipped"

    return pass_or_fail, remark

def openbrowser(driver_setup,path):
    driver_setup.get(path)
    driver_setup.maximize_window()
    time.sleep(3)

def is_element_present(driver_setup):
    try:
        elements = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
        return True
    except Exception as e:
        return False
    
def handlePopup(driver_setup):    
    pop_up = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']").click()