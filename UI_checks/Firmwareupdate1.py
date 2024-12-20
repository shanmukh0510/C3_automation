import json
import time
import yaml
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import logging
import os
import openpyxl
import inspect
import datetime
from openpyxl.styles import Font, PatternFill
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import WebDriverException


def read_file(path):
    with open(path, "r") as rf:
        values = json.load(rf)
    return values

Xpath = read_file('json/Xpath.json')
setting = read_file('json/setting.json')
file_path = "C:\\GRL\\GRL-C3-MP-TPT\\AppData\\AppProperty.json"

ElementXpath = Xpath['Xpath'][setting['Mode']]
Element = setting['Mode']
APIs = Xpath['API'][setting['Mode']]

excel_file = None
ws = None

result_colors = {
    "Pass": "00FF00",  # Green
    "Fail": "FF0000",  # Red
}

# Function to create or open the Excel file
def create_or_open_excel_file():
    global excel_file
    if excel_file is None:
        try:
            current_path = os.path.abspath(os.path.dirname(__file__))
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = os.path.join(current_path, 'Logs', 'Excels', f"{Element}_Firmware_UpdateAutomation{timestamp}.xlsx")

            if not os.path.exists(excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Method Name", "Testdescription", "Result", "Remark(if the test failed or skip)"])
                header_fill = PatternFill(start_color='DDDDDDDD', end_color='DDDDDDDD', fill_type='solid')
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(excel_file)
            else:
                wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell_value = str(cell.value)
                    if cell_value == "Pass":
                        fill_color = "00FF00"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    elif cell_value == "Fail":
                        fill_color = "FF0000"
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        logging.error(f"Cell with unexpected value: {cell_value}")
        except Exception as e:
            logging.error(f"Failed to create or open Excel file: {str(e)}")

def write_to_excel(method_name, testdescription, result, remark):
    try:
        create_or_open_excel_file()  # Ensure Excel file is created or opened

        # Now, you can write data to the Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([method_name, testdescription, result, remark])
        wb.save(excel_file)
    except Exception as e:
        logging.error(f"Failed to write to Excel: {str(e)}")

def get_exe_yaml():
    exe_yaml_dict()
    messages_dict = exe_yaml_dict()
    return messages_dict

def exe_yaml_dict():
    current_path = os.path.abspath(os.path.dirname(__file__))
    file_path = current_path + "\\Resource\\resources.yaml"
    print(file_path)
    yaml_dict = {}
    if os.path.exists(file_path):
        with open(file_path, "r") as resource_yaml_fh:
            yaml_dict = yaml.safe_load(resource_yaml_fh)
    else:
        err_msg = "Specified file {} does not exist ".format(file_path)
        raise Exception(err_msg)
    return yaml_dict

def yaml_msg(value):
    config = get_exe_yaml()
    message = config[value]
    return message

def initialize_browser(driver_setup, mode):
    try:
        driver_setup.get(Xpath['URL'][mode])
        driver_setup.maximize_window()
        logging.info("The browser has been successfully opened, and the page has landed on the connection setup page.")
    except WebDriverException as e:
        logging.error(f"The attempt to open the browser failed: {str(e)}")
        raise

def tc1_open_browser(driver_setup, mode):
    logging.info("*************** TC-1 ***************")
    logging.info("OpenBrowser")
    
    pass_or_fail = "Pass"
    remark = ""
    
    try:
        initialize_browser(driver_setup, mode)
        # Add additional steps if needed after browser initialization
    except Exception as e:
        pass_or_fail = "Fail"
        logging.error(f"{str(e)}")
        remark = f"The attempt to open the browser failed: {str(e)}"

    test_description = "The browser should open and land on the connection setup page."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc2_connect_to_the_ethernet(driver_setup):
    logging.info("*************** TC-2 ***************")
    static_dyanmic = setting['static_dynamic']
    logging.info(static_dyanmic)
    # pass_or_fail = "Pass"  # Initialize as "Pass" by default
    # remark = ""
    # check that start button is disable state.
    time.sleep(5)
    if driver_setup.find_element(By.XPATH, "//img[@class='connection-status-icon']"):
        logging.info("Start button is disable state")
    if static_dyanmic == 'Static':
        pass_or_fail, remark = scanNetworkstate(driver_setup)
        pass_or_fail, remark = defaultIp(driver_setup)
    else:
        pass_or_fail, remark = dynamicIp(driver_setup)

    test_description = "Please ensure that the MPP setup should be properly connected to the Ethernet."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

def tc3_firmware_update(driver_setup):
    logging.info("*************** TC-3 ***************")
    logging.info("firmware_update")
    
    pass_or_fail = "Pass"
    remark = ""
    version_number = versioncheck()
    time.sleep(5)
    if Element == "TPT":
        version = '1.0'
    else:
        version = '3.3'

    if version_number == version:
        try:
            # Firmware update button
            clickupdate = driver_setup.find_element(By.XPATH, "//button[@id='connectionsetup_update_firmware_button']")
            clickupdate.click()
            time.sleep(8)
            try:
                # Update C3 Fimrware / Eload Diagram
                firmwareupdate_popup = driver_setup.find_element(By.XPATH, "//img[@class='popup-img' and @src='./images/Setup Images/MPPTPRFWUpdate.png']")
                if firmwareupdate_popup:
                    logging.info("Update Firmware pop-up is displayed")
                    try:
                        #firmware update done then click ok
                        ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-primary']")
                        if ok_button:
                            ok_button.click()
                            logging.info("In the Update Firmware pop-up Clicked the Ok button.")
                            time.sleep(20)

                            # if the USB cable Unable to access the controller
                            try:
                                if Element == "TPT":
                                    element = driver_setup.find_element(By.XPATH, "//strong[text()='Unable to access the controller. Please reconnect the USB cable and power cycle the GRL-C3-MP-TPT before retrying.']")
                                    text = 'Unable to access the controller. Please reconnect the USB cable and power cycle the GRL-C3-MP-TPT before retrying.'
                                else:
                                    element = driver_setup.find_element(By.XPATH, "//strong[text()='Unable to access the controller. Please reconnect the USB cable and power cycle the GRL-C3-MP-TPR before retrying.']")
                                    text = 'Unable to access the controller. Please reconnect the USB cable and power cycle the GRL-C3-MP-TPR before retrying.'

                                logging.info("Unable to access the controller. Please reconnect the USB cable and power cycle the GRL-C3-MP-TPT before retrying.")
                                logging.warning("USB is not connected")
                                # Get the text content of the element
                                text_content = element.text
                                if text_content == text:
                                    time.sleep(5)
                                    try:
                                        # Find the "Ok" button using its class name
                                        ok_button = driver_setup.find_element(By.CLASS_NAME, "popupButton_Ok")
                                        # Click the "Ok" button
                                        ok_button.click()
                                        logging.info("Unable to access the controller- Ok Pop is clicked")
                                        time.sleep(10)
                                        try:
                                            ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Cancel btn btn-secondary']")
                                            if ok_button:
                                                ok_button.click()
                                                if Element == "TPT":
                                                    logging.info("Latest PPS Firmware version 1.0 is already available in controller so clicked the Ok to ignore the PPS update, or click Cancel to update PPS again.")
                                                    logging.info("Cancel button is clicked")
                                                else:
                                                    logging.info("Latest PPS Firmware version 3.3 is already available in controller so clicked the Ok to ignore the PPS update, or click Cancel to update PPS again.")
                                                    logging.info("Cancel button is clicked")
                                                time.sleep(310)
                                                try:
                                                    ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                                    if ok_button:
                                                        ok_button.click()
                                                        logging.info("ELOAD updated successfully, Click Ok to power cycle the controller.")
                                                        time.sleep(15)
                                                        try:
                                                            ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                                            if ok_button:
                                                                ok_button.click()
                                                                logging.info("Please wait for the controller to finish rebooting. Clicked the Ok button in the popup.")
                                                                logging.info(" FirmwareUpdate is completed.")
                                                        except Exception as e:
                                                            logging.error("Failed to click the Ok button in the controller to finish rebooting popup.")
                                                            pass_or_fail = "Fail"
                                                            remark = "Failed to click the Ok button in the controller to finish rebooting popup."
                                                except Exception as e:
                                                    logging.error("Failed to click the Ok button in the PPS update done pop up.")
                                                    pass_or_fail = "Fail"
                                                    remark = "Failed to click the Ok button in the PPS update done pop up."
                                        except Exception as e:
                                            logging.error("Failed to click the Cancel button in the Latest PPS Firmware popup.")
                                            pass_or_fail = "Fail"
                                            remark = "Failed to click the Ok button in the Latest PPS Firmware popup."
                                    except NoSuchElementException:
                                        # Handle the case where the "Ok" button is not found
                                        logging.error("Unable to access the controller- Ok Pop-up button not found")
                            except NoSuchElementException:
                                logging.info("USB is connected")
                                # Handle the case where the element is not found
                                try:
                                    ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                    if ok_button:
                                        ok_button.click()
                                        logging.info("Clicked the Ok button in the Firmware Updated Successfully popup.")
                                        time.sleep(15)
                                        try:
                                            ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-primary']")
                                            if ok_button:
                                                ok_button.click()
                                                if Element == "TPT":
                                                    logging.info("Latest PPS Firmware version 1.0 is already available in controller so clicked the Ok to ignore the PPS update, or click Cancel to update PPS again.")
                                                    logging.info("Ok button is clicked")
                                                else:
                                                    logging.info("Latest PPS Firmware version 3.3 is already available in controller so clicked the Ok to ignore the PPS update, or click Cancel to update PPS again.")
                                                    logging.info("Ok button is clicked")
                                                time.sleep(15)
                                                try:
                                                    ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                                    if ok_button:
                                                        ok_button.click()
                                                        logging.info("Please wait for the controller to finish rebooting.Clicked the Ok button in the popup.")
                                                except Exception as e:
                                                    logging.error("Failed to click the Ok button in the controller to finish rebooting popup.")
                                                    pass_or_fail = "Fail"
                                                    remark = "Failed to click the Ok button in the controller to finish rebooting popup."
                                        except Exception as e:
                                            logging.error("Failed to click the Ok button in the Latest PPS Firmware popup.")
                                            pass_or_fail = "Fail"
                                            remark = "Failed to click the Ok button in the Latest PPS Firmware popup."
                                except Exception as e:
                                    logging.error("Failed to click the Ok button in the Firmware Updated Successfully popup.")
                                    pass_or_fail = "Fail"
                                    remark = "Failed to click the Ok button in the Firmware Updated Successfully popup."
                    except Exception as e:
                        logging.error("In the Update Firmware pop-up Failed to click the Ok button.")
                        pass_or_fail = "Fail"
                        remark = "In the Update Firmware pop-up Failed to click the Ok button."
            except Exception as e:
                logging.error("Update Firmware pop-up is not displayed.")
                pass_or_fail = "Fail"
                remark= "Update Firmware pop-up is not displayed."
        except Exception as e:
            logging.error("Update Firmware button is not clickable at the moment.")
            pass_or_fail = "Fail"
            remark= "Update Firmware button is not clickable at the moment."
    else:
        logging.warning("Eload is may be upgraded version or else Downgrade version Hence it will auto download for current version")
        try:
            clickupdate = driver_setup.find_element(By.XPATH, "//button[@id='connectionsetup_update_firmware_button']")
            clickupdate.click()
            time.sleep(8)
            try:
                firmwareupdate_popup = driver_setup.find_element(By.XPATH, "//img[@class='popup-img' and @src='./images/Setup Images/MPPTPRFWUpdate.png']")
                if firmwareupdate_popup:
                    logging.info("Update Firmware pop-up is displayed")
                    try:
                        ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-primary']")
                        if ok_button:
                            ok_button.click()
                            logging.info("In the Update Firmware pop-up Clicked the Ok button.")
                            time.sleep(20)
                            try:
                                ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                if ok_button:
                                    logging.info("Firmware Updated Successfully, Click Cancel to proceed PPS Firmware update.")
                                    ok_button.click()
                                    time.sleep(300)
                                    # while True:
                                    #     try:
                                    #         # Check if the loader is still visible
                                    #         loader = driver_setup.find_element(By.CLASS_NAME, "css-ihpzsd-style-Loader")
                                    #         if not loader.is_displayed():
                                    #             print("Loader has stopped.")
                                    #             break  # Exit the loop if the loader has disappeared
                                    #     except NoSuchElementException:
                                    #         print("Loader element not found.")
                                    #         break  # Exit the loop if the loader element is not found

                                    #     # Wait for 20 seconds before checking again
                                    #     time.sleep(20)
                                    try:
                                        ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                        if ok_button:
                                            ok_button.click()
                                            logging.info("PPS firmware updated successfully, Click Ok to power cycle the controller.")
                                            try:
                                                time.sleep(15)
                                                ok_button = driver_setup.find_element(By.XPATH, "//button[@class='popupButtons popupButton_Ok btn btn-success']")
                                                if ok_button:
                                                    ok_button.click()
                                                    logging.info("Clicked the Ok button in the popup Please wait for the controller to finish rebooting.")
                                                    logging.info("PPS firmware updated successfully.")
                                            except NoSuchElementException:
                                                logging.error("Ok button not found in the popup.")
                                    except NoSuchElementException:
                                        logging.error("Failed to click the Ok button in the PPS firmware updated.")
                            except Exception as e:
                                logging.error("Failed to click the Ok button in the Firmware Updated Successfully popup.")
                                pass_or_fail = "Fail"
                                remark = "Failed to click the Ok button in the Firmware Updated Successfully popup."
                    except Exception as e:
                        logging.error("In the Update Firmware pop-up Failed to click the Ok button.")
                        pass_or_fail = "Fail"
                        remark = "In the Update Firmware pop-up Failed to click the Ok button."
            except Exception as e:
                logging.error("Update Firmware pop-up is not displayed.")
                pass_or_fail = "Fail"
                remark= "Update Firmware pop-up is not displayed."
        except Exception as e:
            logging.error("Update Firmware button is not clickable at the moment.")
            pass_or_fail = "Fail"
            remark= "Update Firmware button is not clickable at the moment."

    # reconnect(driver_setup)
    # #version_number1 = versioncheck()

    # #if version_number == version_number1:
    
    # #else:
    #     #logging.error("PPS firmware updated is unsuccessfull.")

    test_description = "Please verify that the Firmware update validation is done."
    method_name = inspect.currentframe().f_code.co_name
    write_to_excel(method_name, test_description, pass_or_fail, remark)

# def reconnect(driver_setup):
#     time.sleep(180)
#     tc2_connect_to_the_ethernet(driver_setup)

def scanNetworkstate(driver_setup):
    remark = ""
    input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])
    # Click on the input element to focus it
    input_element.click()
    time.sleep(3)
    # Select all text in the input field and delete it
    input_element.send_keys(Keys.CONTROL + "a")  # Select all text
    input_element.send_keys(Keys.DELETE)  # Delete the selected text
    # Enter the new IP 
    in_valid = "192.168.255.2"
    input_element.send_keys(in_valid)
    time.sleep(3)
    button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
    button_element.click()
    time.sleep(5)
    try:
        pass_or_fail = "Pass"
        # Find the "Scan Network" and "Connect" buttons
        scan_button = driver_setup.find_element(By.XPATH,  ElementXpath['scanbutton'])
        connect_button = driver_setup.find_element(By.XPATH, ElementXpath['connect_button'])

        # Click the "Scan Network" button
        scan_button.click()

        # Define a wait for the loading icon to appear
        wait = WebDriverWait(driver_setup, 40)  # Adjust the timeout as needed

        # Wait for the loading icon to appear
        try:
            loading_icon = wait.until(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            
            # Check if the "Connect" button is disabled
            if not connect_button.is_enabled():
                #logging.info("'Connect' button is disabled as expected")
                pass
            else:
                #logging.error("Error: 'Connect' button is not disabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not disabled as expected."
            
        except Exception as e:
            logging.error("Loading icon did not appear within the specified time or is not visible.")
            pass_or_fail = "Fail"
            remark= "Loading icon did not appear within the specified time or is not visible."
        # Wait for the loading icon to disappear
        try:
            wait.until_not(EC.visibility_of_element_located((By.XPATH, ElementXpath['loading_icon'])))
            logging.info("Loading icon has disappeared")
            
            # Check if the "Connect" button is enabled
            if connect_button.is_enabled():
                #logging.info("'Connect' button is enabled as expected.")
                pass
            else:
                #logging.error("Error: 'Connect' button is not enabled as expected.")
                pass_or_fail = "Fail"
                remark = "Error: 'Connect' button is not enabled as expected."

        except Exception as e:
            logging.error("Loading icon did not disappear within the specified time or is still visible.")
            pass_or_fail = "Fail"
            remark = "Loading icon did not disappear within the specified time or is still visible." 

    except Exception as e:
        # If OpenBrowser raises an exception, it's a fail
        pass_or_fail = "Fail"
        logging.error(f" The Scan Network button not visible on the UI: {str(e)}")
        remark = " The Scan Network button not visible on the UI"

    return pass_or_fail,remark

def defaultIp(driver_setup):
    check = setting["tester_connected"]
    remark = ""
    if check == True:
        try:
            pass_or_fail = "Pass"
            ip = 'DefaultIP'
            elements = driver_setup.find_elements(By.CSS_SELECTOR, ElementXpath['Ipdropdown'])
            element_req = None
            for element in elements:
                ip_element = element.find_element(By.TAG_NAME,ElementXpath['option'])
                ip_address = ip_element.get_attribute('value')
                Defaultip = yaml_msg("Defaultip")
                if ip_address == Defaultip:
                    element_req = element
            logging.info(f"The Default IP is :{Defaultip}")
            if element_req is not None:  # Check if element_req is assigned
                time.sleep(2)
                element_req.click()   
            connected = False
            while not connected:
                try:
                    # Try to find the image element using the provided XPath
                    image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
                    
                    # If the image is found, set connected to True and break out of the loop
                    connected = True
                    logging.info("Tester is connected with Default IP")

                except Exception as e:
                    try:
                        # If the image is not found, try to click the button
                        button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                        button_element.click()
                        #print("Clicked the connection setup button")
                        time.sleep(7)
                        # You can add additional logic here if needed after clicking the button
                        
                    except Exception as e:
                        # If both the image and button are not found, print a message and wait before trying again
                        logging.info("c3 is not connected, retrying...")
                        time.sleep(5)  # Adjust the sleep duration as needed

        except Exception as e:
            # If OpenBrowser raises an exception, it's a fail
            pass_or_fail = "Fail"
            logging.error(f"DefaultIp failed with error: {str(e)}")
            remark = f"DefaultIp failed with error: {str(e)}"
            #time.sleep(180)
        # Get the calling method's name using inspect
    else:
        logging.info("Tester is not connected so the test is get skipped")
        pass_or_fail = "Skip"
        remark = "Tester is not connected so the test is get skipped"

    return pass_or_fail, remark
       
def dynamicIp(driver_setup):
    remark=""
    input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])
    # Click on the input element to focus it
    input_element.click()
    time.sleep(3)
    # Select all text in the input field and delete it
    input_element.send_keys(Keys.CONTROL + "a")  # Select all text
    input_element.send_keys(Keys.DELETE)  # Delete the selected text
    # Enter the new IP 
    in_valid = "192.168.255.2"
    input_element.send_keys(in_valid)
    time.sleep(3)
    button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
    button_element.click()
    check = setting["tester_connected"]
    if check == True:
        try:
            pass_or_fail = "Pass"
            ip = 'DynamicIP'
            # Find the input element using its class name
            input_element = driver_setup.find_element(By.CLASS_NAME, ElementXpath['Iptextbox'])

            # Click on the input element to focus it
            input_element.click()
            time.sleep(3)
            # Select all text in the input field and delete it
            input_element.send_keys(Keys.CONTROL + "a")  # Select all text
            input_element.send_keys(Keys.DELETE)  # Delete the selected text
            # Enter the new IP 
            in_valid = setting["DynamicIP"]
            input_element.send_keys(in_valid)
            logging.info(in_valid)
            time.sleep(5)
            connected = False
            while not connected:
                try:
                    # Try to find the image element using the provided XPath
                    image_element = driver_setup.find_element(By.XPATH, ElementXpath['connectonimg'])
                    
                    # If the image is found, set connected to True and break out of the loop
                    connected = True
                    logging.info("Tester is connected Dynamically")

                except Exception as e:
                    try:
                        # If the image is not found, try to click the button
                        button_element = driver_setup.find_element(By.XPATH, ElementXpath['Connectbutton'])
                        button_element.click()
                        #print("Clicked the connection setup button")
                        time.sleep(7)
                        # You can add additional logic here if needed after clicking the button
                        
                    except Exception as e:
                        # If both the image and button are not found, print a message and wait before trying again
                        logging.info("c3 is not connected, retrying...")
                        time.sleep(5)  # Adjust the sleep duration as needed

        except Exception as e:
            # If OpenBrowser raises an exception, it's a fail
            pass_or_fail = "Fail"
            logging.error(f"Can't able to find the IP text box or can't able to pass the IP address: {str(e)}")
            remark = f"Can't able to find the IP text box or can't able to pass the IP address: {str(e)}"

    else:
        logging.info("Tester is not connected so the test is get skipped")
        pass_or_fail = "Skip"
        remark = "Tester is not connected so the test is get skipped"

    return pass_or_fail, remark

def versioncheck():
    # Path to the JSON file
    if Element == "TPT":
        file_path = r"C:\GRL\GRL-C3-MP-TPT\AppData\AppProperty.json"
    else:
        file_path = r"C:\GRL\GRL-C3-MP-TPR\AppData\AppProperty.json"

    try:
        # Open the JSON file and load its contents
        with open(file_path, 'r') as file:
            data = json.load(file)
            
        # Access the "CS_FirmwareVersion" object and fetch the "PropertyValue"
        firmware_version = data["CS_FirmwareVersion"]["PropertyValue"]
        print(firmware_version)

        if Element == "TPT":
            # Split the firmware version string and extract the second part
            version_parts = firmware_version.split("/")
            if len(version_parts) > 1:
                version_number = version_parts[1].strip()  # Remove leading/trailing whitespace
                #logging.info(f"Firmware Version Number: {version_number}")  # Corrected logging statement
                return version_number
            else:
                logging.error("Firmware Version Number not found.")
                return None
        else:
            firmware_versions = firmware_version.split('/')
            version_parts = firmware_versions
            if len(version_parts) > 1:
                firmware_version1 = firmware_versions[1].strip()
                print(firmware_version1)
                return firmware_version1
            else:
                logging.error("Firmware Version Number not found.")
                return None
        
    except FileNotFoundError:
        logging.error("File not found.")
        return None
    except KeyError:
        logging.error("Key not found in JSON.")
        return None
    except Exception as e:
        logging.error(f"Error: {e}")
        return None